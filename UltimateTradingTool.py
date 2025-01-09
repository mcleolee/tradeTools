# -*- coding: utf-8 -*-
# TODO 每个函数完了之后clr
# TODO 自动识别IP
# TODO 交易成功之后有没有记录，这个也自动化
# TODO 获取终端的宽度（虽然已经有了）
# TODO 自动显示当时时间和交易阶段，提示应该做怎么操作
# TODO 写一个便捷查询持仓的函数
# TODO 写 def pressAnyKeyToContinue(): 然后替换所有
# TODO 写出 UI

# TODO 在主页打印出实时信号节点和时间！！！
# TODO 把昨日的 ZS 的 format_data 通过 18 复制到浙商的服务器

# TODo 判断今天是不是周一，比如用处在：checkYesterdayDataTo37()，要是周一运行程序就会炸

# TODO 增加功能：Python HTTPS server | use muti thread!
# TODO 增加改变窗口大小的功能
# 输入 SETPARAS 来查询需要修改参数的地方
from generallib import *

import ast
import glob
import math
import multiprocessing

import subprocess
from datetime import datetime, timedelta
import zipfile
# import requests                   # 这个不注释，在 61 上跑不了
# from bs4 import BeautifulSoup     # 这个不注释，在 40 上跑不了
import urllib.request
import csv
import tkinter as tk
from multiprocessing import Process
from tkinter import messagebox

import pandas
import sv_ttk
import re
import time
# import datetime
import threading
import xlrd
import openpyxl
import prettytable

import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from prettytable import PrettyTable
from concurrent.futures import ThreadPoolExecutor, as_completed

import matplotlib.pyplot as plt

# import seaborn as sns


import numpy as np
import cx_Oracle as co


from tqdm import tqdm

import baostock as bs

import sys
from contextlib import contextmanager
from WindPy import w

from pathlib import Path


DEBUG_MODE = 1

is_oracle_init = 0

# isFirstLoginWind = True

# 1表示启用，但这部分代码未完成
HTTP_SERVER = 0
server40addr = "http://192.168.1.40:8000/"
smart_divide_path = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\projects\smart_data_divide.py"
diff_excel_path = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\projects\mergeAllDiffIntoExcelFile.py"
grid_data_analysis_path = r'C:\Users\Administrator\Desktop\网格数据分析\grid_data_analysis_v1.1.py'

os.system("mode con cols=200 lines=70")

product_fund_dict = {
    "CF15": "480151137",
    "FL18": "480149909",
    "HT02XY": "480167623",
    "FL18SCA": "480160777",
    "FL22SCA": "480160777",
    "FL22SCB": "3050003937",
    "FL": "1260016888",
    "PAHF": "320300010625"
}

m = "morning"
m2 = "morning2Two"
a = "afternoon"
a2 = "afternoon2Two"

nom_ = "NoSellingMorning_"
nom2_ = "NoSellingMorning2Two_"
noa_ = "NoSellingAfternoon_"
noa2_ = "NoBuyingAfternoon2Two_"
underline = "_"

sell = "SellOrderList"
buy = "BuyOrderList"
jrcc = "JinRiChiCang"

divide_SC = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SCA\Sell_Buy_List_FL22SC"
divide_SCA = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SCA\Sell_Buy_List_FL22SCA"
divide_SCB = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SCB\Sell_Buy_List_FL22SCB"
ori_SC = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SC\Sell_Buy_List_FL22SC"
ori_SCA = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCA\Sell_Buy_List_FL22SCA"
ori_SCB = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCB\Sell_Buy_List_FL22SCB"

g_yesterday = None

g_realTimeNode = 0

# SETPARAS
# GRID PARAS
grid_info_path = r'C:\Users\Administrator\Desktop\grid_trade\grid_info'
grid_info_file_path = rf'{grid_info_path}/grid_stock_info.csv'
create_holding_info_path = rf'C:\Users\Administrator\Desktop\网格数据分析\build_clear_info\20240617交易股票建仓信息.csv'
clear_holding_info_path = rf'C:\Users\Administrator\Desktop\网格数据分析\build_clear_info\清仓信息.csv'
build_clear_info_path = rf'C:\Users\Administrator\Desktop\网格数据分析\build_clear_info'
trade_data_path = rf"C:\Users\Administrator\Desktop\grid_trade\trade_data"


# 获取文档文件夹路径
documents_path = os.path.join(os.environ['USERPROFILE'], 'Documents')

ins_order_path = rf"C:/Program Files/SmartTrader-Max/InsOrder"
scanTradePath = rf"C:\Users\Administrator\Desktop\兴业证券多账户交易"


# ====================================================================================================================
# ================================================ TKINTER ============================================================
# ====================================================================================================================

# def display_message_label():
#     message = "Hello, this is a message using Label."
#     label.config(text=message)
#
# def display_message_text():
#     message = "Hello, this is a message using Text."
#     text.delete(1.0, tk.END)  # 清空Text组件中的内容
#     text.insert(tk.END, message)
#
# def display_message_text():
#     message = "Hello, this is a message using Text."
#     text.delete(1.0, tk.END)  # 清空Text组件中的内容
#     text.insert(tk.END, message)


def iiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii():
    ...


# ====================================================================================================================
# ================================================ 工具函数 ============================================================
# ====================================================================================================================






def file_exists(file_path):
    """
    @brief Checks if the given file exists.
    @param file_path Path or string representing the file location.
    @return True if the file exists, False otherwise.
    """
    try:
        return Path(file_path).exists()
    except Exception as e:
        print(f"Error checking file existence: {e}")
        return False


def copy_file(original_path, target_path):
    try:
        ret = shutil.copy(original_path, target_path)
        print(f"文件已从 {original_path} 复制到 {target_path}")
        return ret
    except IOError as e:
        print(f"无法复制文件: {e}")


def copy_latest_files(source_folder, destination_folder, fileNumber):
    try:
        # 获取源文件夹中按时间排序的文件列表
        files = sorted(os.listdir(source_folder), key=lambda x: os.path.getmtime(os.path.join(source_folder, x)),
                       reverse=True)

        # 仅复制最新的四个文件
        for file in files[:fileNumber]:
            source_file = os.path.join(source_folder, file)
            destination_file = os.path.join(destination_folder, file)

            # 复制文件
            shutil.copy2(source_file, destination_folder)
            print(f"Copied: {source_file} to {destination_file}")
    except Exception as e:
        printRedMsg(f"Failed to copy files. Error: {e}")


def print_latest_files(source_folder, fileNumber):
    # 打印最新的 fileNumber 个文件名
    try:
        # 获取源文件夹中按时间排序的文件列表
        files = sorted(os.listdir(source_folder), key=lambda x: os.path.getmtime(os.path.join(source_folder, x)),
                       reverse=True)

        # 获取最新的fileNumber个文件名并打印
        for filename in files[:fileNumber]:
            print(filename)
    except Exception as e:
        printRedMsg(f"Failed to print latest files. Error: {e}")


# 复制按时间顺序排列的，有指定字符的前 fileNumber 个文件
def copy_files_with_string_limited(source_folder, destination_folder, string_to_check, fileNumber):
    try:
        # 获取源文件夹中按时间排序的文件列表
        files = sorted(os.listdir(source_folder), key=lambda x: os.path.getmtime(os.path.join(source_folder, x)),
                       reverse=True)

        # 计数器，用于记录已复制的文件数量
        copied_files = 0

        # 遍历文件列表
        for file in files:
            if copied_files >= fileNumber:
                break

            source_file = os.path.join(source_folder, file)

            # 检查文件名是否包含特定字符串
            if string_to_check in file:
                destination_file = os.path.join(destination_folder, file)

                # 复制文件
                shutil.copy2(source_file, destination_folder)
                print(f"Copied: {source_file} to {destination_file}")

                # 增加已复制文件数量
                copied_files += 1

    except Exception as e:
        printRedMsg(f"Failed to copy files. Error: {e}")


# 复制所有含指定字符的文件
def copy_files_with_string_no_limited(source_folder, destination_folder, string_to_check):
    try:
        # 获取源文件夹中按时间排序的文件列表
        files = sorted(os.listdir(source_folder), key=lambda x: os.path.getmtime(os.path.join(source_folder, x)),
                       reverse=True)

        # 遍历文件列表
        for file in files:
            source_file = os.path.join(source_folder, file)

            # 检查文件名是否包含特定字符串
            if string_to_check in file:
                destination_file = os.path.join(destination_folder, file)

                # 复制文件 覆盖目标文件夹中的同名文件
                shutil.copy2(source_file, destination_folder)
                print(f"Copied: {source_file} to {destination_file}")

    except Exception as e:
        printRedMsg(f"Failed to copy files. Error: {e}")


# 复制所有含指定字符的文件的测试函数
def test_copy_files_with_string_no_limited(source_folder, destination_folder, string_to_check):
    try:
        # 获取源文件夹中按时间排序的文件列表
        files = sorted(os.listdir(source_folder), key=lambda x: os.path.getmtime(os.path.join(source_folder, x)),
                       reverse=True)

        # 遍历文件列表
        for file in files:
            source_file = os.path.join(source_folder, file)

            # 检查文件名是否包含特定字符串
            if string_to_check in file:
                destination_file = os.path.join(destination_folder, file)

                # 复制文件
                shutil.copy2(source_file, destination_folder)
                print(f"Copied: {source_file} to {destination_file}")

    except Exception as e:
        print(f"Failed to copy files. Error: {e}")


def erase_folder_contents(folder_path):
    try:
        # 获取文件夹中的所有文件和文件夹
        items = os.listdir(folder_path)

        # 删除每个文件和清空文件夹
        for item in items:
            item_path = os.path.join(folder_path, item)
            if os.path.isfile(item_path):
                os.remove(item_path)
                print(f"Deleted file: {item_path}")
            elif os.path.isdir(item_path):
                # 清空文件夹中的内容
                erase_folder_contents(item_path)
                print(f"Emptied folder: {item_path}")
        return True
    except Exception as e:
        printRedMsg(f"Failed to erase folder contents. Error: {e}")
        return False


def move_files(source_folder, destination_folder):
    try:
        # 移动文件夹中的所有文件到目标文件夹
        shutil.move(source_folder, destination_folder)
        printGreenMsg(f"All files moved from {source_folder} to {destination_folder}")
    except Exception as e:
        printRedMsg(f"Failed to move files. Error: {e}")


def move_all_files_with_string(source_folder, destination_folder, string_to_check):
    try:
        # 确保目标文件夹存在
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)

        # 遍历源文件夹中的所有文件
        for filename in os.listdir(source_folder):
            source_file = os.path.join(source_folder, filename)

            # 检查是否是文件且文件名包含特定字符串
            if os.path.isfile(source_file) and string_to_check in filename:
                destination_file = os.path.join(destination_folder, filename)

                # 移动文件，如果目标文件已存在则覆盖
                shutil.move(source_file, destination_file)
                print(f"Moved: {source_file} to {destination_file}")

        printGreenMsg(f"All files containing '{string_to_check}' moved from {source_folder} to {destination_folder}")

    except Exception as e:
        printRedMsg(f"Failed to move files. Error: {e}")


def copy_files_in_folder(source_folder, destination_folder):
    try:
        # 遍历源文件夹中的所有文件
        for filename in os.listdir(source_folder):
            source_file_path = os.path.join(source_folder, filename)
            destination_file_path = os.path.join(destination_folder, filename)
            # 如果是文件则复制
            if os.path.isfile(source_file_path):
                shutil.copy2(source_file_path, destination_file_path)
                print(f"Copied file: {source_file_path} to {destination_file_path}")
    except Exception as e:
        printRedMsg(f"Failed to copy files. Error: {e}")


def zip_folder(folder_path, zip_path):
    try:
        # 创建一个zip文件
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # 遍历文件夹中的所有文件和子文件夹
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    # 将文件添加到zip文件中
                    zipf.write(file_path, os.path.relpath(file_path, folder_path))

        printGreenMsg(f"Folder {folder_path} successfully zipped to {zip_path}")
    except Exception as e:
        printRedMsg(f"Failed to zip folder. Error: {e}")

def zip_files_in_list(file_list, zip_path):
    """
    @brief  将文件列表压缩到指定路径
    @param  file_list: list - 文件路径列表
    @param  zip_path: str - 压缩文件保存路径
    """
    try:
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file in file_list:
                arcname = os.path.relpath(file, os.path.commonpath(file_list))
                zipf.write(file, arcname)
        printGreenMsg(f"Files successfully zipped to {zip_path}")
    except Exception as e:
        printRedMsg(f"Failed to zip file. Error: {e}")


def unzip_file(zip_path, extract_to):
    try:
        # 创建一个ZipFile对象
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # 解压缩所有文件到指定目标文件夹
            zip_ref.extractall(extract_to)

        printGreenMsg(f"Successfully extracted {zip_path} to {extract_to}")
    except Exception as e:
        printRedMsg(f"Failed to unzip file. Error: {e}")


def create_folder(folder_path):
    try:
        # 创建文件夹
        os.mkdir(folder_path)
        print(f"Folder created at {folder_path}")
    except Exception as e:
        printRedMsg(f"Failed to create folder. Error: {e}")


def ifExist(path):
    if os.path.exists(path):
        print(f"The file {path} exists.")
        return True
    else:
        notExist = "not exist"
        print(f"The file {path} does \033[33m{notExist}\033[0m.")
        return False







def ifContain(text, targetText):
    return targetText in text


def count_files_with_target_field(folder_path, target_field):
    try:
        # 初始化计数器
        count = 0
        # 遍历目标文件夹中的所有文件
        for filename in os.listdir(folder_path):
            # 如果目标字段在文件名中，则计数加一
            if target_field in filename:
                count += 1
        return count
    except Exception as e:
        printRedMsg(f"Failed to count files. Error: {e}")


def printName():
    logo1 = """
          _____                    _____                    _____                    _____                    _____          
         /\    \                  /\    \                  /\    \                  /\    \                  /\    \         
        /::\    \                /::\    \                /::\____\                /::\    \                /::\____\        
       /::::\    \              /::::\    \              /:::/    /               /::::\    \              /::::|   |        
      /::::::\    \            /::::::\    \            /:::/    /               /::::::\    \            /:::::|   |        
     /:::/\:::\    \          /:::/\:::\    \          /:::/    /               /:::/\:::\    \          /::::::|   |        
    /:::/__\:::\    \        /:::/__\:::\    \        /:::/____/               /:::/__\:::\    \        /:::/|::|   |        
   /::::\   \:::\    \      /::::\   \:::\    \       |::|    |               /::::\   \:::\    \      /:::/ |::|   |        
  /::::::\   \:::\    \    /::::::\   \:::\    \      |::|    |     _____    /::::::\   \:::\    \    /:::/  |::|   | _____  
 /:::/\:::\   \:::\____\  /:::/\:::\   \:::\    \     |::|    |    /\    \  /:::/\:::\   \:::\    \  /:::/   |::|   |/\    \ 
/:::/  \:::\   \:::|    |/:::/  \:::\   \:::\____\    |::|    |   /::\____\/:::/__\:::\   \:::\____\/:: /    |::|   /::\____\\
\::/   |::::\  /:::|____|\::/    \:::\  /:::/    /    |::|    |  /:::/    /\:::\   \:::\   \::/    /\::/    /|::|  /:::/    /
 \/____|:::::\/:::/    /  \/____/ \:::\/:::/    /     |::|    | /:::/    /  \:::\   \:::\   \/____/  \/____/ |::| /:::/    / 
       |:::::::::/    /            \::::::/    /      |::|____|/:::/    /    \:::\   \:::\    \              |::|/:::/    /  
       |::|\::::/    /              \::::/    /       |:::::::::::/    /      \:::\   \:::\____\             |::::::/    /   
       |::| \::/____/               /:::/    /        \::::::::::/____/        \:::\   \::/    /             |:::::/    /    
       |::|  ~|                    /:::/    /          ~~~~~~~~~~               \:::\   \/____/              |::::/    /     
       |::|   |                   /:::/    /                                     \:::\    \                  /:::/    /      
       \::|   |                  /:::/    /                                       \:::\____\                /:::/    /       
        \:|   |                  \::/    /                                         \::/    /                \::/    /        
         \|___|                   \/____/                                           \/____/                  \/____/         
                                                                                                                             
    """

    logo2 = """
    
    __________                               
    \______   \_____  ___  __  ____    ____  
     |       _/\__  \ \  \/ /_/ __ \  /    \ 
     |    |   \ / __ \_\   / \  ___/ |   |  \\
     |____|_  /(____  / \_/   \___  >|___|  /
            \/      \/            \/      \/ 
                                                        MADE WITH LOVE 
    """
    print(logo2)


if HTTP_SERVER:
    ...
    # def fetchDataFromHttpServer(server_url):
    #     try:
    #         response = requests.get(server_url)
    #         if response.status_code == 200:
    #             printGreenMsg("Data fetched successfully:")
    #             # print(response.text)
    #             return response.text
    #         else:
    #             printRedMsg(f"Failed to fetch data. Status code: {response.status_code}")
    #     except Exception as e:
    #         printRedMsg(f"An error occurred: {e}")
    #
    #
    # def download_files_from_html(html_content, server_url, download_dir):
    #     try:
    #         soup = BeautifulSoup(html_content, 'html.parser')
    #         links = soup.find_all('a')
    #
    #         for link in links:
    #             href = link.get('href')
    #             # 如果链接是相对路径，拼接成绝对路径
    #             if not href.startswith('http'):
    #                 href = server_url + '/' + href
    #
    #             filename = href.split('/')[-1]
    #             file_path = os.path.join(download_dir, filename)
    #
    #             printBlueMsg(f"Downloading {filename}...")
    #             r = requests.get(href, stream=True)
    #             with open(file_path, 'wb') as f:
    #                 for chunk in r.iter_content(chunk_size=8192):
    #                     if chunk:
    #                         f.write(chunk)
    #             printGreenMsg(f"{filename} downloaded successfully.")
    #
    #     except Exception as e:
    #         printRedMsg(f"An error occurred: {e}")


def get_public_ip():
    try:
        ip_data = urllib.request.urlopen('https://api.ipify.org').read().decode('utf-8')
        return ip_data
    except Exception as e:
        return "Error occurred: {}".format(str(e))


def find_stock_data(file_path, stock_code):
    with open(file_path, 'r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)  # 读取表头
        # print("表头:", end=" ")
        printGreenMsg("\n--------------------------------------------------")
        for item in header:
            print(item.ljust(15), end=" ")  # 对齐表头内容
        print()  # 换行

        isFound = False
        for row in reader:
            if any(stock_code in cell for cell in row):
                # print("找到匹配的行：", end=" ")
                for item in row:
                    print(item.ljust(15), end=" ")  # 对齐每一列的内容
                print()  # 换行
        printGreenMsg("\n--------------------------------------------------\n")
        input("press any key to return FIND DATA mode")
        isFound = True

        if not isFound:
            printRedMsg("未找到匹配的股票代码")
            input("press any key to continue...")


# 这个函数接受文件名和产品号与资金账号的字典作为参数。它首先在文件名中查找包含的关键字，然后根据找到的关键字获取对应的产品号。
# 最后，根据产品号在字典中查找对应的资金账号并输出。
def getInfoFromFileName(file_name, product_fund_dict_):
    keywords = ["CF15", "FL18", "HT02XY", "FL18SCA", "FL22SCB"]

    # 查找文件名中包含的关键字
    found_keyword = None
    for keyword in keywords:
        if keyword in file_name:
            found_keyword = keyword
            break

    if found_keyword is None:
        printRedMsg("未找到匹配的关键字")
        input("")
        return

    # 根据关键字获取对应的产品号
    product_number = None
    if found_keyword in product_fund_dict_:
        product_number = product_fund_dict_[found_keyword]

    if product_number is None:
        printRedMsg("未找到对应的产品号")
        input("")
        return

    print(f"\nYou are using \033[33m{found_keyword}\033[0m and the account number is \033[33m{product_number}\033[0m")

    # 根据产品号获取对应的资金账号
    # if product_number in product_fund_dict_:
    #     fund_account = product_fund_dict_[product_number]
    #     printGreenMsg(f"\nYou are using \033[33m{found_keyword}\033[0m and the account number is \033[33m{product_number}\033[0m\n")
    # else:
    #     printRedMsg("未找到对应的产品号")


def printAccountInfo():
    ...


def printAllAccountInfo():
    os.system("cls")
    isCorrect = enterPasswordToContinue()
    if not isCorrect:
        printRedMsg("Wrong password, contact with Admin")
        input("")
        return
    if isCorrect:
        printGreenMsg("password Correct.")
        input("")
    # 五个产品名
    product_names = ["CF15", "FL18", "HT02XY", "FL18SCA", "FL22SCB"]
    # 五个产品号
    product_numbers = ["480151137", "480149909", "480167623", "480160777", "3050003937"]
    # 五个密码
    passwords = ["297725", "230577", "890898", "123321", "123321"]
    printGreenMsg("\n--------------------------------------------------")

    print("产品名\t\t产品账号\t\t密码")
    for name, number, password in zip(product_names, product_numbers, passwords):
        print(f"{name}\t\t{number}\t\t{password}")
    printGreenMsg("\n--------------------------------------------------\n")

    input("press any key to quit")


def enterPasswordToContinue():
    cp = "147258"
    while True:
        p = input("Enter the password to unlock\n")
        if p == cp:
            return True
        else:
            return False


# 和下面的函数好像一样
# def list_files_with_numbers(path, n):
#     files = os.listdir(path)
#     files = [f for f in files if os.path.isfile(os.path.join(path, f))]  # 过滤出文件
#     files.sort()  # 排序文件列表
#
#     for i, file in enumerate(files[:n]):
#         print(f"{i + 1}: {file}")
#
#     choice = int(input("请选择文件序号（输入数字 1 到 n）: "))
#     if choice < 1 or choice > n:
#         print("无效的选择")
#         return None
#
#     selected_file = os.path.join(path, files[choice - 1])
#     return selected_file


# 如果文件可能很大，我们可以采用逐行读取和写入的方式来处理文件，以避免一次性加载整个文件到内存中
def remove_lines_with_character(file_path, target_character):
    try:
        print(f"正在处理 {target_character} 中...", end=" ")
        # 将目标字符转换为字节对象
        target_byte = target_character.encode()
        temp_file_path = file_path + '.tmp'  # 创建临时文件路径
        # 逐行读取原文件，并过滤掉包含特定字符的行，写入临时文件
        with open(file_path, 'rb') as input_file, open(temp_file_path, 'wb') as output_file:
            for line in input_file:
                if target_byte not in line:
                    output_file.write(line)
        # printGreenMsg("正在逐行读取原文件，并过滤掉包含特定字符的行，写入临时文件中...")
        # 替换原文件

        os.remove(file_path)
        # printGreenMsg("正在替换原文件中...")
        os.rename(temp_file_path, file_path)

        printGreenMsg("处理完成")

    except Exception as e:
        # 如果发生异常，打印错误消息
        printRedMsg(f"处理文件时出现错误：{e}")
        input("")


def rename_all_py_files_and_return_paths(folder_path, new_fund_name):
    file_paths = []

    # 遍历文件夹中的所有文件
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".py"):  # 确保只处理.py文件
                file_path = os.path.join(root, file)
                file_paths.append(file_path)

                # 读取文件内容
                with open(file_path, "r") as f:
                    content = f.read()

                # 使用正则表达式替换文件中的字符串
                new_content = re.sub(r'fund_name=\w+', 'fund_name=' + new_fund_name, content)

                # 将替换后的内容写回文件
                with open(file_path, "w") as f:
                    f.write(new_content)

                # 生成新的文件名（可选）
                new_file_name = file.replace("old_fund_name", new_fund_name)
                new_file_path = os.path.join(root, new_file_name)

                # 重命名文件
                os.rename(file_path, new_file_path)

    return file_paths


def rename_py_files(directory_path, new_extension=".new"):
    """
    :brief 传入一个路径，遍历这个路径下所有py文件，对每个文件进行重命名操作
    Function to rename all .py files in a given directory.
    :param directory_path: Path to the directory containing .py files.
    :param new_extension: New extension for renamed files (default: ".new").
    """
    today = getDate('')
    print(f'today is {today}')
    files = os.listdir(directory_path)
    for filename in files:
        if filename.endswith(".py"):
            old_path = os.path.join(directory_path, filename)
            productName = find_fund_name(old_path)
            if productName:
                new_filename = f"{productName}_{today}{new_extension}"
                new_path = os.path.join(directory_path, new_filename)
                os.rename(old_path, new_path)


def find_fund_name(filepath):
    """
    传入一个文件路径，找到文件里含有fund_name=的一行，返回等号后面的字符串
    Function to find and return the string after "fund_name=" in a specified file.
    :param filepath: Path to the file.
    :return: String after "fund_name=" if found, otherwise None.
    """
    with open(filepath, 'r', encoding='utf-8') as file:
        for line in file:
            match = re.search(r'fund_name=(.*)', line)
            if match:
                return match.group(1).strip()


# Example usage:
# 1. Rename all .py files in the directory "my_directory" with extension ".new"
# rename_py_files("my_directory", ".new")

# 2. Find and return the string after "fund_name=" in the file "example_file.txt"
# fund_name_value = find_fund_name("example_file.txt")
# print(fund_name_value)


# 总是删掉错误的行！！！
def remove_lines_time_in_range_for_order_log(file_path):
    printYellowMsg("in func remove_lines_time_in_range_for_order_log")
    try:
        # 读取文件内容
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        # 用于匹配时间格式的正则表达式
        import re
        time_pattern = re.compile(r'^\d{2}:\d{2}:\d{2}$')

        # 初始化开始索引和结束索引
        start_index = None
        end_index = None

        # 逐行遍历文件内容
        for i, line in enumerate(lines):
            # 判断是否为时间行且不包含其他字符
            if time_pattern.match(line.strip()) and len(line.strip()) == 8:
                if start_index is None:
                    start_index = i
                else:
                    # 发现一个新的时间行，删除上一个时间段的内容
                    if end_index is not None:
                        del lines[start_index + 1:end_index]
                    start_index = i
                end_index = None
            elif start_index is not None:
                # 找到非时间行，更新结束索引
                end_index = i

        # 如果文件末尾存在时间段，则删除最后一个时间段的内容
        if end_index is not None:
            del lines[start_index + 1:end_index]

        # 将处理后的内容写回文件
        with open(file_path, 'w', encoding='uft-8') as file:
            file.writelines(lines)
    except Exception as e:
        # 如果发生异常，打印错误消息
        printRedMsg(f"删除order文件的一系列时间行错误：{e}")
        input("")


def modifyNamePAtoPAHF(path):
    try:
        # 遍历指定目录中的所有 CSV 文件
        for filename in os.listdir(path):
            if filename.endswith(".csv") and "PA" in filename:
                # 构造旧文件路径
                old_file = os.path.join(path, filename)

                # 替换文件名中的 "PA" 为 "PAHF"
                new_filename = filename.replace("PA", "PAHF")

                # 构造新文件路径
                new_file = os.path.join(path, new_filename)

                # 重命名文件
                os.rename(old_file, new_file)
                print(f"Renamed: {old_file} to {new_file}")

        printGreenMsg("Files renamed successfully.")

    except Exception as e:
        printRedMsg(f"Failed to rename files. Error: {e}")


def run_python_file(path):
    os.system(fr"D:\software\python3.6.6\python.exe {path}")


# 如果表头中有重复的字段名称，PrettyTable 会抛出 Field names must be unique 错误。
# 为了解决这个问题，可以在读取表头时检查并处理重复字段名称。例如，可以在字段名称后面添加编号使其唯一。
# 以下是一个改进后的示例
def make_unique(field_names):
    """
    确保字段名称唯一。
    """
    seen = {}
    unique_field_names = []
    for name in field_names:
        if name in seen:
            seen[name] += 1
            unique_field_names.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            unique_field_names.append(name)
    return unique_field_names


def replace_none_with_empty_string(row):
    """
    将行中的 None 替换为空字符串。
    """
    return ['' if cell is None else cell for cell in row]


def read_and_print_xlsx(file_path):
    """
    读取并打印 .xlsx 文件的内容

    参数:
    file_path (str): .xlsx 文件的路径
    """
    printYellowMsg(f"now printing excel: {file_path}")
    try:
        # 尝试加载工作簿
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # 创建 PrettyTable 对象
        table = PrettyTable()

        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        unique_headers = make_unique(headers)
        table.field_names = replace_none_with_empty_string(unique_headers)

        # 添加表格行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            table.add_row(replace_none_with_empty_string(row))

        # 打印表格内容
        print(table)
    except FileNotFoundError:
        printRedMsg(f"File not found: {file_path}. Please check the file path.")
    except Exception as e:
        printRedMsg(f"An error occurred while reading the file: {e}")


def backup_file(file_path):
    """
    Copy a file to the same directory with a new name appended with .backup{today}.
    If a file with that name already exists, append _1, _2, etc. to avoid duplication.

    :param file_path: str, path to the original file
    """
    # Get the directory and the original file name
    dir_path, original_file_name = os.path.split(file_path)
    base_name, ext = os.path.splitext(original_file_name)

    # Get today's date
    today = datetime.now().strftime('%Y%m%d')

    # Construct the new file name
    new_file_name = f"{base_name}.backup{today}{ext}"
    new_file_path = os.path.join(dir_path, new_file_name)

    # Check if the new file name already exists and add a suffix if necessary
    counter = 1
    while os.path.exists(new_file_path):
        new_file_name = f"{base_name}.backup{today}_{counter}{ext}"
        new_file_path = os.path.join(dir_path, new_file_name)
        counter += 1

    # Copy the file
    shutil.copy2(file_path, new_file_path)
    printGreenMsg(f"Backup created: {new_file_path}")


def get_input(prompt, default=None):
    """
    Get user input with a prompt. If no input is provided, return the default value.

    :param prompt: str, the input prompt to display
    :param default: the default value to return if no input is given
    :return: the user input or the default value
    """
    user_input = input(prompt)
    if user_input.strip() == "":
        return default
    return user_input

def getFileName(filePath):
    return os.path.basename(filePath)


def get_days_between_date(start_date, end_date):
    """
    计算两个日期之间的天数间隔。

    参数:
    start_date (str): 起始日期，格式为 'YYYY-MM-DD'
    end_date (str): 结束日期，格式为 'YYYY-MM-DD'

    返回:
    int: 两个日期之间的天数间隔
    """
    # 先从YYYYMMDD转换到YYYY-MM-DD
    start_date = start_date[:4] + '-' + start_date[4:6] + '-' + start_date[6:]
    end_date = end_date[:4] + '-' + end_date[4:6] + '-' + end_date[6:]

    date_format = "%Y-%m-%d"

    # 将日期字符串解析为 datetime 对象
    start_date_obj = datetime.strptime(start_date, date_format)
    end_date_obj = datetime.strptime(end_date, date_format)

    # 计算日期差
    delta = end_date_obj - start_date_obj

    return delta.days


def get_testReturn(ret, start_date, end_date):
    '''
    计算 info 文件中的 test_return
    需要回测文件里的Ret、start_date和end_date
    :return:
    '''
    delta_days = get_days_between_date(start_date, end_date)
    print(f"From {start_date} to {end_date}, there are {delta_days} days.")
    return (ret/delta_days)*365

def get_min_max_price_from_backtesting_daily_data_ADJ(STOCK_CODE, start_date, end_date, file_or_fetch=1):
    '''
    :param start_date, end_date: 时间段
    :param file_or_fetch: 从文件读取(0)还是从在线数据库(1)读取，默认为在线数据库
    :return: 返回股价最小值和最大值
    FIXME 这是一个有问题的函数，需要修改，返回的值很不对
    '''
    global is_oracle_init
    if is_oracle_init == 0:
        # co.init_oracle_client(lib_dir=r"C:\software\oracle\instantclient_21_14") # SETPARAS IDK on which pc
        co.init_oracle_client(lib_dir=r"D:\softwares\oracle\instantclient_21_14") # SETPARAS ON MY PC
        is_oracle_init = is_oracle_init+1
    else:
        ...
    if file_or_fetch == 1:
        # 连接到Oracle数据库，需用户名、密码、ip、端口和数据库名
        conn = co.connect('jg_user08', 'ZDbR@#Zxsj@#sg2y', '109.244.130.220:20000/wind')

        # 创建游标 cursor
        cursor = conn.cursor()
        printGreenMsg("successful connecting to database!")
        def queryTable_FilterBy_S_INFO_WINDCODE(tableName, S_INFO_WINDCODE, cursor):

            if cursor is None:
                raise Exception("Cursor has not been initialized.")

            colName = "S_INFO_WINDCODE"
            tableName = (tableName).upper()
            # 查询表内容
            query = f"""
                SELECT *
                FROM (
                    SELECT *
                    FROM (
                        SELECT *
                        FROM {tableName}
                        WHERE S_INFO_WINDCODE = '{S_INFO_WINDCODE}'
                        ORDER BY TRADE_DT DESC
                    )
                    -- WHERE ROWNUM <= 10
                )
                ORDER BY TRADE_DT ASC
                        """
            cursor.execute(query)
            data = cursor.fetchall()

            # 获取列名
            colnames = [desc[0] for desc in cursor.description]

            # 转换为Pandas DataFrame
            df = pd.DataFrame(data, columns=colnames)

            # save_query_to_csv(df, f"{S_INFO_WINDCODE[:6]}_DAILY_RAW_DATA.csv", f"{raw_data_path}\DAILY")

            return df
            # print(df)


        printBlueMsg(f">>>   Starting to download and process daily data for {STOCK_CODE}...  ")
        RAW_DATA_DAILY = queryTable_FilterBy_S_INFO_WINDCODE("AShareEODPrices", STOCK_CODE, cursor)
        KEEP_COL_FOR_DAILY_BACKTEST = ['TRADE_DT', 'S_DQ_PRECLOSE', 'S_DQ_OPEN', 'S_DQ_HIGH', 'S_DQ_LOW', 'S_DQ_CLOSE',
                                       'S_DQ_VOLUME', 'S_DQ_ADJFACTOR', 'S_DQ_ADJCLOSE']
        RAW_DATA_DAILY_KEEP = RAW_DATA_DAILY[KEEP_COL_FOR_DAILY_BACKTEST]

        # 转换日期列为datetime格式
        RAW_DATA_DAILY_KEEP.loc[:, 'TRADE_DT'] = pd.to_datetime(RAW_DATA_DAILY_KEEP['TRADE_DT'], format='%Y%m%d')

        # 过滤指定时间段的数据
        filtered_data = RAW_DATA_DAILY_KEEP[(RAW_DATA_DAILY_KEEP['TRADE_DT'] >= start_date) & (RAW_DATA_DAILY_KEEP['TRADE_DT'] <= end_date)]

        # 计算最小值和最大值
        # min_price = filtered_data[['S_DQ_OPEN']].min().values[0]
        # max_price = filtered_data[['S_DQ_OPEN']].max().values[0]

        min_price = filtered_data[['S_DQ_ADJCLOSE']].min().values[0]
        max_price = filtered_data[['S_DQ_ADJCLOSE']].max().values[0]

        printBlueMsg(f">>>   Finish processing daily data for {STOCK_CODE}...  ")

        # 关闭游标和 ORACLE 数据库
        cursor.close()
        conn.close()
        # print(min_price, max_price)
        return min_price, max_price

    elif file_or_fetch == 0:
        ...
    else:
        printRedMsg("that's a wrong para.")
        input("")
        return


def get_min_max_price_from_backtesting_daily_data(STOCK_CODE, start_date, end_date, file_or_fetch=1):
    '''
    :param start_date, end_date: 时间段
    :param file_or_fetch: 从文件读取(0)还是从在线数据库(1)读取，默认为在线数据库
    :return: 返回股价最小值和最大值
    '''
    global is_oracle_init
    if is_oracle_init == 0:
        # co.init_oracle_client(lib_dir=r"C:\software\oracle\instantclient_21_14") # SETPARAS IDK on which pc
        co.init_oracle_client(lib_dir=r"D:\softwares\oracle\instantclient_21_14") # SETPARAS ON MY PC
        is_oracle_init = is_oracle_init+1
    else:
        ...
    if file_or_fetch == 1:
        # 连接到Oracle数据库，需用户名、密码、ip、端口和数据库名
        conn = co.connect('jg_user08', 'ZDbR@#Zxsj@#sg2y', '109.244.130.220:20000/wind')

        # 创建游标 cursor
        cursor = conn.cursor()
        printGreenMsg("successful connecting to database!")
        def queryTable_FilterBy_S_INFO_WINDCODE(tableName, S_INFO_WINDCODE, cursor):

            if cursor is None:
                raise Exception("Cursor has not been initialized.")

            colName = "S_INFO_WINDCODE"
            tableName = (tableName).upper()
            # 查询表内容
            query = f"""
                SELECT *
                FROM (
                    SELECT *
                    FROM (
                        SELECT *
                        FROM {tableName}
                        WHERE S_INFO_WINDCODE = '{S_INFO_WINDCODE}'
                        ORDER BY TRADE_DT DESC
                    )
                    -- WHERE ROWNUM <= 10
                )
                ORDER BY TRADE_DT ASC
                        """
            cursor.execute(query)
            data = cursor.fetchall()

            # 获取列名
            colnames = [desc[0] for desc in cursor.description]

            # 转换为Pandas DataFrame
            df = pd.DataFrame(data, columns=colnames)

            # save_query_to_csv(df, f"{S_INFO_WINDCODE[:6]}_DAILY_RAW_DATA.csv", f"{raw_data_path}\DAILY")

            return df
            # print(df)


        printBlueMsg(f">>>   Starting to download and process daily data for {STOCK_CODE}...  ")
        RAW_DATA_DAILY = queryTable_FilterBy_S_INFO_WINDCODE("AShareEODPrices", STOCK_CODE, cursor)
        KEEP_COL_FOR_DAILY_BACKTEST = ['TRADE_DT', 'S_DQ_PRECLOSE', 'S_DQ_OPEN', 'S_DQ_HIGH', 'S_DQ_LOW', 'S_DQ_CLOSE',
                                       'S_DQ_VOLUME', 'S_DQ_ADJFACTOR']
        RAW_DATA_DAILY_KEEP = RAW_DATA_DAILY[KEEP_COL_FOR_DAILY_BACKTEST]

        # 转换日期列为datetime格式
        RAW_DATA_DAILY_KEEP.loc[:, 'TRADE_DT'] = pd.to_datetime(RAW_DATA_DAILY_KEEP['TRADE_DT'], format='%Y%m%d')

        # 过滤指定时间段的数据
        filtered_data = RAW_DATA_DAILY_KEEP[(RAW_DATA_DAILY_KEEP['TRADE_DT'] >= start_date) & (RAW_DATA_DAILY_KEEP['TRADE_DT'] <= end_date)]

        # 计算最小值和最大值
        # min_price = filtered_data[['S_DQ_OPEN']].min().values[0]
        # max_price = filtered_data[['S_DQ_OPEN']].max().values[0]

        min_price = filtered_data[['S_DQ_CLOSE']].min().values[0]
        max_price = filtered_data[['S_DQ_CLOSE']].max().values[0]

        printBlueMsg(f">>>   Finish processing daily data for {STOCK_CODE}...  ")

        # 关闭游标和 ORACLE 数据库
        cursor.close()
        conn.close()
        # print(min_price, max_price)
        return min_price, max_price

    elif file_or_fetch == 0:
        ...
    else:
        printRedMsg("that's a wrong para.")
        input("")
        return


def append_new_element_to_line(line, new_element):
    '''
    按csv的格式在行后添加新元素
    :param line:
    :param new_element:
    :return: new_line
    '''
    new_line = line + ',' + new_element
    return new_line


def append_new_line_to_csv(file_path, new_line):
    """
    将新行添加到 CSV 文件。

    参数:
    file_path (str): CSV 文件的路径。
    new_line (list): 要添加的新行数据。
     'a'（追加模式）和 'r+'（读写模式）
    """
    try:
        # 检测文件编码
        with open(file_path, 'rb') as file:
            result = chardet.detect(file.read())
        encoding = result['encoding']

        # 打开 CSV 文件，使用 'a' 模式追加内容
        with open(file_path, 'a', newline='',encoding=encoding) as file:
            writer = csv.writer(file)
            writer.writerow(new_line)
        printGreenMsg("New line added successfully.")
    except Exception as e:
        printRedMsg(f"Failed to add new line. Error: {e}")


def add_row_to_df_manually(df):
    """
    This function takes a DataFrame and prompts the user to input values for a new row.
    It displays the current headers of the DataFrame and adds the new row to the DataFrame.
    """
    new_row = {}
    print("Current DataFrame headers are: ", list(df.columns))
    try:
        for column in df.columns:
            value = input(f"Enter the value for '{column}': ")
            new_row[column] = value

        # Convert the dictionary to a DataFrame row and append it
        df = df.append(new_row, ignore_index=True)
        printGreenMsg("New row added successfully.")
    except Exception as e:
        printRedMsg(f"Failed to add new row. Error: {e}")
        return -1
    return df

def get_header(df):
    """
    打印 DataFrame 的表头。
    参数:
    df (pd.DataFrame): 要打印表头的 DataFrame。
    """
    # printGreenMsg("DataFrame headers:")
    return(df.columns)

def open_new_console(command):
    if os.name == 'nt':  # 如果是Windows系统
        subprocess.Popen(['start', 'cmd', '/k', command], shell=True)
    else:
        # 其他系统的实现
        subprocess.Popen(['gnome-terminal', '--', 'bash', '-c', command])


def clear_line():
    print('\r', end='')

def is8Digits(date):
    return len(date) == 8

def list_files(directory):
    """
    @brief  获取指定目录下所有文件的路径和文件名
    @param  directory: str - 需要遍历的目录路径
    @return list - 包含所有文件路径和文件名的列表
    """
    files_list = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            files_list.append((file_path, file))
    return files_list



def filter_files_by_keyword_inList(file_list, keyword):
    """
    @brief  筛选包含特定字段的文件路径和文件名
    @param  file_list: list - 文件路径和文件名的列表
    @param  keyword: str - 需要匹配的字段
    @return list - 包含匹配字段的文件路径和文件名的列表
    """
    filtered_files = [file for file in file_list if keyword in file[1]]
    return filtered_files


def extract_paths_from_tuple_list(file_list):
    # 初始化一个空列表来存储路径
    paths = []

    # 遍历列表中的每个元组
    for file_tuple in file_list:
        # 将元组的第一个元素（即文件路径）添加到路径列表中
        paths.append(file_tuple[0])

    # 返回只包含路径的列表
    return paths


def count_weekdays(start_date, end_date):
    """
    @brief 计算两个日期之间的工作日数量
    @param start_date 开始日期
    @param end_date 结束日期
    @return 返回工作日数量
    """
    day_count = (end_date - start_date).days + 1  # 总天数
    weekend_count = 0

    for single_date in (start_date + timedelta(n) for n in range(day_count)):
        if single_date.weekday() >= 5:  # 周六和周日
            weekend_count += 1

    return day_count - weekend_count


def get_today_files(directory):
    # 获取当前日期的年、月、日
    today = datetime.today().date()

    today_files = []

    # 遍历目标目录下的所有文件和子文件夹
    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)

        # 检查是否为文件
        if os.path.isfile(filepath):
            # 获取文件创建时间（Unix时间戳）
            creation_time = os.path.getctime(filepath)
            file_date = datetime.fromtimestamp(creation_time).date()

            # 如果文件创建时间是今天，则添加到列表中
            if file_date == today:
                today_files.append(filepath)

    return today_files





# 上下文管理器
# 用法：        with redirect_stdout(log):
#
#             from WindPy import w
#             log("DOWNLOADING FROM WIND DATABASE...")
#
#             # 连接Wind数据库
#             w.start()
#             # 获取分钟行情数据
#             wind_data = w.wsd(row.ticker,
#                          "close",
#                          str(row.start_date),
#                          str(row.end_date),
#                          PriceAdj='F')
class StreamToLogger:
    def __init__(self, log_func):
        self.log_func = log_func

    def write(self, message):
        if message.strip():  # 忽略空行
            self.log_func(message)

    def flush(self):
        pass

@contextmanager
def redirect_stdout(log_func):
    logger = StreamToLogger(log_func)
    old_stdout = sys.stdout
    sys.stdout = logger
    try:
        yield
    finally:
        sys.stdout = old_stdout
        

def get_daily_data_from_wind(code, start_date, end_date, price_adjust_type="F"):
    """
        # 获取wind日k线数据
        :param code: 股票代码，比如601688.SH
        :param start_date: 开始日期，比如2015-09-10
        :param end_date: 结束日期，比如2022-12-31
        :param price_adjust_type: 复权类型，"F"前复权，"B"后复权，"NA"不复权
        :return:日数据的dataframe，列名有PRE_CLOSE,OPEN,HIGH,LOW,CLOSE,VOLUME,ADJFACTOR
        """

    if code.startswith("0") or code.startswith("3"):
        code = code + ".SZ"
    elif code.startswith("6"):
        code = code + ".SH"

    def disappear(*args, **kwargs):
        ...

    with redirect_stdout(disappear):
        w.start()

    result = w.wsd(f"{code}", "pre_close,open,high,low,close,volume,adjfactor",
                   start_date, end_date, f"PriceAdj={price_adjust_type}")

    datatime = [x.strftime('%Y-%m-%d') for x in result.Times]
    data_columns = [x.upper() for x in result.Fields]
    df = pd.DataFrame(result.Data, index=data_columns, columns=datatime)
    data = df.T
    # print(data)
    def round_fixed(num, d):
        """
        浮点数的四舍五入
        num: 浮点数
        d: d>0，在小数点后面第几位四舍五入；
           d<0，在整数部位四舍五入，d=-1表示将个位数四舍五入，以此类推
        """
        if num == 0:
            a = 0
        else:
            a = round(num + 0.5 / 10 ** 7, d)  # +0.5/10**7是为了避免丢失的精度造成误差
        return a

    data["PRE_CLOSE"] = data["PRE_CLOSE"].apply(lambda x: round_fixed(x, 2))
    data["OPEN"] = data["OPEN"].apply(lambda x: round_fixed(x, 2))
    data["HIGH"] = data["HIGH"].apply(lambda x: round_fixed(x, 2))
    data["LOW"] = data["LOW"].apply(lambda x: round_fixed(x, 2))
    data["CLOSE"] = data["CLOSE"].apply(lambda x: round_fixed(x, 2))
    try:
        data["VOLUME"] = data["VOLUME"].apply(lambda x: int(x))
    except ValueError:  # VOLUME的值可能为空，int函数会报错
        pass

    # 等待3秒
    # time.sleep(3)
    return data

class prt:
    @staticmethod
    def printDataFrameWithMaxRows(df):
        # 设置显示所有行
        # pd.set_option('display.max_rows', None)  # 显示所有行
        # pd.set_option('display.max_columns', None)  # 显示所有列
        # pd.set_option('display.max_colwidth', None)  # 显示所有列宽
        # pd.set_option('display.width', 1000)  # 调整宽度
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        # 设置显示宽度
        pd.set_option('display.width', 1000)  # 设置为适当的值，以便显示足够的列

        printYellowMsg(f"Printing dataframe with max rows")
        print(df)

    @staticmethod
    def printAllColumns(df):
        print(f"Printing all columns in dataframe {df}")
        for col in df.columns:
            print(col)

    @staticmethod
    def print_dict(dictionary):
        for key, value in dictionary.items():
            print(f"{key}: {value}")

    @staticmethod
    def aligned_dict_vertical(d):
        # 计算键和值的最大长度
        max_key_len = max(len(str(key)) for key in d.keys())
        max_value_len = max(len(str(value)) for value in d.values())

        # 生成对齐格式的字符串
        key_format = f"{{:<{max_key_len}}}"
        value_format = f"{{:<{max_value_len}}}"
        row_format = f"{key_format} : {value_format}"

        # 打印标题行
        print(row_format.format("Key", "Value"))
        print("-" * (max_key_len + 3 + max_value_len))

        # 打印字典的键和值
        for key, value in d.items():
            print(row_format.format(key, value))

    @staticmethod
    def aligned_dict_horizontal(d):
        # 计算键和值的最大长度
        max_key_len = max(len(str(key)) for key in d.keys())
        max_value_len = max(len(str(value)) for value in d.values())
        column_width = max(max_key_len, max_value_len)

        # 生成对齐格式的字符串
        key_format = f"{{:<{column_width}}}"
        value_format = f"{{:<{column_width}}}"

        # 打印键作为列标题
        keys_row = " | ".join(key_format.format(key) for key in d.keys())
        print(keys_row)

        # 打印分隔线
        # print("-" * len(keys_row))

        # 打印值
        values_row = " | ".join(value_format.format(value) for value in d.values())
        print(values_row)
        print("")

class data:
    @staticmethod
    def drop_stock_codes(df, deleted_stock):
        """
        删除DataFrame中特定股票代码的行。

        参数:
        df (pd.DataFrame): 包含股票数据的DataFrame。
        deleted_stock (list): 要删除的股票代码列表。

        返回:
        pd.DataFrame: 已删除特定股票代码的DataFrame。
        """
        # 确保删除的股票代码列表是字符串类型
        deleted_stock = [str(code) for code in deleted_stock]

        # 使用isin函数检查股票代码是否在列表中，并返回一个布尔索引
        # 然后使用~（取反）来选择不在列表中的行
        mask = ~df['stock_code'].isin(deleted_stock)

        # 返回过滤后的DataFrame
        return df.loc[mask]

    @staticmethod
    def drop_is_grid_is_0(df):
        # 检查 is_grid 是否为列名
        if "is_grid" not in df.columns:
            raise ValueError(f"is_grid is not a column in the DataFrame")

        # 使用布尔索引来找到 is_grid 列中值为 0 的行，并删除这些行
        df_dropped = df[df["is_grid"] != 0]

        return df_dropped

    @staticmethod
    def drop_targetPositionIsNone(df):
        # 检查 is_grid 是否为列名
        if "target_position_dict" not in df.columns:
            raise ValueError(f"target_position_dict is not a column in the DataFrame")

        # 使用布尔索引来找到 is_grid 列中值为 0 的行，并删除这些行
        df_dropped = df[df["target_position_dict"] == "None"]

        return df_dropped

    @staticmethod
    def drop_columns(df, columns_to_drop):
        """
        删除指定的列

        参数:
        df: pandas.DataFrame
            需要删除列的 DataFrame
        columns_to_drop: list
            需要删除的列名列表

        返回:
        pandas.DataFrame
            删除指定列后的 DataFrame
        """
        return df.drop(columns=columns_to_drop)

    @staticmethod
    def keepColumnsDeleteOthers(df, columns_to_keep):
        """
        保留DataFrame中指定的列，并删除其他所有列。

        参数:
        df (pd.DataFrame): 输入的DataFrame。
        columns_to_keep (list): 需要保留的列名列表。

        返回:
        pd.DataFrame: 只包含指定列的新DataFrame。
        """
        # 检查columns_to_keep中的列是否都在df中
        if not set(columns_to_keep).issubset(df.columns):
            raise ValueError("列名列表中包含不存在于DataFrame中的列")

        # 返回只包含指定列的新DataFrame
        return df[columns_to_keep]

    @staticmethod
    def rename_headers(df, new_headers):
        """
        Rename the headers of a DataFrame.

        Parameters:
        df (pd.DataFrame): The DataFrame whose headers need to be renamed.
        new_headers (list): The list of new header names.

        Returns:
        pd.DataFrame: The DataFrame with renamed headers.
        """
        if len(new_headers) != len(df.columns):
            raise ValueError("The length of new_headers must match the number of columns in the DataFrame.")

        df.columns = new_headers
        return df

    @staticmethod
    def get_df_from_csv(file_path):
        """
        读取CSV文件并转换为DataFrame。

        :param file_path: CSV文件的路径
        :return: 包含CSV数据的DataFrame
        """
        df = pd.read_csv(file_path)
        return df

    @staticmethod
    def read_excel_to_df(file_path, sheet_num):
        """
        读取指定路径和工作表编号的Excel文件，并将其转换为DataFrame。

        参数:
        file_path (str): Excel文件的路径。
        sheet_num (int): 工作表编号，从0开始。

        返回:
        DataFrame: 指定工作表的数据。
        """
        df = pd.read_excel(file_path, sheet_name=sheet_num)
        return df

    @staticmethod
    def getElementFromCol(df, column_name):
        column_values = df[f'{column_name}'].values  # 获取所有元素作为 NumPy 数组
        column_values_list = df[f'{column_name}'].tolist()  # 获取所有元素作为列表
        return column_values_list

    @staticmethod
    def common_stocks(df1, df2):
        """
        比较两个DataFrame，返回它们共同拥有的股票列表。

        参数:
        df1: 第一个DataFrame，必须包含 'stock_code' 列。
        df2: 第二个DataFrame，必须包含 'stock_code' 列。

        返回:
        List: 共同拥有的股票列表。
        """
        # 提取两个DataFrame中的股票代码
        stocks1 = set(df1['stock_code'])
        stocks2 = set(df2['stock_code'])

        # 找到共同拥有的股票
        common = stocks1.intersection(stocks2)

        # 将集合转换为列表返回
        return list(common)

    @staticmethod
    def append_row_to_grid_info_csv(file_path):
        """
        没用！
        Append a row to the end of a CSV file with user input and preset defaults.

        :param file_path: str, path to the CSV file
        """
        # Collect data from the user with defaults where applicable
        new_row = {
            'stock_code': get_input("Enter stock code: ", "AAPL"),
            'judge_grid_start': get_input("Enter judge grid start date (YYYY-MM-DD): ", "2024-01-01"),
            'judge_grid_end': get_input("Enter judge grid end date (YYYY-MM-DD): ", "2024-06-30"),
            'is_grid': get_input("Is grid (yes/no): ", "yes"),
            'test_start_date': get_input("Enter test start date (YYYY-MM-DD): ", "2024-07-01"),
            'test_end_date': get_input("Enter test end date (YYYY-MM-DD): ", "2024-12-31"),
            'best_grid_n': int(get_input("Enter best grid number: ", "10")),
            'min_price': float(get_input("Enter minimum price: ", "120.0")),
            'max_price': float(get_input("Enter maximum price: ", "150.0")),
            'grid_list': get_input("Enter grid list (e.g., [120, 125, 130, 135, 140, 145, 150]): ",
                                   "[120, 125, 130, 135, 140, 145, 150]"),
            'test_ret': float(get_input("Enter test return: ", "0.15")),
            'target_position_dict': get_input(
                "Enter target position dictionary (e.g., {\"2024-07-01\": 100, \"2024-07-02\": 200}): ",
                "{\"2024-07-01\": 100, \"2024-07-02\": 200}")
        }

        # Append the new row to the CSV file
        with open(file_path, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.DictWriter(file, fieldnames=new_row.keys())
            # If the file is empty, write the header
            if file.tell() == 0:
                writer.writeheader()
            writer.writerow(new_row)

    @staticmethod
    def add_row_to_grid_info_csv(csv_path):
        try:
            # 从CSV文件中读取数据
            csv_df = pd.read_csv(csv_path)

            # 提示用户输入Excel文件路径
            excelFilePath = input("请拖入回测结果文件\n")
            # 从Excel文件中读取数据
            # 指定表头在第二行（索引为1）
            excel_df = pd.read_excel(excelFilePath)
            stock_code = os.path.basename(excelFilePath)[:9].replace('_', '.')
            # print(stock_code)

            # 提示用户输入需要的Grid_N值
            grid_n = int(input("请输入股票格子数: \n"))
            # 匹配特定的grid_n行数据
            matched_row = excel_df[excel_df['Grid_N'] == grid_n]

            if matched_row.empty:
                raise ValueError(f"No matching row found for Grid_N={grid_n}")

            # 获取匹配行的字典
            matched_row_dict = matched_row.to_dict(orient='records')[0]

            # YYYYMMDD
            start_date_str = str(matched_row_dict['start_date'])
            end_date_str = str(matched_row_dict['end_date'])

            # 转换为 YYYY_MM_DD

            start_date_format = start_date_str[:4] + '_' + start_date_str[4:6] + '_' + start_date_str[6:]
            end_date_format = end_date_str[:4] + '_' + end_date_str[4:6] + '_' + end_date_str[6:]
            # 创建新的行字典
            new_row_dict = {
                'stock_code': stock_code,
                'judge_grid_start' : 'None',
                'judge_grid_end' : 'None',
                'is_grid' : '1',
                'test_start_date':start_date_format,
                'test_end_date': end_date_format
            }

            new_row_dict['best_grid_n'] = grid_n

            # 去 data 里面找数据文件来计算最小值和最大值 20240627不用了，回测中已经包含
            # min_price, max_price = get_min_max_price_from_backtesting_daily_data(stock_code, matched_row_dict['start_date'], matched_row_dict['end_date'])
            new_row_dict['min_price'] = matched_row_dict['min_price']
            new_row_dict['max_price'] = matched_row_dict['max_price']
            print(f"min price: {matched_row_dict['min_price']}, max price: {matched_row_dict['max_price']}")

            new_row_dict['target_position_dict'] = 'None'
            new_row_dict['grid_list'] = matched_row_dict['Grid_List']

            # 计算ret

            ret_series = matched_row['Ret'].iloc[0]  # 获取'Ret'列的第一个值
            test_return = get_testReturn(ret_series, start_date_str, end_date_str)
            new_row_dict['test_ret'] = test_return
            print(f"test ret: {test_return:.4f}")

            new_row_dict['target_position_dict'] = "None"

            # 确保新行数据与CSV文件的列标题一致
            for column in csv_df.columns:
                if column not in new_row_dict:
                    new_row_dict[column] = input(f"Please enter value for {column}: ")



            # 确认数据
            printYellowMsg("\nCOMFIRM THE DATA:")
            printYellowMsg("====================================")
            prt.print_dict(new_row_dict)
            printYellowMsg("\n====================================")
            decision = input("\n\nenter \"insert\" to insert the data into info file, if not, enter anything else.\n\n")

            if decision == "insert":
                # 将新行数据插入到CSV文件中
                csv_df = csv_df.append(new_row_dict, ignore_index=True)
                csv_df.to_csv(csv_path, index=False)

                print("New row added successfully!")
                input("press enter to continue")
            else:
                input("returning to last menu...")
                return
        except Exception as e:
            printRedMsg(f"error when adding row to grid info csv:{e}")
            input("")
            return

    @staticmethod
    def convert_date_format_at1stCol(df):
        """
        将日期列从YYYYMMDD或YYYY/MM/DD格式转换为YYYY-MM-DD格式,目前只适配YYYYMMDD
        :param df: 输入的DataFrame
        :return: 转换后的DataFrame
        """
        try:
            # 假设日期总是在第一列
            date_column = df.iloc[:, 0].astype(str)

            # 处理YYYYMMDD格式
            mask_yyyymmdd = date_column.str.match(r'^\d{8}$')
            df.iloc[:, 0][mask_yyyymmdd] = pd.to_datetime(date_column[mask_yyyymmdd], format='%Y%m%d').dt.strftime(
                '%Y-%m-%d')

            # 处理YYYY/MM/DD格式
            mask_yyyy_mm_dd = date_column.str.match(r'^\d{4}/\d{2}/\d{2}$')
            df.iloc[:, 0][mask_yyyy_mm_dd] = pd.to_datetime(date_column[mask_yyyy_mm_dd],
                                                                format='%Y/%m/%d').dt.strftime('%Y-%m-%d')

        except Exception as e:
            print(f"Error converting date format: {e}")
            return None

        return df

class plot:
    @staticmethod
    def parse_details(details):
        """
        解析GRID_DETAIL字段，提取价格和数量对。
        :param details: 包含价格和数量对的字符串，格式为 '价格:数量;价格:数量;...'
        :return: 价格和数量对的列表
        """
        price_quantity_pairs = []
        for item in details.split(';'):
            if item:  # 确保每个对不是空的
                price_quantity_pairs.append(item)
        return price_quantity_pairs

    @staticmethod
    def plot_data_grid_paras_txt(df):
        """
        为每个股票ID绘制价格和数量图。
        :param df: 包含股票数据的DataFrame
        """
        num_stocks = len(df)  # 获取股票ID的数量
        # 创建子图，每个子图宽度为2.08英寸（200像素），高度为6英寸，子图之间间隔1英寸
        fig, axs = plt.subplots(1, num_stocks, figsize=(num_stocks * (200 / 96 + 1), 6), sharey=False)
        fig.subplots_adjust(wspace=1, bottom=0.2)  # 设置子图之间的间隔为1英寸

        for ax, stock_id in zip(axs, df['STOCK_ID']):
            # 找到与当前股票ID对应的行
            row = df[df['STOCK_ID'] == stock_id]
            if not row.empty:
                details = row.iloc[0]['GRID_DETAIL']  # 提取GRID_DETAIL字段
                marker_value = int(row.iloc[0]['HOLDING'])  # 提取HOLDING字段

                price_quantity_pairs = plot.parse_details(details)  # 解析价格和数量对
                prices, quantities = zip(*[pair.split(':') for pair in price_quantity_pairs])  # 拆分为价格和数量
                quantities = list(map(int, quantities))  # 将数量转换为整数

                ax.plot(prices, quantities, marker='o', linestyle='')  # 绘制价格和数量的散点图

                # 标记特定的点
                if marker_value in quantities:
                    marker_index = quantities.index(marker_value)  # 找到marker_value在数量列表中的索引
                    ax.scatter(prices[marker_index], quantities[marker_index], color='red', zorder=5)  # 标记红色点
                    ax.text(prices[marker_index], quantities[marker_index],
                            f' {prices[marker_index]}:{quantities[marker_index]}', fontsize=12, color='red')  # 添加文本标签

                # 检查HOLDING值是否在GRID_DETAIL的数量范围内
                if marker_value <= min(quantities) or marker_value >= max(quantities):
                    ax.text(0.5, -0.2, 'BREACHED!!!', color='red', fontsize=15, transform=ax.transAxes, ha='center')

                ax.set_title(f'STOCK ID: {stock_id}')  # 设置子图标题为股票ID
                ax.set_xlabel('Price')  # 设置x轴标签为Price
                ax.set_ylabel('Quantity')  # 设置y轴标签为Quantity
                ax.grid(True)  # 显示网格线
        # 绘制完所有子图后，添加大标题
        plt.suptitle('GRID PARAS', fontsize=16)

        plt.show()  # 显示绘图

    @staticmethod
    def stock_min_max_latest(stock_min_max_latest):
        # 创建子图
        num_stocks = len(stock_min_max_latest)
        fig, axes = plt.subplots(nrows=1, ncols=num_stocks, figsize=(8, 6), sharey=False)
        fig.subplots_adjust(wspace=0.5, top=0.912, right=0.974, bottom=0.1)  # 设置子图之间的间隔为1英寸

        # 处理axes，当只有一个子图时将其转换为列表
        if num_stocks == 1:
            axes = [axes]  # 将单个子图对象转换为列表
            
            
        # 绘制每个子图
        # 这一行遍历了stock_min_max_latest DataFrame 中的每一行，并返回行索引（index）和行数据（row）。enumerate() 函数用于同时获得行索引和行数据，并在每次循环中增加计数器 i。
        for i, (index, row) in enumerate(stock_min_max_latest.iterrows()):
            ax = axes[i]  # 根据当前迭代的计数器 i，选择了对应的子图 ax，以便对其进行操作。
            # 这一行绘制了一个水平区间，水平轴表示价格范围，竖直轴为零。row['min_price'] 和 row['max_price'] 分别表示当前股票的最低价和最高价。
            ax.plot([0, 0], [row['min_price'], row['max_price']], color='gray', linestyle='-', linewidth=2)
            # 根据最新价在价格区间内外使用绿色或红色绘制了一个点。如果最新价在最低价和最高价之间，则使用绿色表示，否则使用红色。
            ax.scatter(0, row['latest_price'],
                       color='green' if row['min_price'] <= row['latest_price'] <= row['max_price'] else 'red')

            # 在最新价格点旁边添加对应的数值
            ax.text(0.05, row['latest_price'], f"{row['latest_price']:.2f}", verticalalignment='center', fontsize=10,
                    color='black')

            ax.set_aspect('auto')
            ax.set_ylim(row['min_price'] - 1, row['max_price'] + 1)  # 设置了竖直轴的范围，确保价格区间的可视化合理。
            ax.set_xlim(-1, 1)  # 设置了水平轴的范围，在这里固定为 (-1, 1)。
            # ax.set_title(row['stock_code']) # 设置了子图的标题，即当前股票的股票代码。
            ax.set_xlabel(row['stock_code'])  # 将股票代码作为 x 轴标签
            ax.set_yticks([row['min_price'], row['max_price']])  # 设置了竖直轴上的刻度，这里是最低价和最高价。
            ax.set_xticks([])  # 由于水平轴的范围已经设置为 (-1, 1)，因此我们在竖直轴上不显示刻度

        # 调整子图之间的间距
        # plt.tight_layout()

        # 绘制完所有子图后，添加大标题
        fig.suptitle('MONITOR ON TRADING STOCK', fontsize=16)

        # 显示图表
        plt.show()

    @staticmethod
    def create_holding(df):
        # 创建子图
        num_stocks = len(df)
        if num_stocks == 0:
            printRedMsg("NO DATA TO DRAW, RETURNING...\n")
            return
        fig, axes = plt.subplots(nrows=1, ncols=num_stocks, figsize=(8, 6), sharey=False)
        fig.subplots_adjust(wspace=0.5, top=0.912, right=0.974)  # 设置子图之间的间隔为1英寸
        # 绘制每个子图
        # 这一行遍历了stock_min_max_latest DataFrame 中的每一行，并返回行索引（index）和行数据（row）。enumerate() 函数用于同时获得行索引和行数据，并在每次循环中增加计数器 i。
        for i, (index, row) in enumerate(df.iterrows()):
            ax = axes[i]  # 根据当前迭代的计数器 i，选择了对应的子图 ax，以便对其进行操作。
            # 这一行绘制了一个水平区间，水平轴表示价格范围，竖直轴为零。row['min_price'] 和 row['max_price'] 分别表示当前股票的最低价和最高价。
            ax.plot([0, 0], [row['min_price'], row['max_price']], color='gray', linestyle='-', linewidth=2)
            ax.plot([0, 0], [row['ten_percent'], row['thirty_percent']], color='green', linestyle='-', linewidth=3)
            # 根据最新价在价格区间内外使用绿色或红色绘制了一个点。如果最新价在最低价和最高价之间，则使用绿色表示，否则使用红色。
            ax.scatter(0, row['close'],
                       color='green' if row['ten_percent'] <= row['close'] <= row['thirty_percent'] else 'red')

            # 在最新价格点旁边添加对应的数值
            ax.text(0.05, row['close'], f"{row['close']:.2f}", verticalalignment='center',
                    fontsize=10,
                    color='black')

            ax.set_aspect('auto')
            ax.set_ylim(row['min_price'] - 1, row['max_price'] + 1)  # 设置了竖直轴的范围，确保价格区间的可视化合理。
            ax.set_xlim(-1, 1)  # 设置了水平轴的范围，在这里固定为 (-1, 1)。
            # ax.set_title(row['stock_code']) # 设置了子图的标题，即当前股票的股票代码。
            ax.set_xlabel(row['stock_code'])  # 将股票代码作为 x 轴标签
            ax.set_yticks([row['min_price'], row['max_price'], row['ten_percent'],
                           row['thirty_percent']])  # 设置了竖直轴上的刻度，这里是最低价和最高价。
            ax.set_xticks([])  # 由于水平轴的范围已经设置为 (-1, 1)，因此我们在竖直轴上不显示刻度

        # 调整子图之间的间距
        plt.tight_layout()

        # 绘制完所有子图后，添加大标题
        fig.suptitle('MONITOR ON TRADING STOCK', fontsize=16)

        # 显示图表
        plt.show()

    @staticmethod
    def create_holding_2(df, part2_stock_10percent_30percent):
        # Check if df is None
        if df is None:
            printGreenMsg("NO STOCK IS IN THE PRICE RANGE OF CREATING HOLDING")
            return
        # 获取所有唯一的ts_code
        ts_codes = df['ts_code'].unique()

        # 检查是否有ts_code
        if len(ts_codes) == 0:
            printGreenMsg("NO STOCK IS IN THE PRICE RANGE OF CREATING HOLDING")
            return

        # 创建子图
        fig, axes = plt.subplots(len(ts_codes), 1, figsize=(12, 8), sharex=False)

        # 检查 axes 是否为单个 AxesSubplot 对象，如果是，将其转换为数组
        if not isinstance(axes, (np.ndarray, list)):
            axes = [axes]

        # 画图
        for i, ts_code in enumerate(ts_codes):
            sub_df = df[df['ts_code'] == ts_code]
            ten_percent = part2_stock_10percent_30percent.loc[
                part2_stock_10percent_30percent['stock_code'] == ts_code, 'ten_percent'].values[0]
            thirty_percent = part2_stock_10percent_30percent.loc[
                part2_stock_10percent_30percent['stock_code'] == ts_code, 'thirty_percent'].values[0]

            ax = axes[i]
            ax.plot(sub_df['trade_date'].str[-4:][::-1], sub_df['close'][::-1], marker='o', label='Close')
            # ax.axhline(y=sub_df['pre_close'].iloc[0], color='b', linestyle='--', label='Pre Close')
            ax.axhline(y=ten_percent, color='g', linestyle='--', label='10%')
            ax.axhline(y=thirty_percent, color='r', linestyle='--', label='30%')
            ax.set_ylim(sub_df['close'].min() - 1, sub_df['close'].max() + 1)
            ax.set_ylabel('Price')
            ax.legend()
            ax.set_title(f'Stock: {ts_code}')

        # 设置公共的x轴标签
        plt.xlabel('Trade Date (MMDD)')

        # 调整布局
        # plt.tight_layout()
        plt.show()


class grid:
    @staticmethod
    def grid_paras_txt_to_dataframe(file_path):
        data = []

        with open(file_path, 'r') as file:
            for line in file:
                parts = line.strip().split('=')
                if len(parts) >= 3 and parts[2]:  # 确保第二个等号后面有内容
                    record_id = parts[0]
                    value = parts[1]
                    details = parts[2]
                    data.append([record_id, value, details])

        df = pd.DataFrame(data, columns=['STOCK_ID', 'HOLDING', 'GRID_DETAIL'])
        return df


def get_fundname(fund_account):
    if fund_account == 1260016888:
        fundname = 'FL'
    elif fund_account == 480149909:
        fundname = 'FL18'
    elif fund_account == 480151137:
        fundname = 'CF15SC'
    elif fund_account == 3050005040:
        fundname = 'CF15XZ'
    elif fund_account == 480160777:
        fundname = 'FL22SCA'
    elif fund_account == 3050003937:
        fundname = 'FL22SCB'
    elif fund_account == 480167623:
        fundname = 'HT02XY'

    return fundname

def get_chineseFundname(fund_account):
    if fund_account == 1260016888:
        return "中天飞泸 2 号"
    elif fund_account == 480149909:
        return "中天飞泸私募投资基金"
    elif fund_account == 480151137 or fund_account == 3050005040:
        return "中天乘风私募"
    elif fund_account == 480160777 or fund_account == 3050003937:
        return "中天飞泸 1 号"
    elif fund_account == 480167623:
        return "中天宏图 2 号"
    else:
        return "Unknown fund account"
        


def iiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii():
    ...


# ====================================================================================================================
# ============================================   主要函数   ===========================================================
# ====================================================================================================================

def before0920processFL22SC():
    oriFl22 = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SC\Sell_Buy_List_FL22SC"
    os.system("cls")
    today = getDate('')
    print(f'today is {today}')

    oriBuy = rf"{oriFl22}\{buy}FL22SC_{today}.csv"
    oriSell = rf"{oriFl22}\{sell}FL22SC_{today}.csv"
    oriJrcc = rf"{oriFl22}\{jrcc}FL22SC_{today}.csv"

    # 20240524 百度网盘同步空间移动到了 37，分单程序都自动运行了
    while False:
        # 检查原始信号在不在，不在直接退
        print("checking if the original signals are exist...")
        if ((ifExist(oriBuy) and ifExist(oriSell) and ifExist(oriJrcc)) == False):
            printRedMsg("There are no original signal!!!\nreturning...")
            input("")
            return
        else:
            printGreenMsg("original signal EXIST")
            input("press any key to continue...")

        # 然后移动到目标文件夹
        printYellowMsg("\nMoving original signals to the divide_order_account path...")
        divideFl22 = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SC\Sell_Buy_List_FL22SC"
        # TODO testing
        copy_file(oriBuy, divideFl22)
        copy_file(oriSell, divideFl22)
        copy_file(oriJrcc, divideFl22)
        printYellowMsg("Move done, now checking if moving is complete...")
        if not (ifExist(rf"{divideFl22}\{buy}FL22SC_{today}.csv") and ifExist(
                rf"{divideFl22}\{sell}FL22SC_{today}.csv") and ifExist(rf"{divideFl22}\{jrcc}FL22SC_{today}.csv")):
            printRedMsg("Moving is not complete!!!\nreturning...")
            input("")
            return
        else:
            printGreenMsg("Moving is complete, press any key to run the 1、盘前订单拆分.py...")
            input(" ")

        # 运行分单脚本
        # FIXME 这个功能不能正常运行 用多线程
        # os.system(r"C:\PROGRA~1\Python36\python.exe" r"C:\Users\Administrator\Desktop\divide_order_account\projects\1、盘前订单拆分.py")
        # input("1、盘前订单拆分.py is ending, press any key to continue...")
        printYellowMsg("Plz run the 1、盘前订单拆分.py by hand, press any key to continue after finishing running...")
        input(" ")

    # 检测拆分是否成功

    # divide 文件夹下的三个文件
    # A
    divideSCAbuy = rf"{divide_SCA}\{buy}FL22SCA_{today}.csv"
    divideSCAsell = rf"{divide_SCA}\{sell}FL22SCA_{today}.csv"
    divideSCAjrcc = rf"{divide_SCA}\{jrcc}FL22SCA_{today}.csv"
    # B
    divideSCBbuy = rf"{divide_SCB}\{buy}FL22SCB_{today}.csv"
    divideSCBsell = rf"{divide_SCB}\{sell}FL22SCB_{today}.csv"
    divideSCBjrcc = rf"{divide_SCB}\{jrcc}FL22SCB_{today}.csv"
    isDivideSCAExist = ifExist(divideSCAbuy) and ifExist(divideSCAsell) and ifExist(divideSCAjrcc)
    isDivideSCBExist = ifExist(divideSCBbuy) and ifExist(divideSCBsell) and ifExist(divideSCBjrcc)
    isDivideSCABExist = isDivideSCAExist and isDivideSCBExist

    if not (isDivideSCABExist):
        printRedMsg("signal dividing fail!!!\nreturning...")
        input("")
        return
    else:
        printGreenMsg("signal dividing is DONE.")
        input(" ")

    printYellowMsg("Copying files, 6 in total...")
    # 传回文件
    # TODO 这里直接
    # TODO testing
    # SCA
    copy_file(divideSCAbuy, ori_SCA)
    copy_file(divideSCAsell, ori_SCA)
    copy_file(divideSCAjrcc, ori_SCA)
    # SCB
    copy_file(divideSCBbuy, ori_SCB)
    copy_file(divideSCBsell, ori_SCB)
    copy_file(divideSCBjrcc, ori_SCB)
    printYellowMsg("Copying finished, examine required...")

    # 检查文件是否复制成功
    # 原始信号文件夹的三个文件
    # A
    oriSCAbuy = rf"{ori_SCA}\{buy}FL22SCA_{today}.csv"
    oriSCAsell = rf"{ori_SCA}\{sell}FL22SCA_{today}.csv"
    oriSCAjrcc = rf"{ori_SCA}\{jrcc}FL22SCA_{today}.csv"
    # B
    oriSCBbuy = rf"{ori_SCB}\{buy}FL22SCB_{today}.csv"
    oriSCBsell = rf"{ori_SCB}\{sell}FL22SCB_{today}.csv"
    oriSCBjrcc = rf"{ori_SCB}\{jrcc}FL22SCB_{today}.csv"
    isOriSCAExist = ifExist(oriSCAbuy) and ifExist(oriSCAsell) and ifExist(oriSCAjrcc)
    isOriSCBExist = ifExist(oriSCBbuy) and ifExist(oriSCBsell) and ifExist(oriSCBjrcc)
    isOriSCABExist = isOriSCAExist and isOriSCBExist

    if not (isOriSCABExist):
        printRedMsg("divided signals copy fail!!!\nreturning...")
        input("")
        return
    else:
        printGreenMsg("divided signals copy DONE.")
        input(" ")

    printGreenMsg("function before0920processFL22SC is ending, press any key to continue...")
    input("")


def realTimeSignalMoveForFL22SC():
    os.system("cls")
    # 先定义一下, 实时信号是直接传到 ori 文件夹的
    today = getDate('')
    print(f'today is {today}')
    nom = rf"{ori_SC}\{nom_}{today}.txt"
    nom2 = rf"{ori_SC}\{nom2_}{today}.txt"
    noa = rf"{ori_SC}\{noa_}{today}.txt"
    noa2 = rf"{ori_SC}\{noa2_}{today}.txt"

    # FIXME 才意识到可以，选择数字之后，比如说2，就直接 nowTimeNode = m2 就可以很方便了。。。
    # FIXME 没有意识到一个信号拆分后可能是A有信号B是no信号！！ 20240430 出现了这个问题，直接爆红，只能手动复制
    # FIXMED 只要早上有一次 No morning, 那接着的每一次都将要复制一次
    # 选择实时信号
    while True:
        printYellowMsg("\n请确认实时信号节点？(1/2/3/4):\n1：morning; 2：morning2Two; 3：afternoon; 4：afternoon2Two\n")
        global g_realTimeNode
        g_realTimeNode = input("")
        if g_realTimeNode == "1":
            rtsBuy = ""
            rtsSell = rf"{ori_SC}\{sell}FL22SC{m}_{today}.csv"
            rtsJRCC = rf"{ori_SC}\{jrcc}FL22SC{m}_{today}.csv"
            rtsNoSignal = nom
            divideAsell = rf"{divide_SCA}\{sell}FL22SCA{m}_{today}.csv"
            divideAjrcc = rf"{divide_SCA}\{jrcc}FL22SCA{m}_{today}.csv"
            divideBsell = rf"{divide_SCB}\{sell}FL22SCB{m}_{today}.csv"
            divideBjrcc = rf"{divide_SCB}\{jrcc}FL22SCB{m}_{today}.csv"
            divideAno = rf"{divide_SCA}\{nom_}{today}.txt"
            divideBno = rf"{divide_SCB}\{nom_}{today}.txt"

            oriAbuy = rf" "
            oriAsell = rf"{ori_SCA}\{sell}FL22SCA{m}_{today}.csv"
            oriAjrcc = rf"{ori_SCA}\{jrcc}FL22SCA{m}_{today}.csv"
            oriAno = rf"{ori_SCA}\{nom_}{today}.txt"

            oriBsell = rf"{ori_SCB}\{sell}FL22SCB{m}_{today}.csv"
            oriBjrcc = rf"{ori_SCB}\{jrcc}FL22SCB{m}_{today}.csv"
            oriBno = rf"{ori_SCB}\{nom_}{today}.txt"

            break
        elif g_realTimeNode == "2":
            rtsBuy = ""
            rtsSell = rf"{ori_SC}\{sell}FL22SC{m2}_{today}.csv"
            rtsJRCC = rf"{ori_SC}\{jrcc}FL22SC{m2}_{today}.csv"
            rtsNoSignal = nom2
            divideAsell = rf"{divide_SCA}\{sell}FL22SCA{m2}_{today}.csv"
            divideAjrcc = rf"{divide_SCA}\{jrcc}FL22SCA{m2}_{today}.csv"
            divideBsell = rf"{divide_SCB}\{sell}FL22SCB{m2}_{today}.csv"
            divideBjrcc = rf"{divide_SCB}\{jrcc}FL22SCB{m2}_{today}.csv"
            divideAno = rf"{divide_SCA}\{nom2_}{today}.txt"
            divideBno = rf"{divide_SCB}\{nom2_}{today}.txt"

            oriAbuy = rf" "
            oriAsell = rf"{ori_SCA}\{sell}FL22SCA{m2}_{today}.csv"
            oriAjrcc = rf"{ori_SCA}\{jrcc}FL22SCA{m2}_{today}.csv"
            oriAno = rf"{ori_SCA}\{nom2_}{today}.txt"

            oriBsell = rf"{ori_SCB}\{sell}FL22SCB{m2}_{today}.csv"
            oriBjrcc = rf"{ori_SCB}\{jrcc}FL22SCB{m2}_{today}.csv"
            oriBno = rf"{ori_SCB}\{nom2_}{today}.txt"
            break
        elif g_realTimeNode == "3":
            rtsBuy = ""
            rtsSell = rf"{ori_SC}\{sell}FL22SC{a}_{today}.csv"
            rtsJRCC = rf"{ori_SC}\{jrcc}FL22SC{a}_{today}.csv"
            rtsNoSignal = noa
            divideAsell = rf"{divide_SCA}\{sell}FL22SCA{a}_{today}.csv"
            divideAjrcc = rf"{divide_SCA}\{jrcc}FL22SCA{a}_{today}.csv"
            divideBsell = rf"{divide_SCB}\{sell}FL22SCB{a}_{today}.csv"
            divideBjrcc = rf"{divide_SCB}\{jrcc}FL22SCB{a}_{today}.csv"
            divideAno = rf"{divide_SCA}\{noa_}{today}.txt"
            divideBno = rf"{divide_SCB}\{noa_}{today}.txt"

            oriAbuy = rf" "
            oriAsell = rf"{ori_SCA}\{sell}FL22SCA{a}_{today}.csv"
            oriAjrcc = rf"{ori_SCA}\{jrcc}FL22SCA{a}_{today}.csv"
            oriAno = rf"{ori_SCA}\{noa_}{today}.txt"

            oriBsell = rf"{ori_SCB}\{sell}FL22SCB{a}_{today}.csv"
            oriBjrcc = rf"{ori_SCB}\{jrcc}FL22SCB{a}_{today}.csv"
            oriBno = rf"{ori_SCB}\{noa_}{today}.txt"

            break
        elif g_realTimeNode == "4":
            rtsSell = ""
            rtsBuy = rf"{ori_SC}\{buy}FL22SC{a2}_{today}.csv"
            rtsJRCC = rf"{ori_SC}\{jrcc}FL22SC{a2}_{today}.csv"
            rtsNoSignal = noa2
            divideAbuy = rf"{divide_SCA}\{buy}FL22SCA{a2}_{today}.csv"
            divideAjrcc = rf"{divide_SCA}\{jrcc}FL22SCA{a2}_{today}.csv"
            divideBbuy = rf"{divide_SCB}\{buy}FL22SCB{a2}_{today}.csv"
            divideBjrcc = rf"{divide_SCB}\{jrcc}FL22SCB{a2}_{today}.csv"
            divideAno = rf"{divide_SCA}\{noa2_}{today}.txt"
            divideBno = rf"{divide_SCB}\{noa2_}{today}.txt"

            oriAbuy = rf"{ori_SCA}\{buy}FL22SCA{a2}_{today}.csv"
            oriAsell = rf" "
            oriAjrcc = rf"{ori_SCA}\{jrcc}FL22SCA{a2}_{today}.csv"
            oriAno = rf"{ori_SCA}\{noa2_}{today}.txt"

            oriBbuy = rf"{ori_SCB}\{buy}FL22SCB{a2}_{today}.csv"
            oriBjrcc = rf"{ori_SCB}\{jrcc}FL22SCB{a2}_{today}.csv"
            oriBno = rf"{ori_SCB}\{noa2_}{today}.txt"
            break
        elif g_realTimeNode == "quit":
            printBlueMsg("returning to main menu...\n")
            input("")
            return
        else:
            printRedMsg("无效的选项，请重新输入！")

    # 检查实时信号到了没有
    printYellowMsg("checking whether the real time signals are arrive...")
    isNoSignal = ifExist(rtsNoSignal)
    isRealTimeSignal = ((ifExist(rtsSell) or ifExist(rtsBuy)) and ifExist(rtsJRCC))
    isRealTimeSignalExist = isNoSignal or isRealTimeSignal
    if not isRealTimeSignalExist:
        printRedMsg("real time signals are NOT arrive!!!\nreturning...")
        input("")
        return
    else:
        printGreenMsg("real time signals ARE arrive.")
        input(" ")

    # # 复制文件进入 divide 文件
    # if isNoSignal:
    #     # 如果是no信号
    #     copy_file(rtsNoSignal, ori_SC)
    # elif isRealTimeSignal:
    #     if choice == "4":
    #         copy_file(rtsBuy, ori_SC)
    #         copy_file(rtsJRCC, ori_SC)
    #     copy_file(rtsSell, ori_SC)
    #     copy_file(rtsJRCC, ori_SC)

    # TODO 等待 30s 让程序拆分文件
    # TODO 检测信号是否复制成功

    # 检测信号是否拆分成功
    if g_realTimeNode == "4":
        isNoDivide = ifExist(divideAno) and ifExist(divideBno)
        isSignalDivide = ifExist(divideAbuy) and ifExist(divideAjrcc) and ifExist(divideBbuy) and ifExist(divideBjrcc)
        isDivide = (isNoDivide) or (isSignalDivide)
        if not isDivide:
            printRedMsg("afternoon2Two signal is not divided!!!\nreturning...")
            input("")
            return
        else:
            printGreenMsg("signal is divided.")
            input(" ")

    else:
        # choice == 1 2 3
        isNoDivide = ifExist(divideAno) and ifExist(divideBno)
        isSignalDivide = ifExist(divideAsell) and ifExist(divideAjrcc) and ifExist(divideBsell) and ifExist(divideBjrcc)
        isDivide = (isNoDivide) or (isSignalDivide)
        if not isDivide:
            printRedMsg("signal is not divided!!!\nreturning...")
            input("")
            return
        else:
            printGreenMsg("signal is divided.")
            input(" ")

    # 把拆分好的信号复制回来
    if g_realTimeNode == "4":
        copy_file(divideAno, ori_SCA)
        copy_file(divideAbuy, ori_SCA)
        copy_file(divideAjrcc, ori_SCA)
        copy_file(divideBbuy, ori_SCB)
        copy_file(divideBjrcc, ori_SCB)
        copy_file(divideBno, ori_SCB)
    else:
        copy_file(divideAsell, ori_SCA)
        copy_file(divideAno, ori_SCA)
        copy_file(divideAjrcc, ori_SCA)
        copy_file(divideBsell, ori_SCB)
        copy_file(divideBjrcc, ori_SCB)
        copy_file(divideBno, ori_SCB)

    # 验证复制回来的信号
    # A
    if g_realTimeNode == "4":
        isNoOri = ifExist(oriAno) and ifExist(oriBno)
        isSignalOri = ifExist(oriAbuy) and ifExist(oriAjrcc) and ifExist(oriBbuy) and ifExist(oriBjrcc)
        isOri = (isNoOri) or (isSignalOri)
        if not isOri:
            printRedMsg("divided afternoon2Two signal is not copy well!!!\nreturning...")
            input("")
            return
        else:
            printGreenMsg("signal is copied")
            input(" ")

    else:
        # choice == 1 2 3
        isNoOri = ifExist(oriAno) and ifExist(oriBno)
        isSignalOri = ifExist(oriAsell) and ifExist(oriAjrcc) and ifExist(oriBsell) and ifExist(oriBjrcc)
        isOri = (isNoOri) or (isSignalOri)
        if not isOri:
            printRedMsg(f"divided signal is not copy well!!!\nreturning...")
            input("")
            return
        else:
            printGreenMsg("signal is copied")
            input(" ")

        # 完成
        printGreenMsg("Action done, returning to main menu...")
        input(" ")


def realTimeSignalMoveForHT02():
    # 包含原始信号和实时信号
    os.system("cls")
    input("Now processing HT02XY")
    today = getDate('')
    print(f'today is {today}')
    divide_HT = r"C:\Users\Administrator\Desktop\divide_order_account\HT02XY\Sell_Buy_List_HT02XY"
    ori_HT = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\HT02XY\Sell_Buy_List_HT02XY"

    nom = rf"{divide_HT}\{nom_}{today}.txt"
    nom2 = rf"{divide_HT}\{nom2_}{today}.txt"
    noa = rf"{divide_HT}\{noa_}{today}.txt"
    noa2 = rf"{divide_HT}\{noa2_}{today}.txt"
    # HT02XYsell,HT02XYjrcc,HT02XYbuy,HT02XYno = None
    # 原始信号
    originalSignalBuy = rf"{divide_HT}\{buy}HT02XY_{today}.csv"
    originalSignalSell = rf"{divide_HT}\{sell}HT02XY_{today}.csv"
    originalSignalJrcc = rf"{divide_HT}\{jrcc}HT02XY_{today}.csv"

    print(g_realTimeNode)
    input("")

    while True:
        # printYellowMsg("\n请确认实时信号节点？(1/2/3/4):\n1：morning; 2：morning2Two; 3：afternoon; 4：afternoon2Two\n")
        # g_realTimeNode = input("")
        if g_realTimeNode == "1":
            HT02XYsell = rf"{divide_HT}\{sell}HT02XY{m}_{today}.csv"
            HT02XYjrcc = rf"{divide_HT}\{jrcc}HT02XY{m}_{today}.csv"
            HT02XYbuy = ""
            HT02XYno = nom
            break
        elif g_realTimeNode == "2":
            HT02XYsell = rf"{divide_HT}\{sell}HT02XY{m2}_{today}.csv"
            HT02XYjrcc = rf"{divide_HT}\{jrcc}HT02XY{m2}_{today}.csv"
            HT02XYbuy = ""
            HT02XYno = nom2
            break
        elif g_realTimeNode == "3":
            HT02XYsell = rf"{divide_HT}\{sell}HT02XY{a}_{today}.csv"
            HT02XYjrcc = rf"{divide_HT}\{jrcc}HT02XY{a}_{today}.csv"
            HT02XYbuy = ""
            HT02XYno = noa
            break
        elif g_realTimeNode == "4":
            HT02XYjrcc = rf"{divide_HT}\{jrcc}HT02XY{a2}_{today}.csv"
            HT02XYsell = ""
            HT02XYbuy = rf"{divide_HT}\{buy}HT02XY{a2}_{today}.csv"
            HT02XYno = noa2
            break
        elif g_realTimeNode == "quit":
            printBlueMsg("returning to main menu...\n")
            input("")
            return
        else:
            printRedMsg("无效的选项，请重新输入！")

    # 检查 HT02XY 拆分了没有
    printYellowMsg("checking whether the real time signals are arrive...")
    isNoSignal = ifExist(HT02XYno)
    isRealTimeSignal = ((ifExist(HT02XYsell) or ifExist(HT02XYbuy)) and ifExist(HT02XYjrcc))
    isHT02XYsignalExist = isNoSignal or isRealTimeSignal
    if not isHT02XYsignalExist:
        printRedMsg("HT02XY signal is not exist!\nreturning...")
        input("")
        return
    else:
        printGreenMsg("HT02XY signals ARE arrive.")
        input("")

    # 复制到扫单文件夹
    if g_realTimeNode == "4":
        copy_file(HT02XYbuy, ori_HT)
        copy_file(HT02XYjrcc, ori_HT)
        copy_file(HT02XYno, ori_HT)
    else:
        copy_file(HT02XYsell, ori_HT)
        copy_file(HT02XYjrcc, ori_HT)
        copy_file(HT02XYno, ori_HT)

    # 验证复制回来的信号
    printYellowMsg("省略验证，请自行检查\n")
    printGreenMsg("Action done, returning to main menu...")
    input(" ")


def dataCollectorOn40():
    today = getDate('')
    print(f'today is {today}')
    global g_yesterday
    if g_yesterday == None:
        g_yesterday = input("请输入上一个交易日的日期")
    printYellowMsg("PLZ CHECK THIS FUNCTION IS ONLY WORKING ON 40")
    printYellowMsg("Deleting data in folder toLZY")
    destinationPath = r"C:\Users\progene014\Desktop\toLZY\data"

    # os.chmod(destinationPath, 0o777)
    if erase_folder_contents(destinationPath):
        printGreenMsg("Data has been deleted.")
        input("")
    else:
        printRedMsg("Data has NOT been deleted, returning to main menu...")
        input(" ")
        return

    # limit_price_file
    limitPricePath = r"D:\limit_price_projects\limit_price_file"
    copy_latest_files(limitPricePath, destinationPath, 1)

    # data
    oriDataSCA = r"D:\hutao\projects\数据分析\FL22SC\data_FL22SCA"
    oriDataSCB = r"D:\hutao\projects\数据分析\FL22SC\data_FL22SCB"
    desDataSCA = r"C:\Users\progene014\Desktop\toLZY\data\A_data"
    desDataSCB = r"C:\Users\progene014\Desktop\toLZY\data\B_data"
    copy_latest_files(oriDataSCA, desDataSCA, 3)
    copy_latest_files(oriDataSCB, desDataSCB, 3)

    # fl22sc 的 format data
    oriDataSC_format = r"D:\hutao\projects\数据分析\FL22SC\format_data"
    desDataA_format = r"C:\Users\progene014\Desktop\toLZY\data\A_format_data"
    desDataB_format = r"C:\Users\progene014\Desktop\toLZY\data\B_format_data"
    copy_files_with_string_limited(oriDataSC_format, desDataA_format, "SCA", 3)
    copy_files_with_string_limited(oriDataSC_format, desDataB_format, "SCB", 3)

    # ht02zs 的 format data
    oriDataHT02_format = r"D:\hutao\projects\数据分析\HT02\format_data"
    desDataHT02ZS_format = r"C:\Users\progene014\Desktop\toLZY\data\HT02ZS_format_data"
    desDataHT02XY_format = r"C:\Users\progene014\Desktop\toLZY\data\HT02XY_format_data"
    copy_files_with_string_limited(oriDataHT02_format, desDataHT02ZS_format, "HT02ZS", 3)
    copy_files_with_string_limited(oriDataHT02_format, desDataHT02XY_format, "HT02XY", 3)

    printYellowMsg("check if files are lastest, making ZIP file now")
    # NQ 修改这个参数以把 zip 文件放到你想要的位置
    whereToZip = rf"C:\Users\progene014\Desktop\toLZY\data{g_yesterday}.zip"
    zip_folder(destinationPath, whereToZip)
    ifExist(whereToZip)
    printGreenMsg("function ending, returning to main menu...")
    input(" ")


# 收盘拆分的导出数据的检查
import time

def checkExportData():
    today = getDate('')
    print(f'today is {today}')
    printYellowMsg("now checking if smt_data exist...")
    
    smtData = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\smt_data"
    asset = rf"{smtData}\asset_{today}.csv"
    trans = rf"{smtData}\transaction_{today}.csv"
    order = rf"{smtData}\order_{today}.csv"
    holdin = rf"{smtData}\holding_{today}.csv"

    # 循环检查文件是否存在
    while True:
        smtDataExist = ifExist(asset) and ifExist(trans) and ifExist(order) and ifExist(holdin)
        if smtDataExist:
            printGreenMsg("smtData is exist.")
            break
        else:
            printRedMsg("smtData is NOT exist, checking again in 1 second...")
            time.sleep(0.2)

    printYellowMsg(f"NOW RUNNING {smart_divide_path}...")

    thread = threading.Thread(target=run_python_file, args=(smart_divide_path,))
    thread.start()
    thread.join()
    printGreenMsg("Data divide program is finished.")

    # 检查是否拆分完毕
    cf15 = count_files_with_target_field(r"C:\Users\Administrator\Desktop\兴业证券多账户交易\CF15\data", today)
    fl18 = count_files_with_target_field(r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL18\data", today)
    ht02zs = count_files_with_target_field(r"C:\Users\Administrator\Desktop\兴业证券多账户交易\HT02XY\data", today)
    fl22sc = count_files_with_target_field(r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCA\data", today)
    fl22xz = count_files_with_target_field(r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCB\data", today)
    print(cf15, fl18, ht02zs, fl22sc, fl22xz)

    printGreenMsg("process done, returning to main menu")
    input("")



def downloadDataFromServer40():
    ...
    # TODO 未完成，又找不到本地文件夹的错误，而且会下载分享的所有文件。
    # download_dir = r"C:\Users\progene12\Desktop\Start Trading\FL22SC数据分析"
    # html_content = fetchDataFromHttpServer(server40addr)
    # if html_content:
    #     if not os.path.exists(download_dir):
    #         os.makedirs(download_dir)
    #
    #     # 下载文件
    #     download_files_from_html(html_content, server40addr, download_dir)


# TODO 转移到37上之后需要改一切! 疯狂！
def copYesterdayData():
    printYellowMsg("接下来将依次进行：\n\tFL22SC -> format_data\n\tFL22SC -> data\n\tHT02 -> format_data")
    # 解压zip文件夹
    startTradePath = r"C:\Users\Administrator\Desktop\startTrade"
    temp = rf"{startTradePath}\data\temp"
    today = getDate('')
    print(f'today is {today}')
    global g_yesterday
    if g_yesterday == None:
        g_yesterday = input("enter yesterday's date")
    zipPath = rf"{startTradePath}\data{g_yesterday}.zip"
    if not ifExist(zipPath):
        printRedMsg(f"data{g_yesterday}.zip is NOT here, do you want to enter a valid date? y/n")
        wantTo = input("")
        if wantTo == "y" or "Y":
            manualDate = input("Enter the date that is last trade day or the date you want for the zip file.")

        elif wantTo == "n":
            printRedMsg(f"data{g_yesterday}.zip is NOT here, returning to main menu...")
            return

    input("")
    # create_folder(temp)
    unzip_file(zipPath, temp)

    printBlueMsg("FL22SCA -> format_data")
    # ! 移动 FL22SC 文件到 divide 文件夹
    # test = rf"{temp}\test"
    des_A_data = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SCA\data"
    des_A_FormatData = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SCA\format_data"
    des_B_data = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SCB\data"
    des_B_FormatData = r"C:\Users\Administrator\Desktop\divide_order_account\FL22SCB\format_data"
    des_limit_price = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\limit_price"
    copy_files_in_folder(rf"{temp}\A_data", des_A_data)
    copy_files_in_folder(rf"{temp}\A_format_data", des_A_FormatData)
    copy_files_in_folder(rf"{temp}\B_data", des_B_data)
    copy_files_in_folder(rf"{temp}\B_format_data", des_B_FormatData)
    copy_file(rf"{temp}\{today}_limit_price.csv", des_limit_price)

    # TODoo yesterday 不是上一个交易日就没法运行了啊啊啊

    # 检验文件是否传输
    # FIXME 220240527 先注释，判断昨日有问题，下面也是

    # if count_files_with_target_field(des_A_data, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des_A_data}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return
    #
    # if count_files_with_target_field(des_A_FormatData, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des_A_FormatData}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return
    #
    # if count_files_with_target_field(des_B_data, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des_B_data}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return
    #
    # if count_files_with_target_field(des_B_FormatData, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des_B_FormatData}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return

    if not ifExist(des_limit_price):
        printRedMsg("limit price file is NOT copied")
    else:
        ...

    printBlueMsg("FL22SCB -> format_data")
    # 移动文件夹到 兴业扫单文件夹 C:\Users\Administrator\Desktop\兴业证券多账户交易
    des2_A_data = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCA\data"
    des2_A_FormatData = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCA\format_data"
    des2_B_data = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCB\data"
    des2_B_FormatData = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\FL22SCB\format_data"
    copy_files_in_folder(rf"{temp}\A_data", des2_A_data)
    copy_files_in_folder(rf"{temp}\A_format_data", des2_A_FormatData)
    copy_files_in_folder(rf"{temp}\B_data", des2_B_data)
    copy_files_in_folder(rf"{temp}\B_format_data", des2_B_FormatData)

    # 检验文件是否传输
    # if count_files_with_target_field(des2_A_data, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des2_A_data}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return
    #
    # if count_files_with_target_field(des2_A_FormatData, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des2_A_FormatData}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return
    #
    # if count_files_with_target_field(des2_B_data, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des2_B_data}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return
    #
    # if count_files_with_target_field(des2_B_FormatData, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des2_B_FormatData}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return

    printBlueMsg("HT02XY -> format_data\nHT02ZS -> format_data")
    # 移动 ht02 的 format_data 到兴业扫单文件夹 C:\Users\Administrator\Desktop\divide
    des_ht02xy_FormatData = r"C:\Users\Administrator\Desktop\divide_order_account\HT02XY\format_data"
    des_ht02zs_FormatData = r"C:\Users\Administrator\Desktop\divide_order_account\HT02ZS\format_data"
    copy_files_in_folder(rf"{temp}\HT02XY_format_data", des_ht02xy_FormatData)
    copy_files_in_folder(rf"{temp}\HT02ZS_format_data", des_ht02zs_FormatData)

    # 检验文件是否传输
    # if count_files_with_target_field(des_ht02xy_FormatData, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des_ht02xy_FormatData}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return
    #
    # if count_files_with_target_field(des_ht02zs_FormatData, g_yesterday) == 3:
    #     printGreenMsg(f"yesterday's data copied to {des_ht02zs_FormatData}")
    # else:
    #     printRedMsg("Data corrupt! returning to main menu...")
    #     input("")
    #     return

    # 操作完之后清空temp文件夹
    printYellowMsg("Deleting temp path and archiving zip file...")
    input("")
    erase_folder_contents(temp)
    move_files(zipPath, r"C:\Users\Administrator\Desktop\startTrade\data")

    printGreenMsg("process done, returning to main menu...")
    input("")


def findData():
    def list_files_with_numbers(path, n):
        files = os.listdir(path)
        files = [f for f in files if os.path.isfile(os.path.join(path, f))]  # 过滤出文件
        files.sort(key=lambda x: os.path.getmtime(os.path.join(path, x)), reverse=True)  # 按更新时间排序，最新的在最前

        for i, file in enumerate(files[:n]):
            print(f"{i + 1}: {file}")

        choice = int(input(f"请选择文件序号（输入数字 1 到 {n}）: "))
        if choice < 1 or choice > n:
            printRedMsg("无效的选择")
            return None

        selected_file = os.path.join(path, files[choice - 1])
        return selected_file

    def quickLookUpSellBuyList(product_choice):
        os.system("cls")
        product_names = ["CF15", "FL18", "HT02XY", "FL18SC", "FL18SCA", "FL22SCB", "FL", "PAHF"]
        productName = product_names[product_choice - 1]
        rootPath = r"C:\Users\Administrator\Desktop\兴业证券多账户交易"
        BSLpath = rf"{rootPath}\{productName}\Sell_Buy_List_{productName}"

        fullPath = list_files_with_numbers(BSLpath, 15)

        if product_choice == "8":
            fullPath = r"D:\Trade\scan_trade_xt\PAHF\Sell_Buy_List_PAHF"
        print(fullPath)

        stock_code = input("请输入你要查询的股票代码：")
        getInfoFromFileName(fullPath, product_fund_dict)
        find_stock_data(fullPath, stock_code)

        # TODO 还没写完

    # TODO 搞format data好像会炸
    while True:
        os.system("cls")
        print(f"\n\t\t  \033[1\033[42;3;31m FIND DATA MODE \033[0m\n")
        # print("请选择功能：")
        print("\n\t1. 查询单个文件")
        print("\t2. 快速查询 Sell_Buy_List")
        # print("quit. 退出")

        choice = input("\n\n\n\n\nEnter the function you want:  ")

        if choice == '1':
            os.system("cls")

            file_path = input("请拖入 CSV 文件：")
            if file_path.lower() == 'quit':
                print("返回主界面")
                return

            stock_code = input("请输入你要查询的股票代码：")
            getInfoFromFileName(file_path, product_fund_dict)
            find_stock_data(file_path, stock_code)
        elif choice == "2":
            # TODO finish!

            print("1. CF15")
            print("2. FL18")
            print("3. HT02XY")
            print("4. FL18SC")
            print("5. FL18SCA")
            print("6. FL22SCB")
            print("7. FL")
            print("8. PAHF")
            productChoice = input("input the product you want.")
            if productChoice == "1":
                quickLookUpSellBuyList(1)
            elif productChoice == "2":
                quickLookUpSellBuyList(2)
            elif productChoice == "3":
                quickLookUpSellBuyList(3)
            elif productChoice == "4":
                quickLookUpSellBuyList(4)
            elif productChoice == "5":
                quickLookUpSellBuyList(5)
            elif productChoice == "6":
                quickLookUpSellBuyList(6)
            elif productChoice == "7":
                quickLookUpSellBuyList(7)
            elif productChoice == "8":
                quickLookUpSellBuyList(8)





        elif choice == "test":
            printYellowMsg("testing for def getInfoFromFileName(file_name, product_fund_dict):")
            x = input()
            getInfoFromFileName(x, product_fund_dict)

            input("press any key to quit testing")
        elif choice.lower() == 'quit':
            print("返回主界面")
            return
        else:
            input("无效选项，请重新输入")


def simpleRiseTopTxt():
    os.system("cls")
    printYellowMsg("\n这个程序不会备份源文件，修改是不可逆的\n")

    file_path = input("请拖入要分析的文件：\n")
    if file_path.lower() == 'quit':
        print("返回主界面")
        return

    print("1. Risestop ----- 监控涨跌停")
    print("2. Unusual ------ 订单异常监控")
    print("3. Order -------- 程序文件")
    print("0. 自定义模式")
    # print("4. ")

    choice = input("\n输入要进行的操作的文件类型: \n")

    if choice == "1":
        printYellowMsg("Auto set the filename with date and file type, \nif not intend to do so, type no.\n")
        isSetName = input("")
        if not isSetName == "no":
            try:
                today = getDate('')
                print(f'today is {today}')
                directory = os.path.dirname(file_path)
                new_path = os.path.join(directory, today + "_risestop.log")
                os.rename(file_path, new_path)
                # printYellowMsg(rf"set the filename to {today}")
                file_path = new_path
            except Exception as e:
                # 如果发生异常，打印错误消息
                printRedMsg(f"重命名错误：{e}")
                input("")
            # print(file_path)
            # input("")
        if not file_path:
            printRedMsg("file_path is bad, return to main menu...")
            input("")
            return

        printGreenMsg("backup the log file...")
        originalLogPath = rf"{file_path}\..\originalLog"
        if not os.path.exists(originalLogPath):
            os.makedirs(originalLogPath)
            printGreenMsg(f"路径 {originalLogPath} 创建成功")
        else:
            print(f"路径 {originalLogPath} 已存在，跳过")
        # C:\Users\Administrator\Desktop\startTrade\Log\Risestop\originalLog
        # shutil.move(file_path, originalLogPath, copy_function=shutil.copy2)
        shutil.copy(file_path, originalLogPath)

        remove_lines_with_character(file_path, "Python")
        remove_lines_with_character(file_path, "for more information")
        remove_lines_with_character(file_path, "now_time")
        remove_lines_with_character(file_path, "nowtime")
        remove_lines_with_character(file_path, "没有一字涨停")
        remove_lines_with_character(file_path, "*******************")
        remove_lines_with_character(file_path, "还未到达")
        remove_lines_with_character(file_path, "已到达")
        remove_lines_with_character(file_path, "cb_error <class 'int'>")
        remove_lines_with_character(file_path, "                        ")

        input("return to main menu...")
        input("")
    elif choice == "2":
        printYellowMsg("Auto set the filename with date and file type, \nif not intend to do so, type no.\n")
        isSetName = input("")
        if not isSetName == "no":
            try:
                today = getDate('')
                print(f'today is {today}')
                directory = os.path.dirname(file_path)
                new_path = os.path.join(directory, today + "_unusual.log")
                os.rename(file_path, new_path)
                printYellowMsg(rf"set the filename to {today}")
                file_path = new_path
            except Exception as e:
                # 如果发生异常，打印错误消息
                printRedMsg(f"重命名错误：{e}")
                input("")
            # print(file_path)
            # input("")
        if not file_path:
            printRedMsg("file_path is bad, return to main menu...")
            input("")
            return

        printGreenMsg("backup the log file...")
        # originalLogPath = r"C:\Users\Administrator\Desktop\startTrade\Log\Unusual\originalLog"
        originalLogPath = rf"{file_path}\..\originalLog"
        if not os.path.exists(originalLogPath):
            os.makedirs(originalLogPath)
            printGreenMsg(f"路径 {originalLogPath} 创建成功")
        else:
            print(f"路径 {originalLogPath} 已存在，跳过")
        # shutil.move(file_path, originalLogPath, copy_function=shutil.copy2)
        shutil.copy(file_path, originalLogPath)

        remove_lines_with_character(file_path, "无异常订单情况")
        remove_lines_with_character(file_path, "现在在")
        remove_lines_with_character(file_path, "now time")
        remove_lines_with_character(file_path, "Python")
        remove_lines_with_character(file_path, "for more information")

        input("return to main menu...")
        input("")
    elif choice == "3":
        # order
        printYellowMsg("!在这个模式下，拖入文件的同文件夹下所有文件将会被处理!\n")
        printYellowMsg("Auto set the filename with date and file type, \nif not intend to do so, type no.\n")

        # 重命名
        isSetName = input("")
        if not isSetName == "no":
            try:
                printBlueMsg("rename now")
                # 重命名
                rename_py_files(file_path + r"\..", ".log")
                printBlueMsg("rename done")
            except Exception as e:
                # 如果发生异常，打印错误消息
                printRedMsg(f"重命名错误：{e}")
                input("")
            # print(file_path)
            # input("")
        # 如果 file_path是空的，就返回
        if not file_path:
            printRedMsg("file_path is bad, return to main menu...")
            input("")
            return

        # 对这个文件夹下的所有 py 文件进行操作
        files = os.listdir(file_path + r"\..")
        today = getDate('')
        print(f'today is {today}')
        for filename in files:
            if today in filename:
                # print(filename)
                orderPath = rf"C:\Users\Administrator\Desktop\startTrade\Log\Order\{filename}"
                printGreenMsg(f"backup the log file {filename}...")
                # originalLogPath = r"C:\Users\Administrator\Desktop\startTrade\Log\Unusual\originalLog"
                originalLogPath = rf"{orderPath}\..\originalLog"
                if not os.path.exists(originalLogPath):
                    os.makedirs(originalLogPath)
                    printGreenMsg(f"路径 {originalLogPath} 创建成功")
                else:
                    print(f"路径 {originalLogPath} 已存在，跳过")
                # print(originalLogPath)
                # print(orderPath)
                # shutil.move(file_path, originalLogPath, copy_function=shutil.copy2)
                try:
                    shutil.copy(orderPath, originalLogPath)
                except Exception as e:
                    # 如果发生异常，打印错误消息
                    printRedMsg(f"移动错误：{e}")
                    input("")

                remove_lines_with_character(orderPath, "交易信号还未到达")
                remove_lines_with_character(orderPath, "Not Trading Time")

        input("return to main menu...")
        input("")

    elif choice == "0":
        printYellowMsg("!在这个模式下，源文件将会被处理!请备份好文件\n")
        printYellowMsg("自定义模式\n")
        input("")

        if not file_path:
            printRedMsg("file_path is bad, return to main menu...")
            input("")
            return
        flag = False
        while not flag:
            x = input("输入要删除的行")
            if x == "quit":
                printYellowMsg("returning to main menu...")
                input("")
                return
            else:
                remove_lines_with_character(file_path, x)

        input("return to main menu...")
        input("")


def baiduToScan():
    today = getDate('')
    print(f'today is {today}')
    baiduSyncdiskPath = r"E:\BaiduSyncdisk"
    scanPath = rf"C:\Users\Administrator\Desktop\兴业证券多账户交易"

    fundListInBaidu = ['FL18', 'CF15', 'FL', 'PA', 'HT02', 'FL22SC']
    fundListInScan = ['FL18', 'CF15', 'FL', 'PA', 'HT02', 'FL22SC', 'HT02XY', 'FL22SCA', 'FL22SCB']
    fundListInDivide = ['HT02XY', 'FL22SCA', 'FL22SCB']
    fundListInBaiduAvaliable = ['FL18', 'CF15', 'FL', 'PA']

    # 先把所有信号复制到扫单文件夹
    for fund in fundListInBaiduAvaliable:
        origin = rf"{baiduSyncdiskPath}/Sell_Buy_List_{fund}"
        if fund == "PA":
            target = r"D:\Trade\scan_trade_xt\PAHF"
            # testTarget = r"C:\Users\Administrator\Desktop\startTrade\test_delete_later\pa"
        else:
            target = rf"{scanPath}/{fund}/Sell_Buy_List_{fund}"
            # testTarget = rf"C:\Users\Administrator\Desktop\startTrade\test_delete_later\{fund}/Sell_Buy_List_{fund}"
        copy_files_with_string_no_limited(origin, target, today)
        # 只有在处理 PA 时才会改名
        if fund == "PA":
            modifyNamePAtoPAHF(target)
            move_all_files_with_string(target, r"D:\Trade\scan_trade_xt\PAHF\Sell_Buy_List_PAHF", today)
            # move_all_files_with_string(testTarget, r"C:\Users\Administrator\Desktop\startTrade\test_delete_later\pa\q", "PAHF")

    input("press enter to exit")


def paToPahf():
    today = getDate('')
    print(f'today is {today}')
    baiduSyncdiskPath = r"E:\BaiduSyncdisk"

    fundPA = 'PA'

    # 先把所有信号复制到扫单文件夹
    origin = rf"{baiduSyncdiskPath}/Sell_Buy_List_{fundPA}"
    target = r"D:\Trade\scan_trade_xt\PAHF"

    copy_files_with_string_no_limited(origin, target, today)
    modifyNamePAtoPAHF(target)
    move_all_files_with_string(target, r"D:\Trade\scan_trade_xt\PAHF\Sell_Buy_List_PAHF", today)
    # move_all_files_with_string(testTarget, r"C:\Users\Administrator\Desktop\startTrade\test_delete_later\pa\q", "PAHF")

    input("press enter to exit")

def ht02ToHt02xy():
    today = getDate('')
    print(f'today is {today}')
    baiduSyncdiskPath = r"E:\BaiduSyncdisk"

    fundPA = 'HT02'

    # 先把所有信号复制到扫单文件夹
    origin = rf"{baiduSyncdiskPath}/Sell_Buy_List_{fundPA}"
    target = r"D:\Trade\scan_trade_xt\PAHF"

    copy_files_with_string_no_limited(origin, target, today)
    modifyNamePAtoPAHF(target)
    move_all_files_with_string(target, r"D:\Trade\scan_trade_xt\PAHF\Sell_Buy_List_PAHF", today)
    # move_all_files_with_string(testTarget, r"C:\Users\Administrator\Desktop\startTrade\test_delete_later\pa\q", "PAHF")

    input("press enter to exit")


def checkYesterdayDataTo37():
    today = getDate('')
    print(f'today is {today}')
    yesterday = getDate('', -1)
    print(rf"yesterday was {yesterday}")
    checkDataList = []
    checkFormatDataList = ['FL22SCA', 'FL22SCB']
    # checkFormatDataList = ['FL22SCA', 'FL22SCB', "HT02XY", 'HT02ZS']
    limitPricePath = r"C:\Users\Administrator\Desktop\兴业证券多账户交易\limit_price"

    for fundname in checkFormatDataList:
        path = rf"C:\Users\Administrator\Desktop\divide_order_account\{fundname}\format_data"
        count = count_files_with_target_field(path, yesterday)
        if count == 3 or count == 7:
            printGreenMsg(f"{fundname}'s format data is prepared.")
        else:
            printRedMsg(f"{fundname}'s format data is NOT prepared, the count is {count}")

    if count_files_with_target_field(limitPricePath, today) == 1:
        printGreenMsg(f"{today}'s limit price data is prepared.")
    else:
        printRedMsg(f"{today}'s limit price data is NOT prepared.")

    format_today = f"DataCorrect{today[:4]}-{today[4:6]}-{today[6:]}"
    if count_files_with_target_field(r"E:\BaiduSyncdisk\Sell_Buy_List_PA", format_today):
        printGreenMsg(f"{today}'s DataCorrect file is generated.")
    else:
        printRedMsg(f"{today}'s DataCorrect file is NOT generated!")

    input("press enter to return to main menu")


def diffGenerateAndCheck():
    today = getDate('')
    print(f'today is {today}')
    printYellowMsg(f"NOW RUNNING {diff_excel_path}")
    input("Press enter to continue...")

    thread = threading.Thread(target=run_python_file, args=(diff_excel_path,))
    thread.start()

    thread.join()
    printGreenMsg("diff excel done generating.")

    # 查看表格
    todayTotalDiffPath = rf"C:\Users\Administrator\Desktop\兴业证券多账户交易\totalDiff\持仓差异整合_{today}.xlsx"
    read_and_print_xlsx(todayTotalDiffPath)

    input("press enter to exit")


# def remove_lines_with_character(file_path, target_character):
#     printGreenMsg("正在处理中...")
#     # 打开文件并逐行读取内容
#     with open(file_path, 'r') as file:
#         lines = file.readlines()
#
#     # 过滤掉包含特定字符的行
#     filtered_lines = [line for line in lines if target_character not in line]
#
#     # 将剩余的行重新写入文件
#     with open(file_path, 'w') as file:
#         file.writelines(filtered_lines)
#
#     printGreenMsg("处理完成")
#     input("")

# def display_current_time():
#     # os.system("cls")
#     os.system('cls' if os.name == 'nt' else 'clear')
#     try:
#         # Define the specific time ranges and their corresponding messages
#         special_time_ranges = [
#             ("00:00", "00:01", "Midnight check!"),
#             ("08:40", "09:20", "Good morning! Time to start your day!"),
#             ("09:20", "09:30", "Real time Signal arrived"),
#             ("09:30", "09:50", "TWAP 1"),
#             ("09:50", "10:10", " - "),
#             ("10:10", "10:30", "TWAP 2 & Morning 1"),
#             ("10:30", "10:50", " - "),
#             ("10:50", "11:00", " - "),
#             ("11:00", "11:20", "TWAP 3 & Morning 2"),
#             ("11:20", "13:30", "Noon break"),
#             ("13:30", "13:40", "TWAP 4"),
#             ("13:40", "14:00", " - "),
#             ("14:00", "14:20", "TWAP 5 & Afternoon 1"),
#             ("14:20", "14:30", " - "),
#             ("14:30", "15:00", "TWAP 6 & Afternoon 2"),
#             ("15:00", "15:30", "Reverse repo"),
#             ("15:30", "17:40", "Wrap up the day!")
#         ]
#
#         while True:
#             now = datetime.now()
#             current_time = now.strftime("%H:%M:%S")
#             current_time_short = now.strftime("%H:%M")
#
#             # Clear the console
#             os.system('cls' if os.name == 'nt' else 'clear')
#
#             # Print the current time
#             print("Current Time:", current_time)
#
#             # Check if current time (HH:MM) is within any of the special time ranges
#             for start, end, message in special_time_ranges:
#                 if start <= current_time_short <= end:
#                     print(f"Special Time Range {start} - {end}: {message}")
#
#             # Wait for 1 second before updating
#             time.sleep(1)
#     except KeyboardInterrupt:
#         ...
#         # 捕捉到键盘中断 (Ctrl+C)，退出循环
#         # print("\n程序已退出")


def gridDataInsight():
    today = getDate('')
    print(f'today is {today}')
    yesterday = getDate('', -1)
    print(f'yesterday is {yesterday}')
    # today = "20240604"  # 调试用，上线后删除
    deleted_stock = ["601669.SH", "301336.SZ"]
    printYellowMsg(f"目前剔除的股票为： {deleted_stock}, 这些股票将暂时剔除")

    # 判断清仓要用的数据
    stock_min_max = None
    holding_latest_price = None
    stock_min_max_gmin_gmax = None # min就是最低价，gmin就是grid min：网格的最低格的价格

    done_data_analysis_today = 1
    today_or_yesterday = 1  # 1: today   2: yesterday


    # SETPARAS 网格数据分析的文件
    dataAnalysisPath = rf"C:\Users\Administrator\Desktop\网格数据分析\data_analysis\{today}网格交易数据分析.xlsx"

    # 判断建仓要用的数据

    def process_Paras():
        # grid_paras.txt
        gridTradePath = r"C:\Users\Administrator\Desktop\grid_trade"  # 修改路径
        # gridTradePath = r"D:\TRADE\grid_trade"  # 调试用，上线后删除
        todayGridParaPath = rf"{gridTradePath}/trade_before_order/{today}_grid_paras.txt"
        todayTradeBeforeOrderPath = rf"{gridTradePath}/trade_before_order/{today}_trade_before_order.csv"

        todayGridPara = grid.grid_paras_txt_to_dataframe(todayGridParaPath)
        prt.printDataFrameWithMaxRows(todayGridPara)

        # 可视化示例
        plot.plot_data_grid_paras_txt(todayGridPara)

    def process_StockInfo():
        # grid_stock_info.csv
        gridStockInfoPath = r"C:\Users\Administrator\Desktop\grid_trade\grid_info\grid_stock_info.csv"
        # gridStockInfoPath = r"D:\TRADE\grid_trade\grid_info\grid_stock_info.csv"  # 调试用，上线后删除
        stockInfo = data.get_df_from_csv(gridStockInfoPath)

        nonlocal stock_min_max, deleted_stock
        # print(f'printing stockInfo : \n{stockInfo}')
        stock_min_max = stockInfo[['stock_code', 'min_price', 'max_price', 'is_grid', 'grid_list']]
        stock_min_max = data.drop_stock_codes(stock_min_max, deleted_stock)  # 删除掉有持仓但是不交易的股票
        # print(f'printing stock_min_max : \n{stock_min_max}')



    def process_DataAnalysis():
        nonlocal today, dataAnalysisPath
        # 数据分析结果
        # dataAnalysisPath = r"D:\TRADE\网格数据分析\data_analysis\20240603网格交易数据分析.xlsx"  # 调试用，上线后删除
        try:
            dataAnalysis_AssetInfo = data.read_excel_to_df(dataAnalysisPath, 0)
            dataAnalysis_HoldingAfterClose = data.read_excel_to_df(dataAnalysisPath, 1)
            dataAnalysis_GridParas = data.read_excel_to_df(dataAnalysisPath, 4)
        except Exception as e:
            printRedMsg(f'read excel failed, maybe you havent done data analysis today?\n{e}')
            # input("press enter to return to main menu")
            return

        nonlocal holding_latest_price
        holding_latest_price = dataAnalysis_HoldingAfterClose[['股票代码', '最新价']]

    # 先判断今天做数据分析了没有
    if ifExist(dataAnalysisPath) == False:
        printRedMsg(f'File is not exist, maybe you havent done data analysis today??\n')
        x = input("\nIf you want to return to main menu,        press 1,\
                   \nif you want to check yesterday's data,     press 2.\
                   \nif you want to run grid data analysis now, press 3.\
                   \nif you want to manually enter date,      , press 4.\n\n")

        if x == "1":
            # done_data_analysis_today = 0
            return
        elif x == "2":
            # nonlocal today
            today_or_yesterday = 2
            today = getDate('', -1)
            print(f'yesterday is {yesterday}')
            # SETPARAS 更新为昨天的路径
            dataAnalysisPath = rf"C:\Users\Administrator\Desktop\网格数据分析\data_analysis\{today}网格交易数据分析.xlsx"
            print(f'reading file: {dataAnalysisPath}')
        elif x == '3':
            ...
        elif x == '4':
            # nonlocal today
            today_or_yesterday = 2
            today = input("enter the date\n")
            # SETPARAS 更新为昨天的路径
            dataAnalysisPath = rf"C:\Users\Administrator\Desktop\网格数据分析\data_analysis\{today}网格交易数据分析.xlsx"
            print(f'reading file: {dataAnalysisPath}')
        elif x == 'quit':
            return

    process_Paras()
    process_StockInfo()
    process_DataAnalysis()

    # 如果今天还没有做数据分析，那就选择返回或者是看昨天的数据分析
    if done_data_analysis_today == 0:
        return

    # 清仓逻辑 ==========================================================================================================
    # 提取股票代码前六位
    stock_min_max['stock_code_prefix'] = stock_min_max['stock_code'].astype(str).str.zfill(6).str[:6]
    holding_latest_price['股票代码_prefix'] = holding_latest_price['股票代码'].astype(str).str.zfill(6).str[:6]

    # 合并两个DataFrame
    stock_min_max_latest = pd.merge(stock_min_max, holding_latest_price, left_on='stock_code_prefix',
                                    right_on='股票代码_prefix')

    # 保留所需列
    stock_min_max_latest = stock_min_max_latest[['stock_code', 'min_price', 'max_price', '最新价']]
    data.rename_headers(stock_min_max_latest, ['stock_code', 'min_price', 'max_price', 'latest_price'])
    tradingStockList = data.getElementFromCol(stock_min_max_latest, "stock_code")

    # 输出结果
    # print(stock_min_max_latest)
    plot.stock_min_max_latest(stock_min_max_latest)

    # 建仓逻辑 ==========================================================================================================
    # print(f'printing stock_min_max : \n{stock_min_max}')
    # 这里是加入了网格的grid_list,提取 grid_list 列中的最小值和最大值，并将它们分别添加为 gmin 和 gmax 列
    def add_gmin_gmax(df):
        # 检查 grid_list 列的类型
        if 'grid_list' not in df.columns:
            raise ValueError("DataFrame must contain 'grid_list' column.")

        # 定义一个辅助函数来处理每个元素
        def parse_and_get_min(x):
            try:
                # 将字符串解析为列表
                parsed_list = ast.literal_eval(x)
                if isinstance(parsed_list, list) and parsed_list:  # 确保解析结果是列表且非空
                    return min(parsed_list)
            except (ValueError, SyntaxError):
                pass
            return None

        def parse_and_get_max(x):
            try:
                # 将字符串解析为列表
                parsed_list = ast.literal_eval(x)
                if isinstance(parsed_list, list) and parsed_list:  # 确保解析结果是列表且非空
                    return max(parsed_list)
            except (ValueError, SyntaxError):
                pass
            return None

        # 应用辅助函数
        df['gmin'] = df['grid_list'].apply(parse_and_get_min)
        df['gmax'] = df['grid_list'].apply(parse_and_get_max)

        return df

    stock_min_max_10percent_30percent = add_gmin_gmax(stock_min_max)
    # print(f'printing stock_min_max_10percent_30percent: \n{stock_min_max_10percent_30percent}')
    stock_min_max_10percent_30percent = stock_min_max_10percent_30percent.assign(
        ten_percent=lambda df: (df['gmax'] - df["gmin"]) * 0.1 + df["gmin"],
        thirty_percent=lambda df: (df['gmax'] - df["gmin"]) * 0.3 + df["gmin"]
    )
    # print(f"\nprinting 1st stock_min_max_10percent_30percent:\n{stock_min_max_10percent_30percent}")

    # 然后剔除掉正在交易的股票，剩下的就是再观察池中的股票拉！
    stock_min_max_10percent_30percent = data.drop_stock_codes(stock_min_max_10percent_30percent, tradingStockList)
    stock_min_max_10percent_30percent = data.drop_is_grid_is_0(stock_min_max_10percent_30percent)
    # print(stock_min_max_10percent_30percent)

    if today_or_yesterday == 1:
        stock_closePrice = getMonitorGridStockInfo()
    elif today_or_yesterday == 2:

        stock_closePrice = getMonitorGridStockInfo_yesterday(today)

    stock_closePrice = stock_closePrice.rename(columns={'ts_code': 'stock_code'})  # 更改股票的列名
    stock_closePrice = data.keepColumnsDeleteOthers(stock_closePrice, ['stock_code', 'close'])  # 只保留两列

    # print(f"\nprinting stock_closePrice:\n{stock_closePrice}")
    # print(f"\nprinting stock_min_max_10percent_30p ercent:\n{stock_min_max_10percent_30percent}")

    df_total = pd.merge(stock_closePrice, stock_min_max_10percent_30percent, on='stock_code', how='inner')

    # print(f"\nprinting df_total:\n{df_total}")
    plot.create_holding(df_total)

    # 建仓逻辑 第二阶段 ==================================================================================================
    # 再次挑选出在建仓区间的股票
    stock_in_the_range = []
    for stock in data.getElementFromCol(df_total, 'stock_code'):
        stock_row = df_total[df_total['stock_code'] == stock]
        close_value = stock_row['close'].values[0]
        ten_percent = stock_row['ten_percent'].values[0]
        thirty_percent = stock_row['thirty_percent'].values[0]

        if ten_percent < close_value < thirty_percent:
            stock_in_the_range.append(stock)

    print(stock_in_the_range)
    the_date_before_20_days = getDate('-', -20)
    df_stock_in_range = []
    df_part2_total = None
    # 获取这些股票的14日收盘价并绘图
    print(f"需要监控建仓的股票为: \n {stock_in_the_range}")
    for index, stock in enumerate(stock_in_the_range, start=1):
        df_single = ts.pro_bar(ts_code=stock, adj='qfq', start_date=the_date_before_20_days, end_date=today)
        df_part2_total = pd.concat([df_part2_total, df_single], ignore_index=True)  # 合并到这里
        print(f"FETCHING DATA --- \033[33m {index:2} / {len(stock_in_the_range)} \033[0m")
        # df = data.keepColumnsDeleteOthers(df, '')
    printGreenMsg("FETCHING COMPLETE!")

    # 获取这些股票的10%和30%的线
    part2_stock_10percent_30percent = stock_min_max_10percent_30percent \
        [stock_min_max_10percent_30percent['stock_code'].isin(stock_in_the_range)] \
        [['stock_code', 'ten_percent', 'thirty_percent']]

    plot.create_holding_2(df_part2_total, part2_stock_10percent_30percent)
    input("PRESS ENTER TO RETURN TO MAIN MENU...")


# 用来获取备选股票池的当天收盘价
def getMonitorGridStockInfo():
    today = getDate('')
    print(f'today is {today}')
    yesterday = getDate('', -1)
    print(f'yesterday is {yesterday}')

    # today = "20240604"  # 调试用，上线后删除
    def getMonitorStockList():
        # 先获取要查询的股票列表！！
        deleted_stock = ["601669.SH", "301336.SZ"]
        printYellowMsg(f"目前剔除的股票为： {deleted_stock}, 这些股票将暂时剔除")
        # info的文件路径 SETPARAS
        info = data.get_df_from_csv(r'C:\Users\Administrator\Desktop\grid_trade\grid_info\grid_stock_info.csv')
        info_isGrid_1 = data.drop_is_grid_is_0(info)
        info_tradingStock = data.drop_targetPositionIsNone(info_isGrid_1)
        return data.getElementFromCol(info_tradingStock, 'stock_code')

    df_total = None
    monitorList = getMonitorStockList()
    print(f"需要监控建仓的股票为: \n {monitorList}")
    for index, stock in enumerate(monitorList, start=1):
        df_single = ts.pro_bar(ts_code=stock, adj='qfq', start_date=today, end_date=today)
        df_total = pd.concat([df_total, df_single], ignore_index=True)  # 合并到这里
        print(f"FETCHING DATA --- \033[33m {index:2} / {len(monitorList)} \033[0m")
        # df = data.keepColumnsDeleteOthers(df, '')
    printGreenMsg("FETCHING COMPLETE!")
    if df_total.empty:
        printRedMsg(f"The DataFrame is empty\nPlease wait for the data update...")
        input("press ENTER to return")
    else:
        print(df_total)
        # 存储当天的info
        # try:
        #     # SETPARAS
        #     savePath = f"./stockData/{today}_Monitor_Grid_Info.csv"
        #     df_total.to_csv(savePath, index=False)
        #     printGreenMsg("File Saved.")
        # except Exception as e:
        #     printRedMsg(f"Failed to save info:{e}")

        time.sleep(1.5)

        # input("press enter to continue\n")
        return df_total


# 用来获取备选股票池的昨天收盘价
def getMonitorGridStockInfo_yesterday(today_or_yesterday=1):
    today = getDate('')
    print(f'today is {today}')
    yesterday = getDate('', -1)
    print(f'yesterday is {yesterday}')

    # 如果昨天不是昨天，就直接拿输入的日期来当成昨天
    if today_or_yesterday != 1:
        today = today_or_yesterday
        yesterday = today_or_yesterday

    # today = "20240604"  # 调试用，上线后删除
    def getMonitorStockList():
        # 先获取要查询的股票列表！！
        deleted_stock = ["601669.SH", "301336.SZ"]
        printYellowMsg(f"目前剔除的股票为： {deleted_stock}, 这些股票将暂时剔除")
        # info的文件路径 SETPARAS
        info = data.get_df_from_csv(r'C:\Users\Administrator\Desktop\grid_trade\grid_info\grid_stock_info.csv')
        info_isGrid_1 = data.drop_is_grid_is_0(info)
        info_tradingStock = data.drop_targetPositionIsNone(info_isGrid_1)
        return data.getElementFromCol(info_tradingStock, 'stock_code')

    df_total = None
    monitorList = getMonitorStockList()
    print(f"需要监控建仓的股票为: \n {monitorList}")
    for index, stock in enumerate(monitorList, start=1):
        df_single = ts.pro_bar(ts_code=stock, adj='qfq', start_date=yesterday, end_date=yesterday)
        df_total = pd.concat([df_total, df_single], ignore_index=True)  # 合并到这里
        print(f"FETCHING DATA --- \033[33m {index:2} / {len(monitorList)} \033[0m")
        # df = data.keepColumnsDeleteOthers(df, '')
    printGreenMsg("FETCHING COMPLETE!")
    if df_total.empty:
        printRedMsg(f"The DataFrame is empty\nPlease wait for the data update...")
        input("press ENTER to return")
    else:
        print(df_total)

    # 存储当天的info
    # try:
    #     # SETPARAS
    #     savePath = f"./stockData/{today}_Monitor_Grid_Info.csv"
    #     df_total.to_csv(savePath, index=False)
    #     printGreenMsg("File Saved.")
    # except Exception as e:
    #     printRedMsg(f"Failed to save info:{e}")

    time.sleep(1.5)

    # input("press enter to continue\n")
    return df_total


def getAllStockInfo():
    ...

def getGridStockPool():
    os.system("cls")
    infoPath = r"C:\Users\Administrator\Desktop\grid_trade\grid_info\grid_stock_info.csv"
    df = pd.read_csv(infoPath)

    df = data.drop_columns(df, ['grid_list','test_ret','judge_grid_end', 'judge_grid_start','test_start_date','test_end_date'])

    df_0 = df[df['is_grid'] == 0]
    df_1 = df[df['is_grid'] == 1]
    df_2 = df[df['is_grid'] == 2]
    df_1_None = df[(df['is_grid'] == 1) & (df['target_position_dict'] == "None")]
    df_1_NotNone = df[(df['is_grid'] == 1) & (df['target_position_dict'] != "None")]

    allDf = [df_0, df_1, df_2, df_1_None, df_1_NotNone]
    for df in allDf:
        df = data.drop_columns(df, ['target_position_dict', 'is_grid'])


    # 打印三个 DataFrame
    printGreenMsg("\n\t\t\t不交易股票的备份池")
    # printYellowMsg(df_0)
    printYellowMsg("此处省略")
    # print("\nDataFrame with is_grid == 1")
    # print(df_1)
    printGreenMsg("\n\t\t\t不交易股票池")
    printYellowMsg(df_2)
    printGreenMsg("\n\t\t\t备选观察股票池")
    printYellowMsg(df_1_None)
    printGreenMsg("\n\t\t\t正常进行网格交易的股票池")
    printYellowMsg(df_1_NotNone)

    input("")

def gridDataAnalysis():
    today = getDate('')
    print(f'today is {today}')
    printYellowMsg(f"NOW RUNNING {grid_data_analysis_path}")
    input("Press enter to continue...")

    thread = threading.Thread(target=run_python_file, args=(grid_data_analysis_path,))
    thread.start()

    thread.join()
    printGreenMsg("Grid data analysis done generating.")

    input("press enter to exit")

def gridDataModify():

    # DEBUG delete later
    # create_holding_info_path = rf'D:\lzy\temp\20240617交易股票建仓信息.csv'
    # clear_holding_info_path = rf'D:\lzy\temp\清仓信息.csv'

    while True:
        printYellowMsg("\n这是修改网格文件参数的功能，请慎重修改，修改后将会有完备的备份\n通过主菜单的 gsp 功能来查看 info 信息\n")
        print("1. 添加一只股票到备选股票池")
        print("2. 添加一条股票建仓信息")
        print("3. 删除一条股票建仓信息")
        print("4. 添加一条股票清仓信息")
        print("5. 删除一条股票清仓信息")
        x = input("\nEnter the operation you want:\n")

        if x == '1':
            print("添加一只股票到备选股票池") 
            printYellowMsg(f'>{grid_info_file_path}< is being modify.')
            backup_file(grid_info_file_path)
            data.add_row_to_grid_info_csv(grid_info_file_path)

            input("")

        elif x == '2':
            print("添加一条股票建仓信息")
            df = pandas.read_csv(create_holding_info_path)

            stock_code = input("请输入股票代码全称：")
            build_price = input("请输入建仓价格：")
            build_quantity = input("请输入建仓数量：")
            distribute_money = input("请输入分配资金：")
            print("现在在 grid_info 文件中查找这只股票的信息")

            # 获取 info 中的相关信息
            df_info = pandas.read_csv(grid_info_file_path)

            # 筛选出这只股票的信息
            df_info = df_info[df_info['stock_code'] == stock_code]

            # 把 stock_code,test_ret,grid_list,target_position_list,build_price,build_quantity,distribute_money,build_date 转换成变量
            test_ret = df_info['test_ret'].values[0]
            grid_list = df_info['grid_list'].values[0]
            target_position_list = df_info['target_position_dict'].values[0]



            # 添加这些列的信息 stock_code,test_ret,grid_list,target_position_list,build_price,build_quantity,distribute_money,build_date
            df = df.append({'stock_code': stock_code,
                            'test_ret': test_ret,
                            'grid_list': grid_list,
                            'target_position_list': target_position_list,
                            'build_price': build_price,
                            'build_quantity': build_quantity,
                            'distribute_money': distribute_money,
                            'build_date': getDate('')},
                           ignore_index=True)
            # prt.printDataFrameWithMaxRows(df)
            df.to_csv(create_holding_info_path, index=False)
            # TODO 添加检查的机制
            input("")

            input("")
        elif x == '3':
            print("删除一条股票建仓信息")

            input("")
        elif x == '4':
            print("添加一条股票清仓信息")

            input("")
        elif x == '5':
            print("删除一条股票清仓信息")

            input("")
        elif x == 'quit':
            break

    input("Press Enter to return to main menu.")


def grid_holding_calculate(stock_code="69"):
    def find_closest_smaller_with_index(arr, p):
        low, high = 0, len(arr) - 1
        closest_value = None
        closest_index = -1

        while low <= high:
            mid = (low + high) // 2

            if arr[mid] < p:
                closest_value = arr[mid]  # 更新最接近且小于 p 的值
                closest_index = mid  # 更新对应的索引
                low = mid + 1  # 继续向右侧搜索
            else:
                high = mid - 1  # 向左侧搜索

        return closest_value, closest_index

    # 倪琴写的，很牛掰
    def buy_stocks(prices, current_price, max_price):
        grid_number = len(prices) - 1
        # 创建持仓对应股数列表，初始均为0
        quantity_list = [0] * grid_number
        # n个档位对应n-1个格子
        grid_number = grid_number - 1

        # 确保current_price为浮点数
        current_price = float(current_price)
        # 计算现价下最大买入数量（100的整数倍）
        max_quantity = math.floor(max_price / current_price) // 100 * 100
        if max_quantity < grid_number * 100:
            print("格子数过多，不能满足最低均分")

        while True:
            # 计算每档持仓差异
            # 均分思想，先为每个格子均分，再将剩下的分配给中间的格子
            aver_quantity = max_quantity // (grid_number * 100)
            remain_quantity = max_quantity % (grid_number * 100)
            remain_quantity_count = remain_quantity // 100
            quantity_list[:-1] = [aver_quantity * 100] * (len(quantity_list) - 1)
            middle_indices = []
            if remain_quantity_count > 0:
                middle_indices = [grid_number // 2] if grid_number % 2 != 0 else [grid_number // 2 - 1, grid_number // 2]

            while len(middle_indices) < remain_quantity_count:
                middle_indices.append(middle_indices[0] - 1)
                middle_indices.sort()
                if len(middle_indices) < remain_quantity_count:
                    middle_indices.append(middle_indices[-1] + 1)
                    middle_indices.sort()

            for i in range(len(quantity_list)):
                if i in middle_indices:
                    quantity_list[i] += 100

            # 计算每档对应持仓数量
            quantity_sum = 0
            reduce_list = quantity_list.copy()
            for i in range(len(quantity_list) - 1):
                type = quantity_list[i]
                quantity_list[i] = max_quantity - quantity_sum
                quantity_sum += type
            # 获取最接近现价的那一档格子的价格及索引
            close_price, close_index = find_closest_smaller_with_index(prices, current_price)
            # 计算需要的花费 前n-1档和第n档分开计算
            all_count = 0
            for i in range(close_index):
                all_count += reduce_list[i] * prices[i]
            all_count += current_price * quantity_list[close_index]
            if all_count < max_price:
                break
            else:
                max_quantity -= 100

        printGreenMsg(f'需要花的金额: {all_count}\n可买的总股数：{quantity_sum}\n可用最大金额：{max_price}\n网格格子数为：{grid_number}')
        printGreenMsg(f"每档股票持仓差异分别为：{quantity_list}\n每格持仓差异：{reduce_list}")


        quantity_dict = {}

        for i in range(len(quantity_list)):
            quantity_dict[prices[i]] = quantity_list[i]

        return quantity_dict

    # 我写的，很傻杯
    def allocate_stocks(prices, max_price, current_price):
        # Calculate maximum shares per grid (must be multiple of 100)
        max_shares = (max_price // current_price) // 100 * 100

        # Initialize result dictionary with each price having 0 shares initially
        result = {price: 0 for price in prices}

        # Find the index of the grid where current_price falls between
        index = 0
        while index < len(prices) - 1 and prices[index + 1] < current_price:
            index += 1

        # Set the middle grid's shares to half of max_shares
        middle_index = len(prices) // 2
        result[prices[middle_index]] = max_shares // 2

        # Calculate shares for grids from middle towards lower prices
        for i in range(middle_index - 1, -1, -1):
            result[prices[i]] = min(result[prices[i + 1]] + 100, max_shares)

        # Calculate shares for grids from middle towards higher prices
        for i in range(middle_index + 1, len(prices)):
            result[prices[i]] = min(result[prices[i - 1]] - 100, max_shares)

        # Ensure minimum and maximum constraints
        result[min(prices, key=lambda x: abs(x - current_price))] = max_shares
        result[max_price] = 0

        return result

    # 如果用户没有传入stock_code, 那么就让用户输入
    if stock_code == "69":
        # 用户输入股票代码
        stock_code = input("enter the stock code:\n")
        # 为了调试，这里直接将股票代码设置为"601688"
        # stock_code = "601688"  # DEBUG!!!
    else:
        stock_code = stock_code[:6]


    # 定义CSV文件的路径，该文件包含网格信息
    # grid_info_file_path = r'D:\lzy\temp\grid_stock_info.csv'  # DEBUG!!!

    # 使用pandas的read_csv函数读取CSV文件，返回一个DataFrame对象
    grid_info = pd.read_csv(grid_info_file_path)

    # 确保DataFrame中的'stock_code'列中的数据是字符串格式
    grid_info['stock_code'] = grid_info['stock_code'].astype(str)

    # 截取'stock_code'列的前六位用于匹配，创建一个新的列'stock_code_prefix'
    grid_info['stock_code_prefix'] = grid_info['stock_code'].str[:6]

    # 在DataFrame中查找'stock_code_prefix'列等于用户输入的股票代码的行，
    # 并选择这些行的'grid_list'列，返回一个Series对象
    grid_prices = grid_info.loc[grid_info['stock_code_prefix'] == stock_code, 'grid_list']
    # prt.printDataFrameWithMaxRows(grid_info)

    # 检查是否找到了匹配的行
    if not grid_prices.empty:
    # 如果找到了匹配的行，选择第一行的'grid_list'值
        grid_prices = grid_prices.iloc[0]
        stock_code_long = grid_info.loc[grid_info['stock_code'].str.contains(stock_code), 'stock_code'].values[0]
        # print(stock_code_long)
    else:
    # 如果没有找到匹配的行，将'grid_prices'设置为空列表
        printRedMsg("No grid prices found for the stock code.")
        grid_prices = []
        stock_code_long = None


    # 将字符串形式的列表转换为实际的列表
    grid_prices_list = ast.literal_eval(grid_prices)

    printGreenMsg(f"\n格子获取成功: \n{grid_prices_list}")

    # 在线获取当前价格
    print("获取实时行情, 连接服务器中...", end="\t")

    try:
        # 东财数据
        df_dc = ts.realtime_quote(ts_code=stock_code_long, src='dc')
        # sina数据  未启用
        df_sina = ts.realtime_quote(ts_code='600000.SH,000001.SZ,000001.SH')
        printGreenMsg("连接成功")

        # prt.printDataFrameWithMaxRows(df_dc)
        now_price = df_dc['PRICE'][0]
        print(f"当前价格为: {now_price}")
    except Exception as e:
        printRedMsg(
            f"\n获取当前价格失败: {e}\n本接口是tushare org版实时接口的顺延，数据来自网络，且不进入tushare服务器，属于爬虫接口，请将tushare升级到1.3.3版本以上。")
        now_price = input("\n请输入当前价格: \n")

    # final = allocate_stocks(grid_prices_list, 40000, 13.13)
    # prt.print_dict(final)
    max_price_available = 40000
    print(f"请确认最大可用金额为：{max_price_available}\n")
    input("")

    final_ = buy_stocks(grid_prices_list, now_price, max_price_available)
    prt.aligned_dict_horizontal(final_)
    printGreenMsg(f'\n请复制目标持仓字典: \n\"{final_}\"')

    input("press enter to return to main menu")

def grid_holding_calculate_free(stock_code="69"):
    def find_closest_smaller_with_index(arr, p):
        low, high = 0, len(arr) - 1
        closest_value = None
        closest_index = -1

        while low <= high:
            mid = (low + high) // 2

            if arr[mid] < p:
                closest_value = arr[mid]  # 更新最接近且小于 p 的值
                closest_index = mid  # 更新对应的索引
                low = mid + 1  # 继续向右侧搜索
            else:
                high = mid - 1  # 向左侧搜索

        return closest_value, closest_index

    # 倪琴写的，很牛掰
    def buy_stocks(prices, current_price, max_price):
        grid_number = len(prices) - 1
        # 创建持仓对应股数列表，初始均为0
        quantity_list = [0] * grid_number
        # n个档位对应n-1个格子
        grid_number = grid_number - 1

        # 确保current_price为浮点数
        current_price = float(current_price)
        # 计算现价下最大买入数量（100的整数倍）
        max_quantity = math.floor(max_price / current_price) // 100 * 100
        if max_quantity < grid_number * 100:
            print("格子数过多，不能满足最低均分")

        while True:
            # 计算每档持仓差异
            # 均分思想，先为每个格子均分，再将剩下的分配给中间的格子
            aver_quantity = max_quantity // (grid_number * 100)
            remain_quantity = max_quantity % (grid_number * 100)
            remain_quantity_count = remain_quantity // 100
            quantity_list[:-1] = [aver_quantity * 100] * (len(quantity_list) - 1)
            middle_indices = []
            if remain_quantity_count > 0:
                middle_indices = [grid_number // 2] if grid_number % 2 != 0 else [grid_number // 2 - 1, grid_number // 2]

            while len(middle_indices) < remain_quantity_count:
                middle_indices.append(middle_indices[0] - 1)
                middle_indices.sort()
                if len(middle_indices) < remain_quantity_count:
                    middle_indices.append(middle_indices[-1] + 1)
                    middle_indices.sort()

            for i in range(len(quantity_list)):
                if i in middle_indices:
                    quantity_list[i] += 100

            # 计算每档对应持仓数量
            quantity_sum = 0
            reduce_list = quantity_list.copy()
            for i in range(len(quantity_list) - 1):
                type = quantity_list[i]
                quantity_list[i] = max_quantity - quantity_sum
                quantity_sum += type
            # 获取最接近现价的那一档格子的价格及索引
            close_price, close_index = find_closest_smaller_with_index(prices, current_price)
            # 计算需要的花费 前n-1档和第n档分开计算
            all_count = 0
            for i in range(close_index):
                all_count += reduce_list[i] * prices[i]
            all_count += current_price * quantity_list[close_index]
            if all_count < max_price:
                break
            else:
                max_quantity -= 100

        printGreenMsg(f'需要花的金额: {all_count}\n可买的总股数：{quantity_sum}\n可用最大金额：{max_price}\n网格格子数为：{grid_number}')
        printGreenMsg(f"每档股票持仓差异分别为：{quantity_list}\n每格持仓差异：{reduce_list}")


        quantity_dict = {}

        for i in range(len(quantity_list)):
            quantity_dict[prices[i]] = quantity_list[i]

        return quantity_dict

    # 我写的，很傻杯
    def allocate_stocks(prices, max_price, current_price):
        # Calculate maximum shares per grid (must be multiple of 100)
        max_shares = (max_price // current_price) // 100 * 100

        # Initialize result dictionary with each price having 0 shares initially
        result = {price: 0 for price in prices}

        # Find the index of the grid where current_price falls between
        index = 0
        while index < len(prices) - 1 and prices[index + 1] < current_price:
            index += 1

        # Set the middle grid's shares to half of max_shares
        middle_index = len(prices) // 2
        result[prices[middle_index]] = max_shares // 2

        # Calculate shares for grids from middle towards lower prices
        for i in range(middle_index - 1, -1, -1):
            result[prices[i]] = min(result[prices[i + 1]] + 100, max_shares)

        # Calculate shares for grids from middle towards higher prices
        for i in range(middle_index + 1, len(prices)):
            result[prices[i]] = min(result[prices[i - 1]] - 100, max_shares)

        # Ensure minimum and maximum constraints
        result[min(prices, key=lambda x: abs(x - current_price))] = max_shares
        result[max_price] = 0

        return result

    # 如果用户没有传入stock_code, 那么就让用户输入
    if stock_code == "69":
        # 用户输入股票代码
        stock_code = input("enter the stock code:\n")
        # 为了调试，这里直接将股票代码设置为"601688"
        # stock_code = "601688"  # DEBUG!!!
    else:
        stock_code = stock_code[:6]


    # 定义CSV文件的路径，该文件包含网格信息
    # grid_info_file_path = r'D:\lzy\temp\grid_stock_info.csv'  # DEBUG!!!

    # 使用pandas的read_csv函数读取CSV文件，返回一个DataFrame对象
    grid_info = pd.read_csv(grid_info_file_path)

    # 确保DataFrame中的'stock_code'列中的数据是字符串格式
    grid_info['stock_code'] = grid_info['stock_code'].astype(str)

    # 截取'stock_code'列的前六位用于匹配，创建一个新的列'stock_code_prefix'
    grid_info['stock_code_prefix'] = grid_info['stock_code'].str[:6]

    # 在DataFrame中查找'stock_code_prefix'列等于用户输入的股票代码的行，
    # 并选择这些行的'grid_list'列，返回一个Series对象
    grid_prices = grid_info.loc[grid_info['stock_code_prefix'] == stock_code, 'grid_list']
    # prt.printDataFrameWithMaxRows(grid_info)

    # 检查是否找到了匹配的行
    if not grid_prices.empty:
    # 如果找到了匹配的行，选择第一行的'grid_list'值
        grid_prices = grid_prices.iloc[0]
        stock_code_long = grid_info.loc[grid_info['stock_code'].str.contains(stock_code), 'stock_code'].values[0]
        # print(stock_code_long)
    else:
    # 如果没有找到匹配的行，将'grid_prices'设置为空列表
        printRedMsg("No grid prices found for the stock code.")
        grid_prices = []
        stock_code_long = None


    # 将字符串形式的列表转换为实际的列表
    grid_prices_list = ast.literal_eval(grid_prices)

    printGreenMsg(f"\n格子获取成功: \n{grid_prices_list}")


    now_price = input("\n请输入当前价格: \n")

    # final = allocate_stocks(grid_prices_list, 40000, 13.13)
    # prt.print_dict(final)
    max_price_available = 40000
    print(f"请确认最大可用金额为：{max_price_available}\n")
    input("")

    final_ = buy_stocks(grid_prices_list, now_price, max_price_available)
    prt.aligned_dict_horizontal(final_)
    printGreenMsg(f'\n请复制目标持仓字典: \n\"{final_}\"')

    # sleep 2s
    time.sleep(2)
    input("press enter to return to main menu")
def reverseRepo():
    print("逆回购下单程序开始")
    asset_path = rf"C:\Program Files\SmartTrader-Max\InsOrder\asset.csv" # SETPARAS
    task_path = rf"C:\Program Files\SmartTrader-Max\InsOrder\task.csv" # SETPARAS

    # asset_path = rf"D:\lzy\temp\asset.csv" # SETPARAS DEBUG
    # task_path = rf"D:\lzy\temp\task.csv" # SETPARAS DEBUG

    with open(asset_path, 'rb') as f:
        result = chardet.detect(f.read())
        print(f"the asset encode is: {result}")
    asset = pd.read_csv(asset_path, encoding=result['encoding'])
    #
    # with open(task_path, 'rb') as f:
    #     result = chardet.detect(f.read())
    #     print(f"the task encode is: {result}")
    #
    # task = pd.read_csv(task_path, encoding=result['encoding'])

    products_for_xy_scan = ['CF15', 'FL18', 'HT02XY', 'FL22SCA', 'FL22SCB', 'FL']

    today = getDate('')
    print(f'today is {today}')
    # 在 today 后面加上999990
    today_order_number = today + '999990'


    # prt.printDataFrameWithMaxRows(asset)
    # 筛选出 资产类别 列为F的行
    asset_filter = asset[asset['资产类别'] == 'F']
    # 筛选出 S2 为 人民币 的行
    asset_filter = asset_filter[asset_filter['S2'] == '人民币']
    asset_filter = data.drop_columns(asset_filter, ['资产类别', '账户类型', 'S1', 'S2', 'S3', 'S5', 'S6', 'S7', 'S8', 'S9'])
    # 去掉最后一列
    asset_filter = asset_filter.iloc[:, :-1]
    # 重命名列名  资金账户         S4        S10 为 资金账户    可用余额    当前时间
    asset_filter.columns = ['资金账户', '可用余额', '当前时间']

    # 添加一列 逆回购数量，是可用余额在千位向下取整再-300 再除以 100
    asset_filter['逆回购数量'] = asset_filter['可用余额'].apply(lambda x: (((x // 100) * 100 - 300) / 100) // 10 * 10 - 800)
    # 转成int类型
    asset_filter['逆回购数量'] = asset_filter['逆回购数量'].astype(int)

    HT02XY_entrust_num = 0
    print(asset_filter)
    # 新建一个 task 的 DataFrame， 结构为#指令编号,下单指令,账户类型,资金账户,证券代码,市场,委托数量,买卖方向,委托价格,委托类别,委托属性,委托编号,本地报单时间
    task = pd.DataFrame(columns=['#指令编号', '下单指令', '账户类型', '资金账户', '证券代码', '市场', '委托数量', '买卖方向', '委托价格', '委托类别', '委托属性', '委托编号', '本地报单时间'])
    for product in products_for_xy_scan:
        # 依次给每个产品添加一个订单号
        today_order_number = str(int(today_order_number) + 1)
        # print(f"{product}'s today_order_number: {today_order_number}")
        pro_num = 0
        if product in product_fund_dict:
            pro_num = product_fund_dict[product]

        # 确保'资金账户'列为整数类型
        asset_filter['资金账户'] = asset_filter['资金账户'].astype(str)

        # 在asset_filter中找到资金账户为pro_num的行，提取逆回购数量为entrust_num类型为int
        filter_row_with_pro_num = asset_filter[asset_filter['资金账户'] == pro_num]
        if not filter_row_with_pro_num.empty:
            entrust_num = int(filter_row_with_pro_num['逆回购数量'].values[0])
            if product == 'HT02XY':
                HT02XY_entrust_num = entrust_num
        else:
            entrust_num = 0

        # 依次给每个产品添加一个行，内容为空
        task = task.append({'#指令编号': today_order_number, '下单指令': 'T', '账户类型': '0', '资金账户': pro_num, '证券代码': '204001', '市场': '1', '委托数量': entrust_num, '买卖方向': '2', '委托价格': '0.1', '委托类别': '0', '委托属性': '0', '委托编号': '0', '本地报单时间': '15:29:59'}, ignore_index=True)

        # 如果 product 在 product_fund_dict 中，就把product_fund_dict对应账户号码填入task的资金账户

    # 打印出完整的下单详情，用户确认后再写入task文件
    printYellowMsg("下单详情如下：")
    prt.printDataFrameWithMaxRows(task)
    printYellowMsg("请确认下单信息，确认无误后输入insert继续")
    x = input("输入insert继续，输入其他任意字符退出\n")
    if x != 'insert':
        printRedMsg("退出下单程序")
        input("")
        return

    # 把task中除了表头的数据追加到task_path中, 用逗号分隔
    task.to_csv(task_path, mode='a', header=False, index=False, sep=',')
    # prt.printDataFrameWithMaxRows(asset_filter)

    # print(asset)
    # print(task)
    printYellowMsg(f" HT02XT 一般无法通过程序下单，请自行逆回购，数量为{HT02XY_entrust_num}")
    input("press enter to return to main menu")

def extremeValueCalculate():
    def calculate_values(sequence):
        # 找到序列中的最大值和最小值
        max_value = max(sequence)
        min_value = min(sequence)

        # 计算 (max - min) * 10% + min 和 (max - min) * 30% + min
        ten_percent_value = (max_value - min_value) * 0.1 + min_value
        thirty_percent_value = (max_value - min_value) * 0.3 + min_value

        # 打印结果
        print(f"{ten_percent_value},{thirty_percent_value}")

    while True:
        # 从控制台输入字符串
        input_string = input("请输入一个数字序列，以逗号分隔：")

        if input_string == "quit":
            print("程序结束")
            input("")
            return

        # 去掉字符串中的方括号
        input_string = input_string.strip('[]')

        # 将输入字符串转换为列表
        sequence = [float(x.strip()) for x in input_string.split(',')]

        # 调用函数计算并打印结果
        calculate_values(sequence)

    input("press enter to return to main menu")

def min_max_value_calculate():
    # stock_code = '002984.SZ'

    stock_codes = ['002138.SZ', '002156.SZ', '600104.SH',
                   '600226.SH', '600529.SH', '600686.SH',
                   '600741.SH', '603501.SH', '688025.SH']
    stock_codes = ['000563.SZ']
    start_date = '20231229'
    end_date = '20240705'

    # 确认信息
    print("===================================")
    print(f"= 股票代码: {stock_codes}")
    print(f"= 开始日期: {start_date}")
    print(f"= 结束日期: {end_date}")
    print("===================================\n")
    x = input("Confirm the information, Press Enter to continue...\n"
          "Enter yummy to enter by hand\n")
    if x == 'yummy':
        os.system('cls')
        stock_codes = input("Enter the stock codes separated by commas: ").split(',')
        start_date = input("Enter the start date(YYYYMMDD): ")
        end_date = input("Enter the end (YYYYMMDD): ")

        # 确认信息
        print("\n===================================")
        print(f"= 股票代码: {stock_codes}")
        print(f"= 开始日期: {start_date}")
        print(f"= 结束日期: {end_date}")
        print("===================================\n")
        input("Confirm the information, Press Enter to continue...\n")

    for stock_code in stock_codes:
        # 去 data 里面找数据文件来计算最小值和最大值
        # min_price, max_price = get_min_max_price_from_backtesting_daily_data_ADJ(stock_code, start_date, end_date)

        df = get_daily_data_from_wind(stock_code[:6], start_date,  end_date)
        min_price = df['CLOSE'].min()
        max_price = df['CLOSE'].max()
        # print(f"min price: {min_price}, max price: {max_price}")
        print(f"{min_price}, {max_price}")

    input("press enter to return to main menu")

def monitorFile():
    file_path = input("Drag the file to monitor: ")
    if not os.path.isfile(file_path):
        print(f"The file {file_path} does not exist.")
        return

    try:
        print(f"Monitoring changes in: {file_path}")
        subprocess.run(["powershell", "Get-Content", "-Path", file_path, "-Wait"], check=True)
    except subprocess.CalledProcessError as e:
        print(f"An error occurred: {e}")

def summaryBacktestResults():
    # 把回测完成的结果xlsx表格总结到一张表格里
    # log("启用功能7: 把回测完成的结果xlsx表格总结到一张表格里")
    try:
        folder_name = input("请输入文件夹名称：")
        # 获取所有CSV文件的文件名列表
        csv_files = glob.glob(rf'D:\TRADE\backtesting\result\{folder_name}\grid_n_test_info_pt0/*.xlsx')
        print(f"开始处理{csv_files}的数据")
        wb = Workbook()

        # 删除默认的sheet
        default_sheet = wb['Sheet']
        # wb.remove(default_sheet)
        for csv_file in tqdm(csv_files, ncols=100): # TODOed 改一下这个进度条
            # print(csv_file)
            df = pandas.read_excel(csv_file, sheet_name='Sheet1')
            sheet_name = csv_file.split("pt0")[-1][1:7]  # 获取文件名作为sheet名称
            # print(sheet_name)

            # 创建一个新的sheet
            ws = wb.create_sheet(title=sheet_name)

            # 将DataFrame写入新的sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            print(f"处理{csv_file}的数据结束")

        # 保存Excel文件
        wb.save(rf'D:\TRADE\backtesting\result\{folder_name}/total.xlsx')
        printGreenMsg(rf"处理结束，保存到 D:\TRADE\backtesting\result\{folder_name}/total.xlsx")
    except Exception as e:
        printRedMsg(f"处理失败: {e}")
        input("press enter to return to main menu")
        return
    input("Press Enter to continue...")

def downloadLimitpriceData():


    # bs.login()
    # from WindPy import w



    def get_code_list_from_csv():
        # 获取当前项目所在目录
        current_dir = os.path.dirname(__file__)

        # CSV文件路径
        csv_file_path = os.path.join(current_dir, 'stock_code_list.csv')

        # 读取CSV文件并转换为列表
        code_list = []
        with open(csv_file_path, 'r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                code_list.append(row[0][:6])  # 假设CSV文件只有一列数据
        # print(code_list)
        return code_list

    def get_stock_code_wind():
        """
        获取wind股票代码
        """
        from WindPy import w
        w.start()
        # 当日日期
        m_today = datetime.now().strftime("%Y-%m-%d")

        # 调用wind接口获取全部A股代码
        result = w.wset("sectorconstituent", f"date={m_today};sectorid=a001010100000000")
        stock_code_list = []
        # 判断代码运行是否正常
        if result.ErrorCode == 0:
            # print(result)
            data = result.Data
            # print(data)
            # print(data[1])
            # 判断数据是否为空
            if len(data) > 0:
                for item in data[1]:
                    # print(item)
                    if item.find("SZ") != -1 or item.find("SH") != -1:
                        stock_code_list.append(item)
            else:
                # print("len(result.Data) == 0")
                return []

        else:
            print("get_stock_code_wind result.ErrorCode != 0, exist error")
            return []

        # 若未成功获取数据，则返回空列表
        if len(stock_code_list) > 0:
            return stock_code_list
        else:
            return []

    def get_std_day(day_str: str) -> str:
        """
        day_str: "yyyymmdd"或者"yyyy/mm/dd"或者"yyyy-mm-dd"或者其它形式
        处理过程：用正则筛选字符串中的全部数字，然后拼凑在一起，如果拼凑结果长度等8，则返回拼凑结果，否则返回""
        """
        num_regex = re.compile(r"\d+")
        find_list = num_regex.findall(day_str)
        # print(find_list)
        ret = "".join(find_list)
        # print(ret)
        return ret

    def Generate_Date(tradeDaysPath, today):
        '''
        tradeDaysPath: 记录当年交易日日期的xlsx
        today: 订单日期，格式"yyyymmdd"
        '''
        # 交易日当天
        try:
            today_slope = today[:4] + "/" + today[4:6] + "/" + today[6:8]
            today_line = today[:4] + "-" + today[4:6] + "-" + today[6:8]

            # 前一个交易日
            df_tradeDate = pd.read_excel(tradeDaysPath)
            df_tradeDate["date"] = df_tradeDate.日期.apply(
                lambda x: get_std_day(x) if type(x) == str else x.strftime("%Y%m%d"))

            yesterday = df_tradeDate.date[df_tradeDate[df_tradeDate.date == today].index[0] - 1]
            yes_slope = yesterday[:4] + "/" + yesterday[4:6] + "/" + yesterday[6:8]
            yes_line = yesterday[:4] + "-" + yesterday[4:6] + "-" + yesterday[6:8]

            print(f"Working on yesterday{yesterday} today{today}")

            return today, today_slope, today_line, yesterday, yes_slope, yes_line
        except Exception as e:
            print(rf"Generate_Date Error: {e}\nMove 2024交易日 file to C:\Users\<USER>\Documents")
            return None

    def round_fixed(num, d):
        """
        浮点数的四舍五入
        num: 浮点数
        d: d>0，在小数点后面第几位四舍五入；
           d<0，在整数部位四舍五入，d=-1表示将个位数四舍五入，以此类推
        """
        if num == 0:
            a = 0
        else:
            a = round(num + 0.5 / 10 ** 7, d)  # +0.5/10**7是为了避免丢失的精度造成误差
        return a

    def disappear(*arg):
        # 不对重定向的消息做任何事情
        ...

    def get_day_k_data_wind_inFunc(code, start_date, end_date, price_adjust_type="F"):
        """
        # 获取wind日k线数据
        :param code: 股票代码，比如601688.SH
        :param start_date: 开始日期，比如2015-09-10
        :param end_date: 结束日期，比如2022-12-31
        :param price_adjust_type: 复权类型，"F"前复权，"B"后复权，"NA"不复权
        :return:日数据的dataframe，列名有PRE_CLOSE,OPEN,HIGH,LOW,CLOSE,VOLUME,ADJFACTOR
        """

        if code.startswith("0") or code.startswith("3"):
            code = code + ".SZ"
        elif code.startswith("6"):
            code = code + ".SH"

        with redirect_stdout(disappear):
            w.start()


        result = w.wsd(f"{code}", "pre_close,open,high,low,close,volume,adjfactor",
                       start_date, end_date, f"PriceAdj={price_adjust_type}")

        datatime = [x.strftime('%Y-%m-%d') for x in result.Times]
        data_columns = [x.upper() for x in result.Fields]
        df = pd.DataFrame(result.Data, index=data_columns, columns=datatime)
        data = df.T
        # print(data)
        data["PRE_CLOSE"] = data["PRE_CLOSE"].apply(lambda x: round_fixed(x, 2))
        data["OPEN"] = data["OPEN"].apply(lambda x: round_fixed(x, 2))
        data["HIGH"] = data["HIGH"].apply(lambda x: round_fixed(x, 2))
        data["LOW"] = data["LOW"].apply(lambda x: round_fixed(x, 2))
        data["CLOSE"] = data["CLOSE"].apply(lambda x: round_fixed(x, 2))
        try:
            data["VOLUME"] = data["VOLUME"].apply(lambda x: int(x))
        except ValueError:  # VOLUME的值可能为空，int函数会报错
            pass

        # 等待3秒
        # time.sleep(3)
        return data

    def get_day_k_data_bs(code, start_day, end_day):
        """
        code:"601998.SH"
        start_day: "20220521"
        end_day: "20220628"
        """

        if code.startswith("0") or code.startswith("3"):
            code = "SZ." + code
        elif code.startswith("6"):
            code = "SH." + code

        stock_code = code
        start_date = start_day[:4] + "-" + start_day[4:6] + "-" + start_day[6:8]  # "2022-05-21"
        end_date = end_day[:4] + "-" + end_day[4:6] + "-" + end_day[6:8]  # "2022-05-21"

        data_fields = "date,preclose,open,high,low,close"
        rs = bs.query_history_k_data_plus(stock_code,
                                          data_fields,
                                          start_date=start_date,
                                          end_date=end_date,
                                          frequency='d',
                                          adjustflag="2")

        # 一次性获取全部数据
        # data = rs.get_data() # 获取全部数据 # 注意获取全部数据后，rs就为空了
        # print(data)

        # 一行一行获取数据
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            # print(type(rs.get_row_data())) # list
            line_list = rs.get_row_data()
            temp_list = []
            for item in line_list:
                if item.find("-") == -1:
                    temp_list.append(round(float(item), 2))
                else:
                    temp_list.append(item)
            data_list.append(temp_list)
        data = pd.DataFrame(data_list, columns=["DATE", "PRE_CLOSE", "OPEN", "HIGH", "LOW", "CLOSE"])
        print(data)
        return data

    # 计算股票的涨跌停价
    def get_rise_down_stop_price(code, last_close_price):
        """
        code:股票代码，形如"601998"或者"601998.SH"
        last_close_price: 昨日收盘价
        """
        if code[:3] == "688" or code[:2] == "30":
            rise_price = round(last_close_price * 1.2, 2)
            down_price = round(last_close_price * 0.8, 2)
        else:
            rise_price = round(last_close_price * 1.1, 2)
            down_price = round(last_close_price * 0.9, 2)
        return rise_price, down_price

    def get_all_stcok_from_wind():
        from WindPy import w
        import pandas as pd
        # 初始化Wind接口
        w.start()

        # 获取全市场A股股票代码
        data = w.wset("SectorConstituent", "sectorId=a001010100000000")

        # 创建DataFrame存储股票代码
        df = pd.DataFrame(data.Data[1])
        df = df[0].str.slice(0, 6)

        # 将DataFrame保存为CSV文件，不包含字段名
        # df.to_csv("stock_code_list.csv", index=False, header=False)

        print(f"已成功获取{len(df)}支股票")

        # 关闭Wind接口
        w.stop()
        return df.to_list()

    def get_next_trade_day(today):
        # 读取 今年的交易日
        df_tradeDate = pd.read_excel(tradeDaysPath)
        # 转换为列表
        tradeDate_list = df_tradeDate.日期.apply(lambda x: get_std_day(x) if type(x) == str else x.strftime("%Y%m%d")).tolist()
        # print(tradeDate_list)
        # 返回今天之后的那个日期
        return tradeDate_list[tradeDate_list.index(today) + 1]




    # 交易日期
    year = str(datetime.now().year)
    tradeDaysPath = rf"{documents_path}/{year}交易日.xlsx"

    today = datetime.now().strftime('%Y%m%d')
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y%m%d')
    next_trade_day = get_next_trade_day(today)
    # 如果是周五，那么下一个交易日是下周一

    notToday = False

    print("This program will use the WIND database, please log in to the WIND account before running the program.")
    # print(fr"Move 2024交易日 file to C:\Users\<USER>\Documents")
    print(f"\n\t         Today -> {today}\n\t      Tomorrow -> {tomorrow}\n\tNext trade day -> {next_trade_day}\n")
    printYellowMsg(f"Auto download {next_trade_day} limit price data, please check if the date is correct")


    input_1 = input("\nEnter \"run\" to run the program, "
                    "\nEnter the date to run specific date,"
                    "otherwise exit\n\n")

    if is8Digits(input_1):
        # 如果输入的是8位数字,那么就要运行这天的涨跌停
        # 但是这个逻辑很奇怪，，？？？为什么是运行 today
        # print("是八位数")
        today = input_1
        next_trade_day = get_next_trade_day(today)
        notToday = True
        # print(today)
    elif input_1 == "run":
        print("Starting program")
        # 停留0.5s
        time.sleep(1)
    else:
        input("returning to main menu")
        return

    w.start()
    # date_select = input("是否自动生成当日日期？(y/n):")
    # if date_select == "y":
    #     today = datetime.now().strftime('%Y%m%d')
    # else:
    #     today = input("手动输入日期 yyyymmdd: ")

    today, today_slope, today_line, yesterday, yes_slope, yes_line = Generate_Date(tradeDaysPath, today)
    print("")
    # print(today, today_slope, today_line, yesterday, yes_slope, yes_line)

    # 1.获取所有买单股票代码
    # stock_code_buy_list = get_code_list_from_csv()

    stock_code_buy_list = get_all_stcok_from_wind()

    # stock_code_buy_list = []
    # print("股票个数：", len(stock_code_buy_list))

    # 2.获取前收盘价和计算涨停价
    # 这里的作用是为了保证不管这个文件里有没有东西，都要把交易程序跑起来的保证。特别是YH
    if len(stock_code_buy_list) == 0:
        stock_code_buy_list.append("601669.SH")
    # print("get start")
    start_time = time.time()

    top_limit_price_dict = dict()

    # 3.生成指定格式的文件A
    save_dir = fr"{documents_path}/limit_price"
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    limit_price_path = f"{save_dir}/{next_trade_day}_limit_price.csv"

    # 时间循环从这里开始：
    target_time = datetime.now().replace(hour=16, minute=30, second=0, microsecond=0)
    morning_start_time = datetime.now().replace(hour=9, minute=30, second=0, microsecond=0)
    if datetime.now() < morning_start_time:
        print("Good morning, forgot to run program yesterday?\nDont worry, I will run it for you\n")
        print("havent done this part, just restart and enter yesterday's date")


    if (datetime.now() < target_time) and not notToday:
        while datetime.now() < target_time:
            # printYellowMsg("It's not time yet")
            # wait till target time
            time.sleep(0.1)
            print("\r", end="")
            print(f"Now time: \033[33m{datetime.now()}\033[0m, Waiting for {target_time}...", end="")


    if (datetime.now() > target_time) and not notToday:
        # 检查数据源
        while True:
            # time.sleep(600) # 10分钟
            time.sleep(3)  # 3s DEBUG
            print(f"Now {datetime.now()}, checking data source:", end="")
            try:

                test_df = get_day_k_data_wind_inFunc("000001", today, today)['VOLUME']
                # 如果 VOLUME 列为 NaN，那么就是没有数据
                if test_df.isnull().values.any():
                    print(" \033[33mData is not updated yet\033[0m", end="")
                else:
                    print(" \033[32mData is updated\033[0m", end="")
                    break
                print("\r", end="")
            except Exception as e:
                print(f"\nFAILED checking data source: {e}")
                continue
    elif not notToday:
        print("It's not time yet")
        # wait till target time
        while datetime.now() < target_time:
            time.sleep(1)
            print(f"Waiting for {target_time}...")

    print("")


    # 从 limit_price_path 读取数据，获取第一列，即所有已经下载的股票，转成liswt
    if os.path.exists(limit_price_path):
        try:
            df = pd.read_csv(limit_price_path)
            limit_price_file_already_exist_stock_list = df["证券代码"].tolist()
            # 把每个元素后缀去掉，换句话说，只留前六位
            limit_price_file_already_exist_stock_list = [x[:6] for x in limit_price_file_already_exist_stock_list]
        except Exception as e:
            print(f"FAILED reading {limit_price_path}: {e}\ncheck file plz, delete file if needed")
            limit_price_file_already_exist_stock_list = []

        # print(f"已经下载的股票代码：{stock_code_buy_list}")
    else:
        # 如果文件不存在，那么就。。什么也不做
        limit_price_file_already_exist_stock_list = []
        print(f"{limit_price_path} file is not exist")


    ret_data = []
    print("\n")
    # 检查是否已经下载过了
    if os.path.exists(limit_price_path):
        if len(limit_price_file_already_exist_stock_list) == 0:
            print("No stock code in limit price file, Will download all stock code")
        elif len(limit_price_file_already_exist_stock_list) == len(stock_code_buy_list):
            print("All stock code in limit price file, No need to download")
            input("press enter to return to main menu")
            return
        elif len(limit_price_file_already_exist_stock_list) > 0:
            print("Some stock code in limit price file, Will download the rest of stock code")
            # 对比 limit_price_file_already_exist_stock_list 和 stock_code_buy_list，如果已经下载过了，就不再下载
            stock_code_buy_list = list(set(stock_code_buy_list) - set(limit_price_file_already_exist_stock_list))
            print(f"left {len(stock_code_buy_list)} left stock code to download")
            print(f"already downloaded {len(limit_price_file_already_exist_stock_list)} stock code")
    else:
        print("No limit price file, Will download all stock code")

    print(f'Downloading {today}\'s data')

    subprocess.Popen([
        "powershell",
        "-Command",
        f'Start-Process powershell -ArgumentList \'-NoExit -Command "Get-Content -Path \\"{limit_price_path}\\" -Wait"\''
    ])

    for index, stock_code in tqdm(enumerate(stock_code_buy_list),
                                  desc="DOWNLOADING",
                                  ncols=100,
                                  bar_format="{l_bar}{bar} | {n_fmt}/{total_fmt} Stocks [Elapsed: {elapsed} | ETA: {remaining}]",
                                  total=len(stock_code_buy_list)
                                  ):

        pre_close_data = get_day_k_data_wind_inFunc(stock_code, today, today)["CLOSE"]
        if pre_close_data.empty:
            print(index, ":", stock_code, "FAILED DOWNLOADING LIMIT PRICE DATA.")
            continue

        # data=get_day_k_data_bs(stock_code, yesterday, yesterday)["CLOSE"]
        # if len(data)>0:
        #     pre_close = data[0]
        # else:
        #     pre_close = 50
        pre_close = pre_close_data[0]
        top_limit_price, down_limit_price = get_rise_down_stop_price(stock_code, pre_close)

        if stock_code.startswith("0") or stock_code.startswith("3"):
            stock_code = stock_code + ".SZ"
        elif stock_code.startswith("6"):
            stock_code = stock_code + ".SH"
        elif stock_code.startswith("8"):
            continue
        # print(pre_close, top_limit_price, down_limit_price)

        # 每获取一条数据就追加写入 CSV 文件
        data_to_append = [[stock_code, pre_close, top_limit_price, down_limit_price]]
        df_data = pd.DataFrame(data=data_to_append, columns=["证券代码", "昨日收盘价", "今日涨停价", "今日跌停价"])
        df_data.to_csv(limit_price_path, mode='a', header=not os.path.exists(limit_price_path), index=False)

        ret_data.append([stock_code, pre_close, top_limit_price, down_limit_price])
        # if index % 1000 == 0:
        #     print(index, ":", stock_code + ":  涨停价  " + str(top_limit_price) + "   跌停价  " + str(down_limit_price))
    # 最终数据框的保存（可以根据需要保留或删除）
    final_df_data = pd.DataFrame(data=ret_data, columns=["证券代码", "昨日收盘价", "今日涨停价", "今日跌停价"])
    final_df_data.to_csv(limit_price_path, mode='a', header=not os.path.exists(limit_price_path), index=False)


    # df_data = pd.DataFrame(data=ret_data, columns=["证券代码", "昨日收盘价", "今日涨停价", "今日跌停价"])
    # df_data.to_csv(limit_price_path, index=False)

    print("End getting limit price")
    # print("获取前收盘价计算今涨停价和跌停价 结束")

    print(f"file saved to {limit_price_path}")
    end_time = time.time()
    print("It takes: ", int(end_time - start_time) / 60, "min")  # 用时 2093 秒=30分钟

    input("press enter to return to main menu")

# this requires no input to run
def downloadLimitpriceData_auto():


    # bs.login()
    # from WindPy import w



    def get_code_list_from_csv():
        # 获取当前项目所在目录
        current_dir = os.path.dirname(__file__)

        # CSV文件路径
        csv_file_path = os.path.join(current_dir, 'stock_code_list.csv')

        # 读取CSV文件并转换为列表
        code_list = []
        with open(csv_file_path, 'r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                code_list.append(row[0][:6])  # 假设CSV文件只有一列数据
        # print(code_list)
        return code_list

    def get_stock_code_wind():
        """
        获取wind股票代码
        """
        from WindPy import w
        w.start()
        # 当日日期
        m_today = datetime.now().strftime("%Y-%m-%d")

        # 调用wind接口获取全部A股代码
        result = w.wset("sectorconstituent", f"date={m_today};sectorid=a001010100000000")
        stock_code_list = []
        # 判断代码运行是否正常
        if result.ErrorCode == 0:
            # print(result)
            data = result.Data
            # print(data)
            # print(data[1])
            # 判断数据是否为空
            if len(data) > 0:
                for item in data[1]:
                    # print(item)
                    if item.find("SZ") != -1 or item.find("SH") != -1:
                        stock_code_list.append(item)
            else:
                # print("len(result.Data) == 0")
                return []

        else:
            print("get_stock_code_wind result.ErrorCode != 0, exist error")
            return []

        # 若未成功获取数据，则返回空列表
        if len(stock_code_list) > 0:
            return stock_code_list
        else:
            return []

    def get_std_day(day_str: str) -> str:
        """
        day_str: "yyyymmdd"或者"yyyy/mm/dd"或者"yyyy-mm-dd"或者其它形式
        处理过程：用正则筛选字符串中的全部数字，然后拼凑在一起，如果拼凑结果长度等8，则返回拼凑结果，否则返回""
        """
        num_regex = re.compile(r"\d+")
        find_list = num_regex.findall(day_str)
        # print(find_list)
        ret = "".join(find_list)
        # print(ret)
        return ret

    def Generate_Date(tradeDaysPath, today):
        '''
        tradeDaysPath: 记录当年交易日日期的xlsx
        today: 订单日期，格式"yyyymmdd"
        '''
        # 交易日当天
        try:
            today_slope = today[:4] + "/" + today[4:6] + "/" + today[6:8]
            today_line = today[:4] + "-" + today[4:6] + "-" + today[6:8]

            # 前一个交易日
            df_tradeDate = pd.read_excel(tradeDaysPath)
            df_tradeDate["date"] = df_tradeDate.日期.apply(
                lambda x: get_std_day(x) if type(x) == str else x.strftime("%Y%m%d"))

            yesterday = df_tradeDate.date[df_tradeDate[df_tradeDate.date == today].index[0] - 1]
            yes_slope = yesterday[:4] + "/" + yesterday[4:6] + "/" + yesterday[6:8]
            yes_line = yesterday[:4] + "-" + yesterday[4:6] + "-" + yesterday[6:8]

            print(f"Working on yesterday{yesterday} today{today}")

            return today, today_slope, today_line, yesterday, yes_slope, yes_line
        except Exception as e:
            print(rf"Generate_Date Error: {e}\nMove 2024交易日 file to C:\Users\<USER>\Documents")
            return None

    def round_fixed(num, d):
        """
        浮点数的四舍五入
        num: 浮点数
        d: d>0，在小数点后面第几位四舍五入；
           d<0，在整数部位四舍五入，d=-1表示将个位数四舍五入，以此类推
        """
        if num == 0:
            a = 0
        else:
            a = round(num + 0.5 / 10 ** 7, d)  # +0.5/10**7是为了避免丢失的精度造成误差
        return a

    def disappear(*arg):
        # 不对重定向的消息做任何事情
        ...

    def get_day_k_data_wind_inFunc(code, start_date, end_date, price_adjust_type="F"):
        """
        # 获取wind日k线数据
        :param code: 股票代码，比如601688.SH
        :param start_date: 开始日期，比如2015-09-10
        :param end_date: 结束日期，比如2022-12-31
        :param price_adjust_type: 复权类型，"F"前复权，"B"后复权，"NA"不复权
        :return:日数据的dataframe，列名有PRE_CLOSE,OPEN,HIGH,LOW,CLOSE,VOLUME,ADJFACTOR
        """

        if code.startswith("0") or code.startswith("3"):
            code = code + ".SZ"
        elif code.startswith("6"):
            code = code + ".SH"

        with redirect_stdout(disappear):
            w.start()


        result = w.wsd(f"{code}", "pre_close,open,high,low,close,volume,adjfactor",
                       start_date, end_date, f"PriceAdj={price_adjust_type}")

        datatime = [x.strftime('%Y-%m-%d') for x in result.Times]
        data_columns = [x.upper() for x in result.Fields]
        df = pd.DataFrame(result.Data, index=data_columns, columns=datatime)
        data = df.T
        # print(data)
        data["PRE_CLOSE"] = data["PRE_CLOSE"].apply(lambda x: round_fixed(x, 2))
        data["OPEN"] = data["OPEN"].apply(lambda x: round_fixed(x, 2))
        data["HIGH"] = data["HIGH"].apply(lambda x: round_fixed(x, 2))
        data["LOW"] = data["LOW"].apply(lambda x: round_fixed(x, 2))
        data["CLOSE"] = data["CLOSE"].apply(lambda x: round_fixed(x, 2))
        try:
            data["VOLUME"] = data["VOLUME"].apply(lambda x: int(x))
        except ValueError:  # VOLUME的值可能为空，int函数会报错
            pass

        # 等待3秒
        # time.sleep(3)
        return data

    def get_day_k_data_bs(code, start_day, end_day):
        """
        code:"601998.SH"
        start_day: "20220521"
        end_day: "20220628"
        """

        if code.startswith("0") or code.startswith("3"):
            code = "SZ." + code
        elif code.startswith("6"):
            code = "SH." + code

        stock_code = code
        start_date = start_day[:4] + "-" + start_day[4:6] + "-" + start_day[6:8]  # "2022-05-21"
        end_date = end_day[:4] + "-" + end_day[4:6] + "-" + end_day[6:8]  # "2022-05-21"

        data_fields = "date,preclose,open,high,low,close"
        rs = bs.query_history_k_data_plus(stock_code,
                                          data_fields,
                                          start_date=start_date,
                                          end_date=end_date,
                                          frequency='d',
                                          adjustflag="2")

        # 一次性获取全部数据
        # data = rs.get_data() # 获取全部数据 # 注意获取全部数据后，rs就为空了
        # print(data)

        # 一行一行获取数据
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            # print(type(rs.get_row_data())) # list
            line_list = rs.get_row_data()
            temp_list = []
            for item in line_list:
                if item.find("-") == -1:
                    temp_list.append(round(float(item), 2))
                else:
                    temp_list.append(item)
            data_list.append(temp_list)
        data = pd.DataFrame(data_list, columns=["DATE", "PRE_CLOSE", "OPEN", "HIGH", "LOW", "CLOSE"])
        print(data)
        return data

    # 计算股票的涨跌停价
    def get_rise_down_stop_price(code, last_close_price):
        """
        code:股票代码，形如"601998"或者"601998.SH"
        last_close_price: 昨日收盘价
        """
        if code[:3] == "688" or code[:2] == "30":
            rise_price = round(last_close_price * 1.2, 2)
            down_price = round(last_close_price * 0.8, 2)
        else:
            rise_price = round(last_close_price * 1.1, 2)
            down_price = round(last_close_price * 0.9, 2)
        return rise_price, down_price

    def get_all_stcok_from_wind():
        from WindPy import w
        import pandas as pd
        # 初始化Wind接口
        w.start()

        # 获取全市场A股股票代码
        data = w.wset("SectorConstituent", "sectorId=a001010100000000")

        # 创建DataFrame存储股票代码
        df = pd.DataFrame(data.Data[1])
        df = df[0].str.slice(0, 6)

        # 将DataFrame保存为CSV文件，不包含字段名
        # df.to_csv("stock_code_list.csv", index=False, header=False)

        print(f"已成功获取{len(df)}支股票")

        # 关闭Wind接口
        w.stop()
        return df.to_list()

    def get_next_trade_day(today):
        # 读取 今年的交易日
        df_tradeDate = pd.read_excel(tradeDaysPath)
        # 转换为列表
        tradeDate_list = df_tradeDate.日期.apply(lambda x: get_std_day(x) if type(x) == str else x.strftime("%Y%m%d")).tolist()
        # print(tradeDate_list)
        # 返回今天之后的那个日期
        return tradeDate_list[tradeDate_list.index(today) + 1]




    # 交易日期
    year = str(datetime.now().year)
    tradeDaysPath = rf"{documents_path}/{year}交易日.xlsx"

    today = datetime.now().strftime('%Y%m%d')
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y%m%d')
    next_trade_day = get_next_trade_day(today)
    # 如果是周五，那么下一个交易日是下周一

    notToday = False

    print("This program will use the WIND database, please log in to the WIND account before running the program.")
    # print(fr"Move 2024交易日 file to C:\Users\<USER>\Documents")
    print(f"\n\t         Today -> {today}\n\t      Tomorrow -> {tomorrow}\n\tNext trade day -> {next_trade_day}\n")
    printYellowMsg(f"Auto download {next_trade_day} limit price data, please check if the date is correct")


    print("Starting program")
    # 停留0.5s
    time.sleep(1)

    w.start()
    # date_select = input("是否自动生成当日日期？(y/n):")
    # if date_select == "y":
    #     today = datetime.now().strftime('%Y%m%d')
    # else:
    #     today = input("手动输入日期 yyyymmdd: ")

    today, today_slope, today_line, yesterday, yes_slope, yes_line = Generate_Date(tradeDaysPath, today)
    print("")
    # print(today, today_slope, today_line, yesterday, yes_slope, yes_line)

    # 1.获取所有买单股票代码
    # stock_code_buy_list = get_code_list_from_csv()

    stock_code_buy_list = get_all_stcok_from_wind()

    # stock_code_buy_list = []
    # print("股票个数：", len(stock_code_buy_list))

    # 2.获取前收盘价和计算涨停价
    # 这里的作用是为了保证不管这个文件里有没有东西，都要把交易程序跑起来的保证。特别是YH
    if len(stock_code_buy_list) == 0:
        stock_code_buy_list.append("601669.SH")
    # print("get start")
    start_time = time.time()

    top_limit_price_dict = dict()

    # 3.生成指定格式的文件A
    save_dir = fr"{documents_path}/limit_price"
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    limit_price_path = f"{save_dir}/{next_trade_day}_limit_price.csv"

    # 时间循环从这里开始：
    target_time = datetime.now().replace(hour=16, minute=30, second=0, microsecond=0)
    morning_start_time = datetime.now().replace(hour=9, minute=30, second=0, microsecond=0)
    if datetime.now() < morning_start_time:
        print("Good morning, forgot to run program yesterday?\nDont worry, I will run it for you\n")
        print("havent done this part, just restart and enter yesterday's date")


    if (datetime.now() < target_time) and not notToday:
        while datetime.now() < target_time:
            # printYellowMsg("It's not time yet")
            # wait till target time
            time.sleep(0.1)
            print("\r", end="")
            print(f"Now time: \033[33m{datetime.now()}\033[0m, Waiting for {target_time}...", end="")


    if (datetime.now() > target_time) and not notToday:
        # 检查数据源
        while True:
            # time.sleep(600) # 10分钟
            time.sleep(3)  # 3s DEBUG
            print(f"Now {datetime.now()}, checking data source:", end="")
            try:

                test_df = get_day_k_data_wind_inFunc("000001", today, today)['VOLUME']
                # 如果 VOLUME 列为 NaN，那么就是没有数据
                if test_df.isnull().values.any():
                    print(" \033[33mData is not updated yet\033[0m", end="")
                else:
                    print(" \033[32mData is updated\033[0m", end="")
                    break
                print("\r", end="")
            except Exception as e:
                print(f"\nFAILED checking data source: {e}")
                continue
    elif not notToday:
        print("It's not time yet")
        # wait till target time
        while datetime.now() < target_time:
            time.sleep(1)
            print(f"Waiting for {target_time}...")

    print("")


    # 从 limit_price_path 读取数据，获取第一列，即所有已经下载的股票，转成liswt
    if os.path.exists(limit_price_path):
        try:
            df = pd.read_csv(limit_price_path)
            limit_price_file_already_exist_stock_list = df["证券代码"].tolist()
            # 把每个元素后缀去掉，换句话说，只留前六位
            limit_price_file_already_exist_stock_list = [x[:6] for x in limit_price_file_already_exist_stock_list]
        except Exception as e:
            print(f"FAILED reading {limit_price_path}: {e}\ncheck file plz, delete file if needed")
            limit_price_file_already_exist_stock_list = []

        # print(f"已经下载的股票代码：{stock_code_buy_list}")
    else:
        # 如果文件不存在，那么就。。什么也不做
        limit_price_file_already_exist_stock_list = []
        print(f"{limit_price_path} file is not exist")


    ret_data = []
    print("\n")
    # 检查是否已经下载过了
    if os.path.exists(limit_price_path):
        if len(limit_price_file_already_exist_stock_list) == 0:
            print("No stock code in limit price file, Will download all stock code")
        elif len(limit_price_file_already_exist_stock_list) == len(stock_code_buy_list):
            print("All stock code in limit price file, No need to download")
            input("press enter to return to main menu")
            return
        elif len(limit_price_file_already_exist_stock_list) > 0:
            print("Some stock code in limit price file, Will download the rest of stock code")
            # 对比 limit_price_file_already_exist_stock_list 和 stock_code_buy_list，如果已经下载过了，就不再下载
            stock_code_buy_list = list(set(stock_code_buy_list) - set(limit_price_file_already_exist_stock_list))
            print(f"left {len(stock_code_buy_list)} left stock code to download")
            print(f"already downloaded {len(limit_price_file_already_exist_stock_list)} stock code")
    else:
        print("No limit price file, Will download all stock code")

    print(f'Downloading {today}\'s data')

    subprocess.Popen([
        "powershell",
        "-Command",
        f'Start-Process powershell -ArgumentList \'-NoExit -Command "Get-Content -Path \\"{limit_price_path}\\" -Wait"\''
    ])

    for index, stock_code in tqdm(enumerate(stock_code_buy_list),
                                  desc="DOWNLOADING",
                                  ncols=100,
                                  bar_format="{l_bar}{bar} | {n_fmt}/{total_fmt} Stocks [Elapsed: {elapsed} | ETA: {remaining}]",
                                  total=len(stock_code_buy_list)
                                  ):

        pre_close_data = get_day_k_data_wind_inFunc(stock_code, today, today)["CLOSE"]
        if pre_close_data.empty:
            print(index, ":", stock_code, "FAILED DOWNLOADING LIMIT PRICE DATA.")
            continue

        # data=get_day_k_data_bs(stock_code, yesterday, yesterday)["CLOSE"]
        # if len(data)>0:
        #     pre_close = data[0]
        # else:
        #     pre_close = 50
        pre_close = pre_close_data[0]
        top_limit_price, down_limit_price = get_rise_down_stop_price(stock_code, pre_close)

        if stock_code.startswith("0") or stock_code.startswith("3"):
            stock_code = stock_code + ".SZ"
        elif stock_code.startswith("6"):
            stock_code = stock_code + ".SH"
        elif stock_code.startswith("8"):
            continue
        # print(pre_close, top_limit_price, down_limit_price)

        # 每获取一条数据就追加写入 CSV 文件
        data_to_append = [[stock_code, pre_close, top_limit_price, down_limit_price]]
        df_data = pd.DataFrame(data=data_to_append, columns=["证券代码", "昨日收盘价", "今日涨停价", "今日跌停价"])
        df_data.to_csv(limit_price_path, mode='a', header=not os.path.exists(limit_price_path), index=False)

        ret_data.append([stock_code, pre_close, top_limit_price, down_limit_price])
        # if index % 1000 == 0:
        #     print(index, ":", stock_code + ":  涨停价  " + str(top_limit_price) + "   跌停价  " + str(down_limit_price))
    # 最终数据框的保存（可以根据需要保留或删除）
    final_df_data = pd.DataFrame(data=ret_data, columns=["证券代码", "昨日收盘价", "今日涨停价", "今日跌停价"])
    final_df_data.to_csv(limit_price_path, mode='a', header=not os.path.exists(limit_price_path), index=False)


    # df_data = pd.DataFrame(data=ret_data, columns=["证券代码", "昨日收盘价", "今日涨停价", "今日跌停价"])
    # df_data.to_csv(limit_price_path, index=False)

    print("End getting limit price")
    # print("获取前收盘价计算今涨停价和跌停价 结束")

    print(f"file saved to {limit_price_path}")
    end_time = time.time()
    print("It takes: ", int(end_time - start_time) / 60, "min")  # 用时 2093 秒=30分钟

    input("press enter to return to main menu")


def pyinstaller():
    printYellowMsg("该功能用时可能比较长，你可以让它在后台运行，该程序将立即返回主菜单")

    path = input("请拖入需要打包的 py 文件:\n")

    subprocess.Popen([
        "powershell",
        "-Command",
        f'Start-Process powershell -ArgumentList \'-NoExit -Command "pyinstaller \\"{path}\\" \''
    ])


def checkFileEncoding():
    asset_path = input("请拖入需要查看编码的文件: \n").strip().strip('"')

    if not os.path.isfile(asset_path):
        print("错误: 请输入有效的文件路径")
        input("按回车键返回主菜单")
        return

    try:
        with open(asset_path, 'rb') as f:
            raw_data = f.read()
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            confidence = result['confidence']
            print(f"文件编码检测结果: {encoding} (置信度: {confidence * 100:.2f}%)")
    except Exception as e:
        print(f"错误: 无法读取文件. 详细信息: {e}")

    input("\n按回车键返回主菜单")


def gridTradeTimes():
    # 读取指定日期范围的文件
    # 文件名格式：20240711-151022证券成交表.csv
    # 前八位是日期，再其中提取关键词”成交“
    # 把这些csv文件合并到一个文件里，再转成 df
    # 全局变量 trade_data_path
    # 筛选“证券代码”列中为目标股票代码的行
    trade_code = input("请输入股票代码：")
    print("\n请输入日期开始范围，或者输入对应数字快捷查询：")
    # 如果输入的是8位数字，那么就是日期，继续让用户输入结束日期
    print("1. 最近一个月")
    print("2. 最近三个月")
    print("3. 最近半年")
    print("4. 最近一年")
    print("5. 全部时间\n")

    # 获取用户选择的日期范围
    choice = input("请输入选择：")
    end_date = datetime.now()

    if choice == "1":
        start_date = end_date - timedelta(days=30)
    elif choice == "2":
        start_date = end_date - timedelta(days=90)
    elif choice == "3":
        start_date = end_date - timedelta(days=180)
    elif choice == "4":
        start_date = end_date - timedelta(days=365)
    elif choice == "5":
        start_date = datetime.min
    else:
        printRedMsg("无效的选择，请重新运行功能")
        return

    # 查找指定日期范围内的文件
    files = [f for f in os.listdir(trade_data_path) if f.endswith('.csv') and '成交' in f]

    # 筛选符合日期范围的文件
    selected_files = []
    date_pattern = re.compile(r'(\d{8})')

    for file in files:
        match = date_pattern.search(file)
        if match:
            file_date_str = match.group(1)
            try:
                file_date = datetime.strptime(file_date_str, "%Y%m%d")
                if start_date <= file_date <= end_date:
                    selected_files.append(os.path.join(trade_data_path, file))
            except ValueError:
                printRedMsg(f"文件 {file} 的日期格式无效。")

    if not selected_files:
        printRedMsg("没有找到符合条件的文件。")
        return

    printGreenMsg(f"找到{len(selected_files)}个符合条件的文件, 现在开始处理...")

    # 合并CSV文件并转换为DataFrame
    df_list = [pd.read_csv(file, encoding='gbk') for file in selected_files]
    combined_df = pd.concat(df_list, ignore_index=True)

    # 确保证券代码列为字符串类型，然后补充前导零
    combined_df['证券代码'] = combined_df['证券代码'].astype(str).str.zfill(6)

    # 筛选“证券代码”列中为目标股票代码的行
    filtered_df = combined_df[combined_df['证券代码'] == str(trade_code)]
    # 输出结果有多少行

    show_df = filtered_df[['证券代码', '证券名称', '成交数量', '成交金额', '买卖方向', '成交价格']]
    # 改名
    show_df.columns = ['stock_code', 'stock_name', 'transaction_volume', 'transaction_amount', 'direction', 'transaction_price']
    # show_df = filtered_df[['stock_code', 'stock_name', 'transaction_volume', 'transaction_amount', 'direction', 'transaction_price']]
    print(show_df)

    printGreenMsg(f"\n在 {str(start_date)[:10]} ~ {str(end_date)[:10]} 之间找到{len(filtered_df)}次交易数据")

    latest_file = None
    latest_date = ""

    # 读取 build_clear_info_path 中带“建仓”二字的文件
    # 文件名前8位是日期，筛选出最新的那个
    try:
        latest_file = None
        latest_date = ""

        for filename in os.listdir(build_clear_info_path):
            if "建仓" in filename:
                date_str = filename[:8]
                if len(date_str) == 8 and date_str.isdigit():
                    if date_str > latest_date:
                        latest_date = date_str
                        latest_file = os.path.join(build_clear_info_path, filename)

        build_info_df = pd.read_csv(latest_file, encoding='gbk')

        build_info_df_1 = build_info_df[build_info_df['stock_code'].apply(lambda x: str(x)[:6] == str(trade_code))]

        # 提取 build_date 并转换为字符串格式
        build_date = str(build_info_df_1['build_date'].values[0])

        # 将 build_date 转换为日期对象
        build_date = datetime.strptime(build_date, "%Y%m%d")

        printGreenMsg(f"{trade_code} 的建仓日期为: {build_date.strftime('%Y-%m-%d')}")

        # 计算 mean_days
        if start_date <= build_date <= end_date:
            mean_days = count_weekdays(build_date, end_date)
        else:
            mean_days = count_weekdays(start_date, end_date)

        # 计算平均每天交易多少次，每月交易多少次
        mean_times = len(filtered_df) / mean_days
        mean_times_month = mean_times * 21
        printGreenMsg(f"平均每天交易次数: {mean_times:.2f}")
        printGreenMsg(f"平均每月交易次数: {mean_times_month:.2f}")
    except Exception as e:
        printRedMsg(f"计算失败: {e}")
        input("按回车键返回主菜单")
        return



    input("press enter to return to main menu")

def countWeekdays():
    try:
        # 输入开始日期和结束日期
        start_date = input("请输入开始日期(格式: YYYYMMDD): ")
        end_date = input("请输入结束日期(格式: YYYYMMDD): ")
        # 转换为日期对象
        start_date = datetime.strptime(start_date, "%Y%m%d")
        end_date = datetime.strptime(end_date, "%Y%m%d")
        printGreenMsg(f'THERE ARE {count_weekdays(start_date, end_date)} DAYS')
    except Exception as e:
        printRedMsg(f"ERROR: {e}")
    input("\npress enter to return to main menu")

def copyData():
    data_analysis_path = rf"D:/hutao/project/数据分析"
    limit_price_path = rf"D:/limit_price_projects/limit_price_file"
    divide_order_account_path = rf'C:\Users\Administrator\Desktop\divide_order_account'
    original_signal_path = rf'E:\BaiduSyncdisk'

    multi_account_trade_path = rf"C:\Users\Administrator\Desktop\兴业证券多账户交易"
    pahf_data_path = rf"D:\Trade\scan_trade_xt\PAHF\data"

    divide_product = ["FL22SCA", 'FL22SCB', 'HT02XY', 'HT02ZS']
    products_in_syncdisk = ['FL22SC', 'CF15', 'FL', 'FL18', 'HT02', 'PA', 'YL17']
    products_in_multi_account_trade = ['FL22SCA','FL22SCB', 'CF15', 'FL', 'FL18', 'HT02XY','HT02ZS']


    # 确认时间
    while True:
        today = getDate('')
        print(f'today is {today}')

        # 读取 今年的交易日
        def get_std_day(day_str: str) -> str:
            """
            day_str: "yyyymmdd"或者"yyyy/mm/dd"或者"yyyy-mm-dd"或者其它形式
            处理过程：用正则筛选字符串中的全部数字，然后拼凑在一起，如果拼凑结果长度等8，则返回拼凑结果，否则返回""
            """
            num_regex = re.compile(r"\d+")
            find_list = num_regex.findall(day_str)
            # print(find_list)
            ret = "".join(find_list)
            # print(ret)
            return ret

        df_tradeDate = pd.read_excel(rf"{documents_path}/2024交易日.xlsx")
        # 转换为列表
        tradeDate_list = df_tradeDate.日期.apply(
            lambda x: get_std_day(x) if type(x) == str else x.strftime("%Y%m%d")).tolist()
        # 返回今天之后的那个日期
        next_trade_day = tradeDate_list[tradeDate_list.index(today) + 1]
        last_trade_day = tradeDate_list[tradeDate_list.index(today) - 1]
        print(f" CONFIRM DATE BY PRESS ENTER: {today}")
        print(' THE DATE THAT IS LAST DAY OF THE NEWEST LIMIT PRICE FILE')
        print(" IF NOT, ENTER DATE YOU WANT:")
        choice = input("\n")
        # 如果按下enter就继续，否则就是输入的日期
        if choice == "":
            date = today
            break
        elif is8Digits(choice):
            date = choice
            break
        else:
            printRedMsg("INVALID DATE")
            os.system("cls")

    # 先选择当前机器
    print("\n SELECT CURRENT MACHINE: ")
    print("666")
    print("18")
    print("37")
    print("15")
    machineCode = input("\n")
    os.system("cls")
    printGreenMsg(f" CURRENT MACHINE:\t{machineCode}    \n  \
                     LAST TRADE DAY: \t{last_trade_day} \n  \
                     CURRENT DATE:   \t{date}           \n  \
                     NEXT TRADE DAY: \t{next_trade_day} \n  \
                  ")

    zipOrUnzip = input("ZIP OR UNZIP, enter 'z' or 'u': ")

    # 选择机器，打包文件

    filesNeedToPack = []  # 需要打包的文件路径列表
    if machineCode == "666":
        if zipOrUnzip == 'z':
            limit_price_files = filter_files_by_keyword_inList(list_files(data_analysis_path), next_trade_day)
            print(f'limit price file:\n{limit_price_files}')

            filesNeedToPack.extend(limit_price_files)

            format_data_files = filter_files_by_keyword_inList(list_files(data_analysis_path), date)
            print(f'format data file:\n{format_data_files}')
            filesNeedToPack.extend(format_data_files)

            print(f' Files will be pack:\n{filesNeedToPack}')
            zip_files_in_list(extract_paths_from_tuple_list(filesNeedToPack), f"./data/{date}_data.zip")
        elif zipOrUnzip == 'u':
            zip_file = input("Drag the data zip file in:\n")
            unzip_file(zip_file, f"./temp/{date}_data")
    if machineCode == "37":
        if zipOrUnzip == 'u':
            # 拖入zip文件
            zip_file = input("Drag the data zip file in:\n")
            unzip_file(zip_file, f"./temp/{date}_data")
            data_file_path = f"./temp/{date}_data"
            list_files(data_file_path)
            for product in divide_product:
                copy_files_with_string_no_limited(extract_paths_from_tuple_list(data_file_path), f'{divide_order_account_path}/{product}/format_data', product)
            copy_files_with_string_no_limited(extract_paths_from_tuple_list(data_file_path), r'C:\Users\Administrator\Desktop\兴业证券多账户交易\limit_price', 'limit_price')
        elif zipOrUnzip == 'z':
            ori_files = filter_files_by_keyword_inList(list_files(divide_order_account_path), date)
            filesNeedToPack.extend(ori_files)
            multi_account_data_files = filter_files_by_keyword_inList(list_files(divide_order_account_path), date)
            print(f'multi_account_data_files:\n{multi_account_data_files}')

            input("")


    if machineCode == 'test':
        files = filter_files_by_keyword_inList(list_files(r'C:\Users\progene12\share\每日持仓差异'), '202407')
        print(files)
        zip_files_in_list(extract_paths_from_tuple_list(files), rf"C:\Users\progene12\Downloads/202407_data.zip")
        input("")

    if machineCode == "18":
        # 拖入zip文件
        zip_file = input("Drag the data zip file in:\n")

    # copy(choice)

    input("")

def runRO():
    def run_python_file_runRO_only(python_executable, path):
        command = f'"{python_executable}" "{path}"'
        print(f"Executing command: {command}")  # 调试信息
        try:
            subprocess.run(command, check=True, shell=True)
        except subprocess.CalledProcessError as e:
            print(f"Error occurred: {e}")

    python_executable = r"D:\softwares\python3.6.6\python.exe"
    roPath = r"D:\TRADE\backtesting\src\runOracle.py"
    thread = threading.Thread(target=run_python_file_runRO_only, args=(python_executable,roPath,))
    thread.start()
    # 等待线程结束
    thread.join()
    input("PRESS ENTER TO RETURN TO MAIN MENU")

def progress_bar(waitSec):
    """
    Function to display the progress bar in a separate process.
    """
    import time as t
    for _ in tqdm(range(100), ncols=100):
        t.sleep(waitSec / 100)

def monitorFuturesSignal():
    import os
    import pandas as pd
    from datetime import datetime, time
    import time as t

    """
    Monitor futures signals for multiple products and store the trading signals into dataframes.
    """

    def read_latest_position_file(directory):
        """
        Read the latest position file in the given directory and return the data in a DataFrame.
        读取最新的，文件名包含PositionFuture的文件
        内容为：
            "ticker","Numb.Short"
            "IC2411.CFE",1
            "IM2411.CFE",1
            ""
            "Margin",298710.72
            "ExposPerc","54.8%"
            ""
            "PreparedCash",814184
        提取其中的信息
        """

        def parsePosition(file_content):
            """
            @brief 解析文件内容并打印 ticker 和 Numb.Short 的值
            @param file_content 文件内容（由 f.read() 传入）
            """
            lines = file_content.strip().splitlines()

            # 读取第二行到空行之间的内容
            for line in lines[1:]:
                # 遇到空行或 "Margin" 停止
                if line.strip() == "\" \"" or "Margin" in line:
                    break
                # 使用 .split(',') 解析，确保结果长度正确
                parts = line.replace('"', '').split(',')
                if len(parts) == 2:  # 确保有两个元素
                    ticker, num_short = parts
                    print(f'{ticker} -> {num_short}')
                else:
                    ...
                    # print(f"Skipping line due to unexpected format: {line}")


        # Step 1: Identify the latest file that contains 'PositionFuture' in the filename

        latest_file = None
        latest_time = None

        for filename in os.listdir(directory):
            if 'PositionFuture' in filename:
                filepath = os.path.join(directory, filename)
                file_time = os.path.getmtime(filepath)
                if latest_time is None or file_time > latest_time:
                    latest_time = file_time
                    latest_file = filepath

        # If no file is found, return an empty DataFrame
        if latest_file is None:
            prt.redMsg(f"No position file found in {directory}")

        # print the whole file
        with open(latest_file, 'r') as f:
            content = f.read()
            # 如果文字中有No Future 那么就输出NO POSITION CURRENTLY
            if "No Future" in content:
                printBlueMsg(f"NO POSTION CURRENTLY")
            else:
                # printBoldMsg(f"POSITION FILE: {latest_file}")
                parsePosition(content)

            f.seek(0)
            lines = f.readlines()
            for line in lines[-1:]:
                # 把line解析，把,分割，提取后面的文本，转换为小数
                prepared_cash = float(line.split(",")[1])
                pc5 = prepared_cash * 0.05
                printCyanMsg(f'Prepare Cash: {prepared_cash:.2f}     {pc5:.2f}')
            return




    # 设置终端的大小
    os.system('mode con cols=80 lines=30')

    # 产品及其对应的目录
    products = {
        'YL17': r'E:\BaiduSyncdisk\Sell_Buy_List_YL17',
        'CF15': r'E:\BaiduSyncdisk\Sell_Buy_List_CF15',
        'HT02': r'E:\BaiduSyncdisk\Sell_Buy_List_HT02'
    }

    printGreenMsg("\n\n\t\t\t Booting up the futures signal monitor...\n\n")
    waitSec = 3


    # 初始化每个产品的信号DataFrame
    last_morning_dfs = {product: pd.DataFrame() for product in products}
    last_afternoon_dfs = {product: pd.DataFrame() for product in products}

    while True:
        today_str = datetime.now().strftime('%Y%m%d')
        current_time = datetime.now().time()
        # within_monitoring_period = time(10, 5) <= current_time <= time(10, 20) or time(14, 30) <= current_time <= time(15, 42)
        within_monitoring_period = 1
        if within_monitoring_period:
            try:
                # 在监控时间段内
                t.sleep(waitSec)  # 每隔waitSec秒监控一次
                os.system('cls')  # 清除控制台
                print("\n\tMonitoring futures signals...\n")

                # 遍历每个产品
                for product, directory in products.items():

                    printPurpleMsg(f'-> {product} <-')
                    read_latest_position_file(directory)
                    # 过滤出包含今天日期的文件
                    files = [f for f in os.listdir(directory) if today_str in f and 'Future' in f]

                    for file in files:
                        # 判断是否为交易信号
                        if 'Trading' in file and 'No' not in file:
                            # 有morning信号
                            if 'morning' in file:
                                last_morning_dfs[product] = pd.read_csv(os.path.join(directory, file))
                                printGreenMsg(f"{product} Morning trading signal detected: {file}")
                                printYellowMsg(last_morning_dfs[product])
                            # 有afternoon信号
                            elif 'afternoon' in file:
                                last_afternoon_dfs[product] = pd.read_csv(os.path.join(directory, file))
                                printGreenMsg(f"{product} Afternoon trading signal detected: {file}")
                                printYellowMsg(last_afternoon_dfs[product])
                            else:
                                printRedMsg(f"Invalid trading signal file: {file}")
                        elif 'No' in file and 'Trading' in file:
                            # 没有morning信号
                            if 'Morning' in file:
                                printBlueMsg(f"{product} No morning trading signal detected: {file}")
                            # 没有afternoon信号
                            elif 'Afternoon' in file:
                                printBlueMsg(f"{product} No afternoon trading signal detected: {file}")
                            else:
                                printRedMsg(f"Invalid trading signal file: {file}")

            except Exception as e:
                printRedMsg(f"Error occurred: {e}")
        else:
            # 不在监控时间段，显示最近一次的监控结果
            t.sleep(waitSec)  # 每隔waitSec秒显示一次结果
            os.system('cls')
            print("\n\tNot within monitoring period. Showing the last detected signals...\n")

            # 显示每个产品的最后一次信号
            for product in products:
                if not last_morning_dfs[product].empty:
                    printYellowMsg(f"Last {product} morning trading signal:")
                    printYellowMsg(last_morning_dfs[product])
                if not last_afternoon_dfs[product].empty:
                    printYellowMsg(f"Last {product} afternoon trading signal:")
                    printYellowMsg(last_afternoon_dfs[product])

        # # 检查用户输入是否退出
        # if input("").strip().lower() == 'quit':
        #     printGreenMsg("Exiting the monitor...")
        #     return

def futuresData():
    """
    这个功能帮助导出期货数据，

    遍历产品列表，提示用户当前应该进行到哪个产品。一旦用户将文件导出到目标文件夹，程序就监控到这个文件，
    然后根据文件的表头来判断是什么数据，然后将这个文件重命名为对应的文件名。
    当导出了两个文件后，就生成asset文件，让用户粘贴相应文字进入，保存文件，然后重置count，继续下一个产品。

    :return:
    """
    

    def monitor_and_process_products(monitorPath, futuresProducts):
        today = getDate('')
        known_files = set(os.listdir(monitorPath))  # 已知文件列表

        for product in futuresProducts:
            fileCount = 0  # 记录导出的文件数量
            printPurpleMsg(f"-> NOW EXPORT {product} <-")
            printYellowMsg(f'Already exported {fileCount} files')

            while fileCount < 2:
                current_files = set(os.listdir(monitorPath))
                new_files = current_files - known_files
                known_files = current_files

                if not new_files:
                    time.sleep(1)
                    continue

                for file_name in new_files:
                    file_path = os.path.join(monitorPath, file_name)
                    try:
                        # 假设文件是 CSV 格式
                        file.saveAsUft8(file_path)
                        df = pd.read_csv(file_path, encoding='utf-8')


                        # 如果文件表头包含 '成交合约'，则重命名为交易文件
                        if '成交合约' in df.columns:
                            transaction_file = os.path.join(monitorPath, product, f"transaction_{today}.csv")
                            df.to_csv(transaction_file, index=False)
                            printGreenMsg(f"Transaction data exported to {transaction_file}")
                            # 删除源文件
                            os.remove(file_path)
                            fileCount += 1
                        # 如果文件表头包含 '持仓合约'，则重命名为持仓文件
                        elif '持仓合约' in df.columns:
                            holding_file = os.path.join(monitorPath, product, f"holding_{today}.csv")
                            df.to_csv(holding_file, index=False)
                            printGreenMsg(f"Holding data exported to {holding_file}")
                            # 删除源文件
                            os.remove(file_path)
                            fileCount += 1

                        # 当导出两个文件后，跳出循环生成 asset 文件
                        if fileCount == 2:
                            printGreenMsg(f"Exported all {fileCount} files for {product}")
                            assetPath = os.path.join(monitorPath, product)
                            generate_asset_file(assetPath, today)
                            break
                        else:
                            printYellowMsg(f'Already exported {fileCount} files, left {2 - fileCount} files to export')
                    except Exception as e:
                        print(f"Error processing file '{file_name}': {e}")

                # 更新已知文件列表
                known_files = current_files

    def generate_asset_file(path, today):
        asset_file = os.path.join(path, f"asset_{today}.txt")
        printCyanMsg(f'Please copy asset information, end with "q"')

        assetInfo = []
        while True:
            line = input()
            if line.strip().lower() == "q":
                break
            assetInfo.append(line)

        with open(asset_file, 'w', encoding='utf-8') as f:
            f.write("\n".join(assetInfo))
        printGreenMsg(f"Asset file saved to {asset_file}")

    monitorPath = r'D:\TRADE\FuturesData'
    futuresProducts = ['HT02']  # 期货产品列表
    monitor_and_process_products(monitorPath, futuresProducts)

    print("=== End of futures data export ===")
    input("")

def monitorXYPosition():
    ###################### 1.账户信息 #######################################
    # 设置终端的大小
    os.system('mode con cols=80 lines=43')
    # 产品账号

    ###################### 2.獲取資金情況 #######################################
    current_data_path = ins_order_path
    desired_order = ['CF15XZ', 'CF15SC', 'FL18', 'HT02XY', 'FL22SCA', 'FL22SCB', 'FL']
    while True:
        try:
            os.system('cls')
            # print("")
            ReallySharePath = "C:/Program Files/SmartTrader-Max/InsOrder/asset.csv"
            # print("ReallySharePath", ReallySharePath)
            try:
                assetData_total = pd.read_csv(ReallySharePath, encoding="gbk")
            except:
                assetData_total = pd.read_csv(ReallySharePath, encoding="utf-8")

            fund_account_list = assetData_total['资金账户'].drop_duplicates().to_list()
            # print(fund_account_list)
            
            # 根据 desired_order 对 fund_account_list 进行排序
            # 使用 get_fundname 将号码映射为基金名称，然后根据 desired_order 排序
            fund_account_list = sorted(
                fund_account_list,
                key=lambda x: desired_order.index(get_fundname(x)) if get_fundname(x) in desired_order else len(desired_order)
            )
            
            for account in fund_account_list:
                print("\n")
                fundname = get_fundname(account)
                chineseFundname = get_chineseFundname(account)
                assetData = assetData_total[assetData_total['资金账户'] == account]
                assetData = assetData.reset_index()
                assetData = assetData.iloc[-1]
                # print(assetData)

                mv = assetData["S8"]  # 市值
                tv = assetData["S5"]  # 总资产
                am = assetData["S4"]  # 可用资金
                FL_sp = mv / tv
                print(f"-> {fundname}  \t [{account}] \t {chineseFundname}")
                print(f"持仓比例:{FL_sp * 100:.2f}%   市值:{mv:.2f}   总资产:{tv:.2f}   可用资金:{am:.2f} ")
                if FL_sp <= 0.85:
                    fill_money = tv * 0.85 - mv  # am - 0.15 * tv
                    printRedMsg(f"仓位资金缺口：{fill_money:.2f}")
                    # fill_money = (int(fill_money / 10000) + 10) * 10000
                    # print("买入ETF 510300 、 512100 、 510500")
                    # common.logger.info("买入ETF 510300 512100 和 510500")
                    # print(f"建议这两个ETF各自都买 {fill_money / 3}")
                    # common.logger.info(f"建议这两个ETF各自都买 {fill_money / 3}")
                    # input("press Enter to Continue")
                else:
                    printGreenMsg(f"{fundname}持仓大于85%")
        except Exception as e:
            printRedMsg(f"Error: {e}")

        time.sleep(2)



def dataCollector():
    def erase_folder_contents(folder_path):
        try:
            # 获取文件夹中的所有文件
            files = os.listdir(folder_path)

            # 删除每个文件
            for filename in files:
                file_path = os.path.join(folder_path, filename)
                os.remove(file_path)
                print(f"Deleted file: {file_path}")
        except Exception as e:
            print(f"Failed to erase folder contents. Error: {e}")

    def copy_latest_files(source_folder, destination_folder):
        try:
            # 获取源文件夹中按时间排序的文件列表
            files = sorted(os.listdir(source_folder), key=lambda x: os.path.getmtime(os.path.join(source_folder, x)),
                           reverse=True)

            # 仅复制最新的四个文件
            for file in files[:4]:
                source_file = os.path.join(source_folder, file)
                destination_file = os.path.join(destination_folder, file)

                # 复制文件
                shutil.copy2(source_file, destination_folder)
                print(f"Copied: {source_file} to {destination_file}")
        except Exception as e:
            print(f"Failed to copy files. Error: {e}")

        # try:
        #     # 清空目标文件夹
        #     for file in os.listdir(destination_folder):
        #         destination_file = os.path.join(destination_folder, file)
        #         os.remove(destination_file)
        #         print(f"Deleted: {destination_file}")
        # except Exception as e:
        #     print(f"Failed to delete files. Error: {e}")
    folder1 = r"C:\Users\progene12\share\dataCollector\GRID"
    folder2 = r"C:\Users\progene12\share\dataCollector\PAHF"
    folder3 = r"C:\Users\progene12\share\dataCollector\YL17"
    erase_folder_contents(folder1)
    erase_folder_contents(folder2)
    erase_folder_contents(folder3)

    # 示例用法
    source_folder_1 = r"D:\TRADE\grid_trade\trade_data"
    source_folder_2 = r"D:\TRADE\scan_trade_xt\PAHF\data"
    source_folder_3 = r"D:\TRADE\scan_trade_xt\YL17YH\data"

    destination_folder_1 = r"C:\Users\progene12\share\dataCollector\GRID"
    destination_folder_2 = r"C:\Users\progene12\share\dataCollector\PAHF"
    destination_folder_3 = r"C:\Users\progene12\share\dataCollector\YL17"

    copy_latest_files(source_folder_1, destination_folder_1)
    copy_latest_files(source_folder_2, destination_folder_2)
    copy_latest_files(source_folder_3, destination_folder_3)

    print("\n")
    # os.system(".\\C:\\Users\\progene12\\share\\httpServer.bat")

    # # 指定.bat文件的完整路径
    bat_file_path = "C:\\Users\\progene12\\share\\httpServer.bat"
    #
    # # 构建命令列表，使用Popen运行.bat文件
    # try:
    #     # 启动进程
    #     process = subprocess.Popen(bat_file_path, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    #     # 等待进程执行完成并获取输出结果
    #     stdout, stderr = process.communicate()
    #     # 打印输出结果
    #     if stdout:
    #         print("Output from the bat file:")
    #         print(stdout.decode())
    #         print("\nThe server is ONLINE\n")
    #     # 如果有错误信息，打印错误信息
    #     if stderr:
    #         print("Error from the bat file:")
    #         print(stderr.decode())
    # except Exception as e:
    #     print(f"An error occurred while trying to run the bat file: {e}")
    print("\nThe server is ONLINE\n")
    os.system(bat_file_path)

    input("Press Enter to exit...")


def orderETF():
    ...
#     print(order_format_st(480151137, 510300, ))


def convertFileEncoding():
    def convert_files_encoding(folder_path, target_encoding='utf-8'):
        """
        Convert all files in a specified folder to the target encoding.

        @param folder_path: Path to the folder containing the files.
        @param target_encoding: Target encoding to save the files (default is UTF-8).
        """
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)

            # 只处理文件，跳过文件夹
            if os.path.isfile(file_path):
                try:
                    # 读取文件内容
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()

                    # 以目标编码写入文件
                    with open(file_path, 'w', encoding=target_encoding) as f:
                        f.write(content)

                    printGreenMsg(f"Converted '{filename}' to {target_encoding} encoding.")
                except Exception as e:
                    printRedMsg(f"Error converting '{filename}': {e}")
    try:
        printYellowMsg(f"该功能将文件夹中的所有文件转换保存为指定的编码格式。")
        folder_path = input("Drag the folder here:\n")
        target_encoding = input("Enter the target encoding (default is UTF-8):\n") or 'utf-8'
        convert_files_encoding(folder_path, target_encoding)
    except Exception as e:
        printRedMsg(f"Error occurred: {e} MVLMLZ")
    input("")

def orderMonitor():
    # 用于检查

    # 委托状态
    # 0  未报
    # 1  待报
    # 2  已报
    # 3  已报待撤
    # 4  部成待撤
    # 5  部撤
    # 6  已撤
    # 7  部成
    # 8  已成
    # 9  废单

    # orders的表头：
    # #指令编号  委托时间 	委托编号  	账户类型 	资金账号   	证券代码 	证券名称 	证券市场 	委托数量 	委托价格 	委托类别 	买卖方向 	委托属性 	成交数量      	成交均价    	撤单数量 	委托状态 	错误信息  当前时间

    # task的表头 
    # #指令编号	下单指令	账户类型	资金账户	证券代码	市场	委托数量	买卖方向	委托价格	委托类别	委托属性	委托编号	本地报单时间
    os.system('mode con cols=80 lines=90')


    CF15 = ScanTrade('CF15')
    FL18 = ScanTrade('FL18')
    HT02XY = ScanTrade('HT02XY')
    FL22SCA = ScanTrade('FL22SCA')
    FL22SCB = ScanTrade('FL22SCB')
    FL = ScanTrade('FL')

    products = [CF15, FL18, HT02XY, FL22SCA, FL22SCB, FL]
    for product in products:
        print("-" * 40)
        print(product)
        # print(product.signal_list)

    


    # while True:
    #     try:
    #         os.system('cls')
    #         ordersPath = current_data_path + "/orders.csv"
    #         taskPath = current_data_path + "/Task.csv"
    #         # print("ReallySharePath", ReallySharePath)
    #         try:
    #             ordersDf = pd.read_csv(ordersPath, encoding="gbk")
    #             taskDf = pd.read_csv(taskPath, encoding="gbk")
    #         except:
    #             ordersDf = pd.read_csv(ordersPath, encoding="utf-8")
    #             taskDf = pd.read_csv(taskPath, encoding="utf-8")

    #         # print(fund_account_list)
    #         for account in fund_account_list:
    #             print("\n")
    #             fundname = get_fundname(account)
    #     except Exception as e:
    #         printRedMsg(f"Error: {e}")
# 
        # time.sleep(5)

    input("press Enter to exit")

def iiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii():
    ...


def afterMain():
    while True:
        os.system("cls")
        menu()
        choice = input("请输入选项数字：")

        if choice == "9":
            before0920processFL22SC()
        elif choice == "1":
            checkYesterdayDataTo37()
        elif choice == "10":
            realTimeSignalMoveForFL22SC()
            realTimeSignalMoveForHT02()
        elif choice == "13":
            dataCollectorOn40()
        elif choice == "3":
            checkExportData()
        elif choice == "8":
            findData()
        elif choice == "4":
            printAllAccountInfo()
        elif choice == "7":
            simpleRiseTopTxt()
            # display_current_time()
        elif choice == "99":
            ...
            # downloadDataFromServer40()
        elif choice == "12":
            copYesterdayData()
        elif choice == "9":
            paToPahf()
        elif choice == "11":
            baiduToScan()
        elif choice == "2":
            diffGenerateAndCheck()
        elif choice == "0":
            os.system("cls")
            print("See you tmr.\n")
            print("\n")
            printName()
            x = input("")
            break
        elif choice == "g":
            os.system("cls")
            gridDataInsight()
        elif choice == "ts":
            getMonitorGridStockInfo()
        elif choice == "ga":
            getAllStockInfo()
        elif choice == 'rs':
            ...
        elif choice == 'gsp':
            getGridStockPool()
        elif choice == 'gda':
            gridDataAnalysis()
        elif choice == 'gdm':
            os.system("cls")
            gridDataModify()
        elif choice == 'gc':
            os.system("cls")
            grid_holding_calculate()
        elif choice == 'gc -f': # free
            os.system("cls")
            grid_holding_calculate_free()
        elif choice == '5':
            os.system('cls')
            reverseRepo()
        elif choice == '6':
            os.system('cls')
            monitorFuturesSignal()
        elif choice == 'evc':
            os.system('cls')
            extremeValueCalculate()
        elif choice == 'mmc':
            os.system('cls')
            min_max_value_calculate()
        elif choice == 'mf':
            monitorFile()
        elif choice == 'sbr':
            os.system('cls')
            summaryBacktestResults()
        elif choice == 'dld':
            os.system('cls')
            downloadLimitpriceData()
        elif choice == 'pyi':
            os.system('cls')
            pyinstaller()
        elif choice == 'ce':
            os.system('cls')
            checkFileEncoding()
        elif choice == 'gtt':
            os.system('cls')
            gridTradeTimes()
        elif choice == 'cwd':
            countWeekdays()
        elif choice == 'cd':
            os.system('cls')
            copyData()
        elif choice == 'ro':
            runRO()
        elif choice == 'fd':
            os.system('cls')
            os.system("mode con cols=200 lines=120")
            # print("期货数据处理")
            futuresData()
        elif choice == 'mp':
            os.system('cls')
            monitorXYPosition()

        elif choice == 'dc':
            dataCollector()
        elif choice == 'oe':
            os.system('cls')
            orderETF()

        elif choice == 'cfe':
            os.system('cls')
            convertFileEncoding()

        elif choice == 'om':
            os.system('cls')
            orderMonitor()




        elif choice == "test":




            input("")
        else:
            printRedMsg("无效的选项，请重新输入！")
            input("\n")
            os.system("cls")


def main():
    # global root
    # root = tk.Tk()
    # root.title("Ultimate Trading Tool")
    #
    # # 创建菜单按钮
    # menu_btn = tk.Button(root, text="打开主菜单", command=afterMain)
    # menu_btn.pack(pady=20)
    #
    # output_text = tk.Text(root, height=10, width=50)
    # output_text.pack(padx=10, pady=10)
    #
    # root.mainloop()
    global g_yesterday
    temp = getDate('', -1)
    while 0:
        x = input("Wrong or Right? y/n \n")
        if x == "y" or "":
            g_yesterday = temp
            break
        elif x == "n":
            g_yesterday = input("enter yesterday's date (format: YYYYMMDD):")
            break
        else:
            print("enter the right choice")
    afterMain()


def menu():
    # TODOed 写一个可以便捷查询真实持仓的小程序
    # ipAddr = get_public_ip()

    print("--------------------------------------------------------------")
    print(f"\t\t\t  \033[1;42;31m MAIN MENU \033[0m")

    print("  1. 检查 昨日数据分析的数据 & 云盘是否正常工作")
    print("  2. Diff 持仓差异分析并查看")
    print("  3. 收盘拆分导出数据的拆分和检查 ")
    print("  4. 查看各个账号密码")
    print("  5. 逆回购下单")
    print("  6. 监控期货信号")

    print(" fd. 期货数据处理")
    print(" mp. 监控兴业扫单产品的持仓比例")
    print(" oe. 下单ETF")
    print(" om. 监控订单")
    # print("  7. 监控下单信号")

    # print("  8. 查询数据")
    # print("  9. PA -> PAHF -> Move")
    # print("  7. 存档并简化 LOG 记录")

    print("  0. 退出")
    print(" ")
    print("  g. 网格数据分析")
    print(" ts. 获取今日网格备选池股票收盘价")
    print(" ga. 获取全市场今日收盘价")

    print("gsp. 输出当前股票池")
    print("gda. 网格数据分析")
    print("gdm. 网格数据修改")
    print(" gc. 网格持股计算和格子计算")
    print("gc-f 网格持股计算和格子计算(自定义)")
    print("evc. 极值计算")
    print("mmc. 最大最小值计算")
    print(" mf. 监控文件")
    print("sbr. 汇总回测结果到一个表格")
    print("dld. 下载涨跌停数据")
    print("pyi. 一键生成pyinstaller可执行文件")
    print(" ce. 查看文件编码")
    print("gtt. 查看网格交易次数")
    print("cwd. 计算一段时间的工作日")
    print(" cd. 复制数据")
    print(" ro. 运行 RunOracle 程序")

    print(" dc. 数据收集(这是过时的功能)")
    print(" cfe. 转换文件夹中所有文件编码")
    # print("")
    # print(f"\033[33m已停用功能: \033[0m")
    # print(" rs. 重新调整窗口大小")
    # print("xxx. 时钟")
    # print("  9. 拆分 FL22SC, HT02 的原始信号并移回源路径 *已停用*")
    # print(" 10. 移动拆分后的 FL22SC, HT02 的实时信号 *已停用*")
    # print(" 11. 百度网盘同步空间 -> 扫单文件夹 *已停用*")
    # print(" 12. 昨日数据分析的数据 -> 分单文件夹 *已停用* ")
    # print(" 13. 自动整理数据分析的数据 *已停用*")

    # colorful version
    # printGreenMsg  ("  1. 检查 昨日数据分析的数据 & 云盘是否正常工作")
    # printGreenMsg  ("  2. PA -> PAHF -> Move")
    # printGreenMsg  ("  3. 收盘拆分导出数据的拆分和检查 ")
    # printGreenMsg  ("  4. Diff 持仓差异分析并查看")
    # printGreenMsg  ("  5. 存档并简化 LOG 记录")
    # printYellowMsg ("  6. 查询数据")
    # printGreenMsg  ("  7. 查看各个账号密码")
    #
    # printGreenMsg  ("  0. 退出")
    # print(" ")
    #
    # printGreenMsg  ("  g. 网格数据分析")
    # printYellowMsg (" ts. 获取今日网格备选池股票收盘价")
    # printYellowMsg (" ga. 获取全市场今日收盘价")
    # printGreenMsg  ("gsp. 输出当前股票池")
    # printYellowMsg ("gda. 网格数据分析")
    # print(" gc. 网格持股计算和格子计算")
    # print("gdm. 网格数据修改")
    #
    # print("")
    #
    # print("")
    # print(f"\033[33m已停用功能: \033[0m")
    # printRedMsg    (" rs. 重新调整窗口大小")
    #
    # printRedMsg    ("  8. 时钟")
    # printRedMsg("  9. 拆分 FL22SC, HT02 的原始信号并移回源路径 *已停用*")
    # printRedMsg(" 10. 移动拆分后的 FL22SC, HT02 的实时信号 *已停用*")
    # printRedMsg(" 11. 百度网盘同步空间 -> 扫单文件夹 *已停用*")
    # printRedMsg(" 12. 昨日数据分析的数据 -> 分单文件夹 *已停用* ")
    # printRedMsg(" 13. 自动整理数据分析的数据 *已停用*")



    # print("10. 从另一台机器上的 HTTP 服务器上 fetch 文件(已删除)         ")

    # printGreenMsg(f"\n这台机器的 IP 地址是： {ipAddr}")
    # printYellowMsg("\n适用于 61 的功能: 1, 2, 3, 4")
    # printYellowMsg("适用于 40 的功能: 5")
    # printYellowMsg("适用于本机的功能: 9")
    print("--------------------------------------------------------------")


if __name__ == "__main__":
    os.system("cls")
    main()


